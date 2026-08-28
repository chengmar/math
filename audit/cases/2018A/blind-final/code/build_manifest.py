"""Build the final secondary output manifest from an explicit allowlist.

Run this only after run_all.py, verify_outputs.py, and
check_reproducibility.py.  The manifest deliberately excludes itself.
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

import openpyxl
import yaml


WORKSPACE = Path(__file__).resolve().parents[1]
RESULTS = WORKSPACE / "results"
OUTPUT = RESULTS / "output_manifest.json"

ALLOWLIST = (
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "results/calibrated_parameters.yaml",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/claims.yaml",
    "results/<SOURCE_FILE_REDACTED>",
    "results/gap_structure_scenarios.yaml",
    "results/independent_verification.json",
    "results/<SOURCE_FILE_REDACTED>",
    "results/initial_condition_scenarios.yaml",
    "results/input_audit.json",
    "results/key_results.yaml",
    "results/<SOURCE_FILE_REDACTED>",
    "results/mesh_study.yaml",
    "results/model_comparison.yaml",
    "results/<SOURCE_FILE_REDACTED>",
    "results/objective_weight_scenarios.yaml",
    "results/optimization_q2.yaml",
    "results/optimization_q3.yaml",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/parameter_diagnostics.yaml",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/problem1_summary.yaml",
    "results/reproducibility_check.json",
    "results/robustness_summary.yaml",
    "results/<SOURCE_FILE_REDACTED>",
    "results/validation_summary.yaml",
    "results/verification.json",
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def update_normalized(digest: Any, value: Any) -> None:
    digest.update(
        json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            default=str,
        ).encode("utf-8")
    )
    digest.update(b"\n")


def workbook_semantic_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sheet_name in workbook.sheetnames:
        update_normalized(digest, {"sheet": sheet_name})
        for row in workbook[sheet_name].iter_rows(values_only=True):
            update_normalized(digest, list(row))
    workbook.close()
    return digest.hexdigest()


def main() -> None:
    allowed = {Path(relative).as_posix() for relative in ALLOWLIST}
    missing = sorted(relative for relative in allowed if not (WORKSPACE / relative).is_file())
    observed = {
        path.relative_to(WORKSPACE).as_posix()
        for directory in (RESULTS, WORKSPACE / "figures")
        for path in directory.iterdir()
        if path.is_file() and path.resolve() != OUTPUT.resolve()
    }
    unexpected = sorted(observed - allowed)
    if missing or unexpected:
        print(
            json.dumps(
                {"status": "fail", "missing": missing, "unexpected": unexpected},
                ensure_ascii=False,
            )
        )
        raise SystemExit(1)

    files = []
    for relative in sorted(allowed):
        path = WORKSPACE / relative
        files.append(
            {
                "path": relative,
                "sha256": sha256(path),
                "size_bytes": path.stat().st_size,
            }
        )
    report = {
        "schema_version": 2,
        "phase": "blind-revision",
        "status": "pass",
        "generation_order": "after run_all, verification, and semantic rerun",
        "allowlist_entry_count": len(allowed),
        "allowlist_status": "pass",
        "manifest_self_hash_included": False,
        "problem1_workbook_semantic_sha256": workbook_semantic_sha256(
            RESULTS / "<SOURCE_FILE_REDACTED>"
        ),
        "interpretation": "Current byte hashes plus a canonical workbook-cell digest; neither is a claim of physical or mathematical truth.",
        "files": files,
    }
    OUTPUT.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    solution_path = WORKSPACE / "solution-report.yaml"
    solution = yaml.safe_load(solution_path.read_text(encoding="utf-8"))
    solution["coverage"]["output_manifest"] = "pass"
    solution_path.write_text(
        yaml.safe_dump(solution, allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )
    reproducibility_path = WORKSPACE / "reproducibility.yaml"
    reproducibility = yaml.safe_load(
        reproducibility_path.read_text(encoding="utf-8")
    )
    reproducibility["output_manifest_status"] = "pass"
    reproducibility["output_manifest_entry_count"] = len(files)
    reproducibility["output_manifest_self_hash_included"] = False
    reproducibility_path.write_text(
        yaml.safe_dump(reproducibility, allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )
    trace_path = WORKSPACE / "revision-traceability.yaml"
    trace = yaml.safe_load(trace_path.read_text(encoding="utf-8"))
    for finding in trace["findings"]:
        if finding["id"] == "AUD-004":
            finding["closure_status"] = "pass"
            finding["reason"] = "Final explicit allowlist manifest excludes itself and records a canonical workbook digest."
    trace_path.write_text(
        yaml.safe_dump(trace, allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "status": "pass",
                "entry_count": len(files),
                "manifest_self_hash_included": False,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
