"""Perform a real second run and compare canonical numerical content."""

from __future__ import annotations

import hashlib
import json
import subprocess
import sys
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import yaml


WORKSPACE = Path(__file__).resolve().parents[1]
RESULTS = WORKSPACE / "results"


def update_hash(digest: Any, value: Any) -> None:
    digest.update(json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8"))
    digest.update(b"\n")


def semantic_digest() -> str:
    digest = hashlib.sha256()
    parsed_files = [
        RESULTS / "key_results.yaml",
        RESULTS / "calibrated_parameters.yaml",
        RESULTS / "optimization_q2.yaml",
        RESULTS / "optimization_q3.yaml",
        RESULTS / "validation_summary.yaml",
        RESULTS / "robustness_summary.yaml",
        RESULTS / "parameter_diagnostics.yaml",
        RESULTS / "mesh_study.yaml",
        RESULTS / "initial_condition_scenarios.yaml",
        RESULTS / "objective_weight_scenarios.yaml",
        RESULTS / "gap_structure_scenarios.yaml",
        RESULTS / "claims.yaml",
    ]
    for path in parsed_files:
        update_hash(digest, path.name)
        update_hash(digest, yaml.safe_load(path.read_text(encoding="utf-8")))
    byte_stable_files = [
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "<SOURCE_FILE_REDACTED>",
    ]
    for path in byte_stable_files:
        digest.update(path.name.encode("utf-8"))
        digest.update(path.read_bytes())

    # Wall-clock time is recorded for the mesh trade-off but is not expected
    # to be byte-identical.  All scientific columns remain in the digest.
    mesh = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>").drop(columns=["elapsed_s"])
    update_hash(digest, "<SOURCE_FILE_REDACTED>_without_elapsed_s")
    update_hash(digest, mesh.to_dict(orient="records"))

    workbook = openpyxl.load_workbook(RESULTS / "<SOURCE_FILE_REDACTED>", data_only=True, read_only=True)
    for sheet_name in workbook.sheetnames:
        update_hash(digest, sheet_name)
        for row in workbook[sheet_name].iter_rows(values_only=True):
            update_hash(digest, list(row))
    return digest.hexdigest()


def main() -> None:
    before = semantic_digest()
    run = subprocess.run(
        [sys.executable, str(WORKSPACE / "code" / "run_all.py")],
        cwd=WORKSPACE,
        check=False,
        text=True,
        capture_output=True,
    )
    after = semantic_digest() if run.returncode == 0 else None
    status = "pass" if run.returncode == 0 and before == after else "fail"

    verification = subprocess.run(
        [sys.executable, str(WORKSPACE / "code" / "verify_outputs.py")],
        cwd=WORKSPACE,
        check=False,
        text=True,
        capture_output=True,
    )
    verification_status = "pass" if verification.returncode == 0 else "fail"

    report = {
        "schema_version": 2,
        "phase": "blind-revision",
        "status": status,
        "first_semantic_sha256": before,
        "second_semantic_sha256": after,
        "run_all_return_code": run.returncode,
        "run_all_stdout": run.stdout.strip(),
        "run_all_stderr": run.stderr.strip(),
        "post_run_verification_status": verification_status,
        "verification_stdout": verification.stdout.strip(),
        "verification_stderr": verification.stderr.strip(),
        "interpretation": "Semantic equality of generated numbers and workbook cells; not a proof of mathematical or external validity.",
    }
    (RESULTS / "reproducibility_check.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    reproducibility_path = WORKSPACE / "reproducibility.yaml"
    reproducibility = yaml.safe_load(reproducibility_path.read_text(encoding="utf-8"))
    reproducibility["second_run_status"] = status
    reproducibility["semantic_sha256_first"] = before
    reproducibility["semantic_sha256_second"] = after
    reproducibility["post_run_verification_status"] = verification_status
    reproducibility_path.write_text(
        yaml.safe_dump(reproducibility, allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )

    solution_path = WORKSPACE / "solution-report.yaml"
    solution = yaml.safe_load(solution_path.read_text(encoding="utf-8"))
    solution["coverage"]["code_rerun"] = status
    solution["coverage"]["paper_result_consistency"] = verification_status
    solution["coverage"]["independent_verification"] = verification_status
    solution_path.write_text(
        yaml.safe_dump(solution, allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )

    trace_path = WORKSPACE / "revision-traceability.yaml"
    trace = yaml.safe_load(trace_path.read_text(encoding="utf-8"))
    for finding in trace["findings"]:
        if finding["id"] == "AUD-007" and verification_status == "pass":
            finding["closure_status"] = "pass"
            finding["reason"] = "Standalone Radau/full-grid verification and anchored paper claims passed."
    trace_path.write_text(
        yaml.safe_dump(trace, allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )
    print(json.dumps({"status": status, "post_run_verification_status": verification_status, "semantic_sha256": after}, ensure_ascii=False))
    if status != "pass" or verification_status != "pass":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
