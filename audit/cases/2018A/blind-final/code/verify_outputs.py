"""Verify generated artifacts, including an independent raw-input reference."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
import yaml


WORKSPACE = Path(__file__).resolve().parents[1]
RESULTS = WORKSPACE / "results"
PAPER = WORKSPACE / "paper"
PAPER_MD = PAPER / "paper.md"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def claim_region(text: str, anchor: str) -> str | None:
    pattern = re.compile(
        rf"(?:<!--|%)\s*{re.escape(anchor)}\s+BEGIN\s*(?:-->)?"
        rf"(?P<body>.*?)"
        rf"(?:<!--|%)\s*{re.escape(anchor)}\s+END\s*(?:-->)?",
        re.DOTALL,
    )
    match = pattern.search(text)
    return None if match is None else match.group("body")


def value_present(region: str, value: float, decimals: int) -> bool:
    return f"{float(value):.{decimals}f}" in region


def main() -> None:
    checks: list[dict[str, Any]] = []

    def add(check_id: str, status: str, evidence: Any) -> None:
        if status not in {"pass", "fail", "needs_review"}:
            raise ValueError(f"invalid status {status}")
        checks.append({"id": check_id, "status": status, "evidence": evidence})

    required = [
        WORKSPACE / "problem-analysis.md",
        WORKSPACE / "data-audit.md",
        WORKSPACE / "assumptions.yaml",
        WORKSPACE / "variables.yaml",
        WORKSPACE / "model-selection.md",
        WORKSPACE / "solution-report.yaml",
        WORKSPACE / "reproducibility.yaml",
        WORKSPACE / "revision-traceability.yaml",
        WORKSPACE / "provenance.yaml",
        WORKSPACE / "code" / "thermal_model.py",
        WORKSPACE / "code" / "run_all.py",
        WORKSPACE / "code" / "independent_reference.py",
        WORKSPACE / "code" / "build_manifest.py",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "key_results.yaml",
        RESULTS / "claims.yaml",
        PAPER_MD,
        PAPER / "main.tex",
    ]
    missing = [str(path.relative_to(WORKSPACE)) for path in required if not path.is_file()]
    add("required_deliverables", "pass" if not missing else "fail", {"missing": missing})

    input_audit = json.loads((RESULTS / "input_audit.json").read_text(encoding="utf-8"))
    actual_problem_hash = sha256(
        WORKSPACE / "input" / "problem" / "<SOURCE_FILE_REDACTED>"
    )
    actual_data_hash = sha256(
        WORKSPACE / "input" / "data" / "<SOURCE_FILE_REDACTED>"
    )
    input_ok = (
        actual_problem_hash == input_audit["problem_sha256"]
        and actual_data_hash == input_audit["data_sha256"]
    )
    add(
        "input_identity",
        "pass" if input_ok else "fail",
        {
            "problem_sha256": actual_problem_hash,
            "data_sha256": actual_data_hash,
            "interpretation": "identity only, not mathematical correctness",
        },
    )
    material_pipeline = input_audit.get("material_pipeline", {})
    material_pipeline_ok = (
        material_pipeline.get("status") == "pass"
        and material_pipeline.get("solver_source") == "attachment_1_typed_objects"
        and material_pipeline.get("hard_coded_solver_materials_used") is False
        and material_pipeline.get("layer_order") == ["I", "II", "III", "IV"]
    )
    add(
        "attachment_materials_are_solver_input",
        "pass" if material_pipeline_ok else "fail",
        material_pipeline,
    )

    workbook = openpyxl.load_workbook(
        RESULTS / "<SOURCE_FILE_REDACTED>", data_only=True, read_only=True
    )
    expected_sheets = {
        "temperature_distribution",
        "spatial_nodes",
        "skin_comparison",
        "parameters",
    }
    sheet_ok = set(workbook.sheetnames) == expected_sheets
    distribution = workbook["temperature_distribution"]
    iterator = distribution.iter_rows(values_only=True)
    header = list(next(iterator))
    first = list(next(iterator))
    last = first
    row_count = 2
    for row in iterator:
        last = list(row)
        row_count += 1
    workbook.close()
    skin = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>")
    workbook_ok = (
        sheet_ok
        and row_count == 5402
        and len(header) == 60
        and float(first[0]) == 0.0
        and float(last[0]) == 5400.0
        and float(np.max(np.abs(np.asarray(first[1:], dtype=float) - 37.0))) <= 1e-12
        and abs(float(last[-1]) - float(skin.iloc[-1]["predicted_c"])) <= 1e-9
    )
    add(
        "problem1_workbook_semantics",
        "pass" if workbook_ok else "fail",
        {
            "sheets": sorted(expected_sheets),
            "rows": row_count,
            "columns": len(header),
            "last_skin_difference_from_csv_c": abs(
                float(last[-1]) - float(skin.iloc[-1]["predicted_c"])
            ),
        },
    )

    comparison = load_yaml(RESULTS / "model_comparison.yaml")
    candidate = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>")
    means = candidate.groupby("model")["rmse_c"].mean()
    best_model = str(means.idxmin())
    add(
        "candidate_selection",
        "pass"
        if best_model == comparison["selected_model"] == "finite_volume"
        else "fail",
        {"mean_rmse_c": means.to_dict(), "selected": comparison["selected_model"]},
    )

    mesh = load_yaml(RESULTS / "mesh_study.yaml")
    add(
        "mesh_decision_invariance",
        str(mesh["manufacturing_decision_invariance_status"]),
        mesh,
    )
    diagnostics = load_yaml(RESULTS / "parameter_diagnostics.yaml")
    add(
        "calibration_multistart",
        str(diagnostics["multistart_status"]),
        {
            "count": diagnostics["multistart_count"],
            "h_out_spread": diagnostics["multistart_h_out_spread_w_m2k"],
            "h_in_spread": diagnostics["multistart_h_in_spread_w_m2k"],
        },
    )
    add(
        "residual_independence_assumption",
        str(diagnostics["independent_second_residual_assumption_status"]),
        {
            "lag1": diagnostics["residual_lag1_correlation"],
            "durbin_watson": diagnostics["durbin_watson"],
            "interpretation": "A fail rejects IID-per-second uncertainty calculations; it does not reject the deterministic fit.",
        },
    )

    independent_run = subprocess.run(
        [sys.executable, str(WORKSPACE / "code" / "independent_reference.py")],
        cwd=WORKSPACE,
        text=True,
        capture_output=True,
        check=False,
    )
    independent_report = (
        json.loads((RESULTS / "independent_verification.json").read_text(encoding="utf-8"))
        if (RESULTS / "independent_verification.json").is_file()
        else {"status": "fail", "reason": "report missing"}
    )
    add(
        "independent_raw_input_reference",
        "pass"
        if independent_run.returncode == 0 and independent_report.get("status") == "pass"
        else "fail",
        {
            "return_code": independent_run.returncode,
            "stdout": independent_run.stdout.strip(),
            "stderr": independent_run.stderr.strip(),
            "report_status": independent_report.get("status"),
            "failed_checks": independent_report.get("failed_checks", []),
        },
    )

    claims_document = load_yaml(RESULTS / "claims.yaml")
    markdown = PAPER_MD.read_text(encoding="utf-8")
    tex = "\n".join(
        path.read_text(encoding="utf-8")
        for path in sorted(PAPER.rglob("*.tex"))
    )
    claim_failures: list[dict[str, Any]] = []
    for claim in claims_document["claims"]:
        anchor = str(claim["paper_anchor"])
        regions = {"markdown": claim_region(markdown, anchor), "tex": claim_region(tex, anchor)}
        for paper_kind, region in regions.items():
            reasons: list[str] = []
            if region is None:
                reasons.append("anchor region missing")
            else:
                if not value_present(
                    region, float(claim["value"]), int(claim["display_decimals"])
                ):
                    reasons.append("rounded source value missing")
                missing_terms = [
                    term for term in claim.get("required_terms", []) if str(term) not in region
                ]
                present_forbidden = [
                    term for term in claim.get("forbidden_terms", []) if str(term) in region
                ]
                if missing_terms:
                    reasons.append(f"required semantic terms missing: {missing_terms}")
                if present_forbidden:
                    reasons.append(f"forbidden semantic terms present: {present_forbidden}")
                unit = str(claim["unit"])
                unit_ok = {
                    "mm": "mm" in region,
                    "s": "s" in region or "秒" in region,
                    "degC": "℃" in region or "circ" in region,
                }.get(unit, unit in region)
                if not unit_ok:
                    reasons.append(f"unit {unit} missing")
            if reasons:
                claim_failures.append(
                    {"claim": claim["id"], "paper": paper_kind, "reasons": reasons}
                )
    add(
        "paper_structured_claims",
        "pass" if not claim_failures else "fail",
        {
            "claim_count": len(claims_document["claims"]),
            "checked_paper_forms": ["markdown", "tex"],
            "failures": claim_failures,
        },
    )

    provenance = load_yaml(WORKSPACE / "provenance.yaml")
    provenance_ok = (
        provenance.get("phase") == "blind-revision"
        and provenance.get("automatic_retrieval", {}).get("cards") == []
        and provenance.get("sources_used")
        == ["input", "blind-v1", "audit", "$cumcm-a-solve"]
        and provenance.get("external_freeze_performed") is False
    )
    add("provenance_consistency", "pass" if provenance_ok else "fail", provenance)

    initial = load_yaml(RESULTS / "initial_condition_scenarios.yaml")
    gap = load_yaml(RESULTS / "gap_structure_scenarios.yaml")
    objective = load_yaml(RESULTS / "objective_weight_scenarios.yaml")
    add("authorized_initial_condition", "needs_review", initial)
    add("gap_structure_identifiability", "needs_review", gap)
    add("stakeholder_objective", "needs_review", objective)
    add(
        "external_experimental_validity",
        "needs_review",
        {"reason": "No measurements at 65 C, 80 C, or alternative gap thicknesses."},
    )

    tex_engine = next(
        (name for name in ("xelatex", "latexmk", "lualatex", "pdflatex") if shutil.which(name)),
        None,
    )
    add(
        "paper_compilation",
        "needs_review",
        {
            "available_engine": tex_engine,
            "compile_attempted": False,
            "reason": "No engine is available." if tex_engine is None else "Compilation is a separate final check.",
        },
    )
    add(
        "mathematical_correctness_claim",
        "needs_review",
        {"reason": "Automated agreement does not prove physical model truth or external safety."},
    )

    critical_ids = {
        "required_deliverables",
        "input_identity",
        "attachment_materials_are_solver_input",
        "problem1_workbook_semantics",
        "candidate_selection",
        "mesh_decision_invariance",
        "calibration_multistart",
        "independent_raw_input_reference",
        "paper_structured_claims",
        "provenance_consistency",
    }
    critical_failed = [
        item["id"]
        for item in checks
        if item["id"] in critical_ids and item["status"] != "pass"
    ]
    report = {
        "schema_version": 2,
        "phase": "blind-revision",
        "internal_verification_status": "pass" if not critical_failed else "fail",
        "overall_scientific_status": "needs_review",
        "critical_failures": critical_failed,
        "checks": checks,
    }
    (RESULTS / "verification.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "internal_verification_status": report["internal_verification_status"],
                "critical_failures": critical_failed,
            },
            ensure_ascii=False,
        )
    )
    if critical_failed:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
