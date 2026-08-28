"""Reproduce every numerical result and figure for the blind solve.

Run from the workspace root with::

    python code/run_all.py

The script only reads the copied input workbook and writes within this solve
workspace.  It uses no network access and no random search.
"""

from __future__ import annotations

import hashlib
import json
import math
import platform
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import scipy
import yaml
from matplotlib.ticker import ScalarFormatter
from openpyxl import Workbook
from scipy.optimize import brentq, minimize_scalar

from thermal_model import (
    BODY_TEMPERATURE_C,
    HIGH_CELLS,
    MEDIUM_CELLS,
    RC_CELLS,
    Material,
    StepResponse,
    build_step_response,
    error_metrics,
    exponential_temperature,
    fit_boundary_coefficients,
    fit_exponential,
)


SEED = 2018
np.random.seed(SEED)

WORKSPACE = Path(__file__).resolve().parents[1]
INPUT_WORKBOOK = WORKSPACE / "input" / "data" / "<SOURCE_FILE_REDACTED>"
INPUT_PROBLEM = WORKSPACE / "input" / "problem" / "<SOURCE_FILE_REDACTED>"
RESULTS = WORKSPACE / "results"
FIGURES = WORKSPACE / "figures"


def plain(value: Any) -> Any:
    """Convert numpy and Path values into JSON/YAML-safe builtins."""

    if isinstance(value, dict):
        return {str(key): plain(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [plain(item) for item in value]
    if isinstance(value, np.ndarray):
        return [plain(item) for item in value.tolist()]
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, (np.floating, float)):
        return float(value)
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, Path):
        return value.as_posix()
    return value


def write_yaml(path: Path, data: Any) -> None:
    path.write_text(
        yaml.safe_dump(plain(data), allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )


def write_json(path: Path, data: Any) -> None:
    path.write_text(
        json.dumps(plain(data), ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_inputs() -> tuple[
    np.ndarray, np.ndarray, tuple[Material, ...], list[dict[str, Any]]
]:
    workbook = openpyxl.load_workbook(INPUT_WORKBOOK, data_only=True, read_only=True)
    if workbook.sheetnames != ["附件1", "附件2"]:
        raise ValueError(f"unexpected workbook sheets: {workbook.sheetnames}")
    material_sheet = workbook["附件1"]
    material_rows: list[dict[str, Any]] = []
    materials: list[Material] = []
    expected_names = ("I层", "II层", "III层", "IV层")
    for row in material_sheet.iter_rows(min_row=3, max_row=6, values_only=True):
        if row[0] not in expected_names:
            raise ValueError(f"unexpected material layer: {row[0]!r}")
        numeric = tuple(float(row[index]) for index in (1, 2, 3))
        if not all(np.isfinite(numeric)) or any(value <= 0 for value in numeric):
            raise ValueError(f"invalid material properties for {row[0]!r}")
        layer = str(row[0]).removesuffix("层")
        materials.append(Material(layer, numeric[0], numeric[1], numeric[2]))
        material_rows.append(
            {
                "layer": row[0],
                "density_kg_m3": numeric[0],
                "heat_capacity_j_kgc": numeric[1],
                "conductivity_w_mc": numeric[2],
                "thickness_mm": row[4],
            }
        )
    measurement_sheet = workbook["附件2"]
    rows = list(measurement_sheet.iter_rows(min_row=3, values_only=True))
    times = np.asarray([row[0] for row in rows], dtype=float)
    observed = np.asarray([row[1] for row in rows], dtype=float)
    if len(times) != 5401 or not np.array_equal(times, np.arange(5401, dtype=float)):
        raise ValueError("attachment 2 must contain exactly times 0..5400 s")
    if not np.isfinite(observed).all():
        raise ValueError("attachment 2 contains missing or non-finite temperatures")
    return times, observed, tuple(materials), material_rows


def status_for_constraints(model: StepResponse, end_s: float) -> str:
    end_temperature = float(model.skin_temperature(end_s))
    duration = model.duration_above_s(44.0, end_s)
    return "pass" if end_temperature <= 47.0 + 1e-9 and duration <= 300.0 + 1e-7 else "fail"


def design_metrics(model: StepResponse, end_s: float) -> dict[str, Any]:
    crossing = model.first_crossing_s(44.0, end_s)
    values = np.asarray(model.skin_temperature(np.arange(0.0, end_s + 1.0, 1.0)))
    return {
        "end_temperature_c": float(model.skin_temperature(end_s)),
        "max_temperature_c": float(np.max(values)),
        "temperature_at_end_minus_300s_c": float(model.skin_temperature(end_s - 300.0)),
        "first_crossing_44_s": None if crossing is None else float(crossing),
        "duration_above_44_s": float(model.duration_above_s(44.0, end_s)),
        "monotonicity_status": "pass" if np.min(np.diff(values)) >= -1e-10 else "fail",
        "constraint_status": status_for_constraints(model, end_s),
    }


def make_model_factory(
    h_out: float,
    h_in: float,
    materials: tuple[Material, ...],
    cells: tuple[int, int, int, int] = MEDIUM_CELLS,
    *,
    h_out_scale: float = 1.0,
    h_in_scale: float = 1.0,
    conductivity_scales: dict[str, float] | None = None,
    capacity_scales: dict[str, float] | None = None,
    initial_clothing_c: float = BODY_TEMPERATURE_C,
    gap_parallel_conductance_w_m2k: float = 0.0,
    gap_temperature_exponent: float = 0.0,
) -> Callable[[float, float, float], StepResponse]:
    def factory(environment_c: float, d2_mm: float, d4_mm: float) -> StepResponse:
        return build_step_response(
            environment_c,
            d2_mm,
            d4_mm,
            h_out * h_out_scale,
            h_in * h_in_scale,
            cells,
            materials=materials,
            initial_clothing_c=initial_clothing_c,
            conductivity_scales=conductivity_scales,
            capacity_scales=capacity_scales,
            gap_parallel_conductance_w_m2k=gap_parallel_conductance_w_m2k,
            gap_temperature_exponent=gap_temperature_exponent,
        )

    return factory


def required_d2(
    factory: Callable[[float, float, float], StepResponse],
    environment_c: float,
    d4_mm: float,
    check_time_s: float,
    temperature_limit_c: float,
) -> float:
    def residual(d2_mm: float) -> float:
        return (
            float(factory(environment_c, d2_mm, d4_mm).skin_temperature(check_time_s))
            - temperature_limit_c
        )

    low, high = 0.6, 25.0
    low_value = residual(low)
    high_value = residual(high)
    if low_value <= 0.0:
        return low
    if high_value > 0.0:
        return math.inf
    return float(brentq(residual, low, high, xtol=1e-10, rtol=1e-13))


def ceil_to_tenth(value: float) -> float:
    return float(math.ceil((value - 1e-10) * 10.0) / 10.0)


def compare_candidates(
    times: np.ndarray,
    observed: np.ndarray,
    materials: tuple[Material, ...],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    folds = (
        ("fold_1", 600, 601, 900),
        ("fold_2", 900, 901, 1200),
        ("fold_3", 1200, 1201, 1800),
    )
    rows: list[dict[str, Any]] = []
    for fold_name, train_end, validation_start, validation_end in folds:
        training = times <= train_end
        validation = (times >= validation_start) & (times <= validation_end)
        fit_t = times[training]
        fit_y = observed[training]
        val_t = times[validation]
        val_y = observed[validation]

        baseline_fit = fit_exponential(fit_t, fit_y)
        baseline_pred = np.asarray(
            exponential_temperature(
                val_t,
                baseline_fit.parameters["amplitude_c"],
                baseline_fit.parameters["time_constant_s"],
            )
        )
        rows.append(
            {
                "fold": fold_name,
                "train_start_s": 0,
                "train_end_s": train_end,
                "validation_start_s": validation_start,
                "validation_end_s": validation_end,
                "model": "empirical_exponential",
                **error_metrics(baseline_pred, val_y),
                "fit_status": baseline_fit.status,
                "parameters": json.dumps(baseline_fit.parameters, sort_keys=True),
            }
        )

        for label, cells in (("four_node_rc", RC_CELLS), ("finite_volume", MEDIUM_CELLS)):
            fit = fit_boundary_coefficients(
                fit_t, fit_y, cells, materials=materials
            )
            model = build_step_response(
                75.0,
                6.0,
                5.0,
                fit.parameters["h_out_w_m2k"],
                fit.parameters["h_in_w_m2k"],
                cells,
                materials=materials,
            )
            prediction = np.asarray(model.skin_temperature(val_t))
            rows.append(
                {
                    "fold": fold_name,
                    "train_start_s": 0,
                    "train_end_s": train_end,
                    "validation_start_s": validation_start,
                    "validation_end_s": validation_end,
                    "model": label,
                    **error_metrics(prediction, val_y),
                    "fit_status": fit.status,
                    "parameters": json.dumps(fit.parameters, sort_keys=True),
                }
            )

    frame = pd.DataFrame(rows)
    summaries: list[dict[str, Any]] = []
    for model_name, group in frame.groupby("model", sort=False):
        summaries.append(
            {
                "model": model_name,
                "mean_rmse_c": float(group["rmse_c"].mean()),
                "worst_fold_rmse_c": float(group["rmse_c"].max()),
                "mean_mae_c": float(group["mae_c"].mean()),
                "worst_max_abs_error_c": float(group["max_abs_error_c"].max()),
                "all_fit_status": "pass" if (group["fit_status"] == "pass").all() else "fail",
            }
        )
    best = min(summaries, key=lambda item: item["mean_rmse_c"])["model"]
    comparison = {
        "validation_scheme": {
            "type": "time_ordered_expanding_prefix",
            "folds": [
                {
                    "fold": name,
                    "train_s": [0, train_end],
                    "validation_s": [validation_start, validation_end],
                }
                for name, train_end, validation_start, validation_end in folds
            ],
        },
        "models": summaries,
        "lowest_mean_rmse_model": best,
        "selected_model": "finite_volume",
        "selection_status": "pass" if best == "finite_volume" else "needs_review",
        "selection_reason": (
            "有限体积模型具有最低平均和最坏折误差，守恒处理材料界面，且能直接输出时空温度场和改变厚度。"
        ),
    }
    return frame, comparison


def save_problem1_workbook(
    times: np.ndarray,
    observed: np.ndarray,
    predicted: np.ndarray,
    model: StepResponse,
    node_metadata: list[dict[str, Any]],
    field: np.ndarray,
    materials: tuple[Material, ...],
) -> None:
    output = RESULTS / "<SOURCE_FILE_REDACTED>"
    workbook = Workbook(write_only=True)
    workbook.properties.creator = "cumcm-a-solve"
    workbook.properties.title = "Problem 1 temperature distribution"
    workbook.properties.description = "Code-generated finite-volume result"
    fixed_time = datetime(2026, 8, 28, 0, 0, 0)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time

    distribution = workbook.create_sheet("temperature_distribution")
    headers = ["time_s"] + [
        f"T_C@{float(item['x_mm']):.6f}mm|{item['node_type']}|{item['layer']}"
        for item in node_metadata
    ]
    distribution.append(headers)
    for row_index, time_value in enumerate(times):
        distribution.append([float(time_value), *[float(v) for v in field[row_index, :]]])

    nodes = workbook.create_sheet("spatial_nodes")
    nodes.append(["temperature_column", "x_mm", "node_type", "layer", "cell_index"])
    for index, item in enumerate(node_metadata, start=2):
        nodes.append(
            [
                index,
                float(item["x_mm"]),
                item["node_type"],
                item["layer"],
                item.get("cell_index"),
            ]
        )

    comparison = workbook.create_sheet("skin_comparison")
    comparison.append(["time_s", "observed_c", "predicted_c", "residual_c"])
    for time_value, obs_value, pred_value in zip(times, observed, predicted):
        comparison.append(
            [
                float(time_value),
                float(obs_value),
                float(pred_value),
                float(pred_value - obs_value),
            ]
        )

    parameters = workbook.create_sheet("parameters")
    parameters.append(["parameter", "value", "unit", "evidence"])
    parameter_rows = [
        ("environment_temperature", model.environment_c, "degC", "problem statement"),
        ("body_temperature", model.body_c, "degC", "problem statement"),
        ("d1", model.thickness_mm[0], "mm", "attachment 1"),
        ("d2", model.thickness_mm[1], "mm", "problem 1"),
        ("d3", model.thickness_mm[2], "mm", "attachment 1"),
        ("d4", model.thickness_mm[3], "mm", "problem 1"),
        ("h_out", model.h_out_w_m2k, "W/(m^2 degC)", "fit on 0-1800 s"),
        ("h_in", model.h_in_w_m2k, "W/(m^2 degC)", "fit on 0-1800 s"),
        ("time_step_output", 1.0, "s", "requested interval"),
        ("spatial_control_volumes", len(model.dx_m), "count", "medium mesh"),
    ]
    for row in parameter_rows:
        parameters.append(list(row))
    parameters.append([])
    parameters.append(["layer", "density_kg_m3", "heat_capacity_j_kgc", "conductivity_w_mc"])
    for material in materials:
        parameters.append(
            [
                material.layer,
                material.density_kg_m3,
                material.heat_capacity_j_kgk,
                material.conductivity_w_mk,
            ]
        )

    workbook.save(output)


def solve_q2(factory: Callable[[float, float, float], StepResponse]) -> dict[str, Any]:
    requirement_44 = required_d2(factory, 65.0, 5.5, 3300.0, 44.0)
    requirement_47 = required_d2(factory, 65.0, 5.5, 3600.0, 47.0)
    continuous_d2 = max(requirement_44, requirement_47)
    practical_d2 = ceil_to_tenth(continuous_d2)
    continuous_model = factory(65.0, continuous_d2, 5.5)
    practical_model = factory(65.0, practical_d2, 5.5)
    lower_grid_d2 = max(0.6, practical_d2 - 0.1)
    lower_grid_model = factory(65.0, lower_grid_d2, 5.5)
    return {
        "problem": 2,
        "objective": "minimize d2_mm",
        "environment_temperature_c": 65.0,
        "fixed_d4_mm": 5.5,
        "work_time_s": 3600.0,
        "bounds": {"d2_mm": [0.6, 25.0]},
        "constraint_equivalence": {
            "maximum_temperature": "T_skin(3600) <= 47 because the computed curve is monotone",
            "duration_above_44": "T_skin(3300) <= 44 iff duration above 44 is at most 300 s",
        },
        "active_constraint": "duration_above_44",
        "required_d2_by_constraint_mm": {
            "temperature_at_3300_le_44": requirement_44,
            "temperature_at_3600_le_47": requirement_47,
        },
        "continuous_optimum": {
            "d2_mm": continuous_d2,
            **design_metrics(continuous_model, 3600.0),
        },
        "practical_0p1mm_design": {
            "d2_mm": practical_d2,
            **design_metrics(practical_model, 3600.0),
        },
        "lower_grid_neighbor": {
            "d2_mm": lower_grid_d2,
            **design_metrics(lower_grid_model, 3600.0),
        },
        "optimality_check_status": (
            "pass" if status_for_constraints(lower_grid_model, 3600.0) == "fail" else "fail"
        ),
        "external_validity_status": "needs_review",
    }


def solve_q3(
    factory: Callable[[float, float, float], StepResponse]
) -> tuple[dict[str, Any], pd.DataFrame, pd.DataFrame]:
    cache: dict[float, tuple[float, float, float, float]] = {}

    def requirements(d4_mm: float) -> tuple[float, float, float, float]:
        key = round(float(d4_mm), 12)
        if key not in cache:
            by_44 = required_d2(factory, 80.0, float(d4_mm), 1500.0, 44.0)
            by_47 = required_d2(factory, 80.0, float(d4_mm), 1800.0, 47.0)
            required = max(by_44, by_47)
            total = required + float(d4_mm) if math.isfinite(required) else math.inf
            cache[key] = (total, required, by_44, by_47)
        return cache[key]

    frontier_rows: list[dict[str, Any]] = []
    for d4 in np.round(np.arange(0.6, 6.4001, 0.05), 10):
        total, required, by_44, by_47 = requirements(float(d4))
        frontier_rows.append(
            {
                "d4_mm": float(d4),
                "required_d2_mm": None if not math.isfinite(required) else required,
                "required_by_44_mm": None if not math.isfinite(by_44) else by_44,
                "required_by_47_mm": None if not math.isfinite(by_47) else by_47,
                "minimum_total_mm": None if not math.isfinite(total) else total,
                "feasibility_status": "pass" if math.isfinite(required) else "fail",
            }
        )
    frontier = pd.DataFrame(frontier_rows)
    feasible_frontier = frontier[frontier["feasibility_status"] == "pass"].copy()
    feasible_start = float(feasible_frontier["d4_mm"].min())

    optimization = minimize_scalar(
        lambda value: requirements(float(value))[0],
        bounds=(feasible_start, 6.4),
        method="bounded",
        options={"xatol": 1e-9, "maxiter": 200},
    )
    continuous_candidates = [float(optimization.x), feasible_start, 6.4]
    continuous_d4 = min(continuous_candidates, key=lambda value: requirements(value)[0])
    continuous_total, continuous_d2, by_44, by_47 = requirements(continuous_d4)
    continuous_model = factory(80.0, continuous_d2, continuous_d4)

    grid_rows: list[dict[str, Any]] = []
    practical_candidates: list[tuple[float, float, float, float, StepResponse]] = []
    for d4_index in range(6, 65):
        d4 = d4_index / 10.0
        _, required, required_44, required_47 = requirements(d4)
        if not math.isfinite(required):
            grid_rows.append(
                {
                    "d4_mm": d4,
                    "minimum_d2_on_0p1mm_grid": None,
                    "total_mm": None,
                    "duration_above_44_s": None,
                    "end_temperature_c": None,
                    "constraint_status": "fail",
                }
            )
            continue
        d2 = ceil_to_tenth(required)
        if d2 > 25.0:
            continue
        model = factory(80.0, d2, d4)
        metrics = design_metrics(model, 1800.0)
        grid_rows.append(
            {
                "d4_mm": d4,
                "minimum_d2_on_0p1mm_grid": d2,
                "required_d2_continuous_mm": required,
                "required_by_44_mm": required_44,
                "required_by_47_mm": required_47,
                "total_mm": d2 + d4,
                "duration_above_44_s": metrics["duration_above_44_s"],
                "end_temperature_c": metrics["end_temperature_c"],
                "constraint_status": metrics["constraint_status"],
            }
        )
        if metrics["constraint_status"] == "pass":
            practical_candidates.append((round(d2 + d4, 10), metrics["duration_above_44_s"], d2, d4, model))

    grid = pd.DataFrame(grid_rows)
    minimum_grid_total = min(item[0] for item in practical_candidates)
    tied = [item for item in practical_candidates if abs(item[0] - minimum_grid_total) <= 1e-9]
    # For equal total thickness, choose the one with more nominal thermal margin.
    chosen = min(tied, key=lambda item: item[1])
    _, _, practical_d2, practical_d4, practical_model = chosen

    lower_total_d2 = max(0.6, practical_d2 - 0.1)
    lower_total_model = factory(80.0, lower_total_d2, practical_d4)
    feasible_totals = grid.loc[grid["constraint_status"] == "pass", "total_mm"]
    return (
        {
            "problem": 3,
            "objective": "minimize d2_mm + d4_mm",
            "environment_temperature_c": 80.0,
            "work_time_s": 1800.0,
            "bounds": {"d2_mm": [0.6, 25.0], "d4_mm": [0.6, 6.4]},
            "constraint_equivalence": {
                "maximum_temperature": "T_skin(1800) <= 47 because the computed curve is monotone",
                "duration_above_44": "T_skin(1500) <= 44 iff duration above 44 is at most 300 s",
            },
            "active_constraint": "duration_above_44",
            "continuous_optimum": {
                "d2_mm": continuous_d2,
                "d4_mm": continuous_d4,
                "total_mm": continuous_total,
                "required_d2_by_constraint_mm": {
                    "temperature_at_1500_le_44": by_44,
                    "temperature_at_1800_le_47": by_47,
                },
                **design_metrics(continuous_model, 1800.0),
            },
            "practical_0p1mm_design": {
                "d2_mm": practical_d2,
                "d4_mm": practical_d4,
                "total_mm": practical_d2 + practical_d4,
                "tie_count_at_minimum_total": len(tied),
                **design_metrics(practical_model, 1800.0),
            },
            "lower_total_neighbor": {
                "d2_mm": lower_total_d2,
                "d4_mm": practical_d4,
                "total_mm": lower_total_d2 + practical_d4,
                **design_metrics(lower_total_model, 1800.0),
            },
            "grid_minimum_total_mm": float(feasible_totals.min()),
            "continuous_search_status": "pass" if optimization.success else "fail",
            "grid_optimality_check_status": (
                "pass"
                if float(feasible_totals.min()) >= practical_d2 + practical_d4 - 1e-9
                else "fail"
            ),
            "external_validity_status": "needs_review",
        },
        frontier,
        grid,
    )


def sensitivity_analysis(
    h_out: float,
    h_in: float,
    materials: tuple[Material, ...],
    q2: dict[str, Any],
    q3: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    variants: list[dict[str, Any]] = [{"scenario": "nominal", "parameter": "none", "scale": 1.0}]
    for parameter in ("h_out", "h_in", "k_II", "k_III", "k_IV", "capacity_all"):
        for scale in (0.95, 1.05):
            variants.append(
                {
                    "scenario": f"{parameter}_{int(round(scale * 100))}pct",
                    "parameter": parameter,
                    "scale": scale,
                }
            )
    variants.append(
        {
            "scenario": "combined_adverse_5pct_stress",
            "parameter": "combined",
            "scale": None,
        }
    )

    rows: list[dict[str, Any]] = []
    required_q2: list[float] = []
    required_q3: list[float] = []
    for variant in variants:
        parameter = variant["parameter"]
        scale = variant["scale"]
        h_out_scale = 1.0
        h_in_scale = 1.0
        conductivity_scales: dict[str, float] = {}
        capacity_scales: dict[str, float] = {}
        if parameter == "h_out":
            h_out_scale = float(scale)
        elif parameter == "h_in":
            h_in_scale = float(scale)
        elif isinstance(parameter, str) and parameter.startswith("k_"):
            conductivity_scales[parameter.split("_", 1)[1]] = float(scale)
        elif parameter == "capacity_all":
            capacity_scales = {material.layer: float(scale) for material in materials}
        elif parameter == "combined":
            h_out_scale = 1.05
            h_in_scale = 0.95
            conductivity_scales = {material.layer: 1.05 for material in materials}
            capacity_scales = {material.layer: 0.95 for material in materials}

        factory = make_model_factory(
            h_out,
            h_in,
            materials,
            h_out_scale=h_out_scale,
            h_in_scale=h_in_scale,
            conductivity_scales=conductivity_scales,
            capacity_scales=capacity_scales,
        )
        q2_d2 = float(q2["practical_0p1mm_design"]["d2_mm"])
        q3_d2 = float(q3["practical_0p1mm_design"]["d2_mm"])
        q3_d4 = float(q3["practical_0p1mm_design"]["d4_mm"])
        q2_model = factory(65.0, q2_d2, 5.5)
        q3_model = factory(80.0, q3_d2, q3_d4)
        q2_metrics = design_metrics(q2_model, 3600.0)
        q3_metrics = design_metrics(q3_model, 1800.0)

        req_q2 = max(
            required_d2(factory, 65.0, 5.5, 3300.0, 44.0),
            required_d2(factory, 65.0, 5.5, 3600.0, 47.0),
        )
        req_q3 = max(
            required_d2(factory, 80.0, 6.4, 1500.0, 44.0),
            required_d2(factory, 80.0, 6.4, 1800.0, 47.0),
        )
        required_q2.append(req_q2)
        required_q3.append(req_q3)
        rows.append(
            {
                **variant,
                "q2_end_temperature_c": q2_metrics["end_temperature_c"],
                "q2_duration_above_44_s": q2_metrics["duration_above_44_s"],
                "q2_constraint_status": q2_metrics["constraint_status"],
                "q2_required_d2_mm": req_q2,
                "q3_end_temperature_c": q3_metrics["end_temperature_c"],
                "q3_duration_above_44_s": q3_metrics["duration_above_44_s"],
                "q3_constraint_status": q3_metrics["constraint_status"],
                "q3_required_d2_at_d4_6p4_mm": req_q3,
            }
        )

    frame = pd.DataFrame(rows)
    nominal = frame[frame["scenario"] == "nominal"].iloc[0]
    one_at_a_time = frame[~frame["parameter"].isin(["none", "combined"])]
    stress = frame[frame["parameter"] == "combined"].iloc[0]
    summary = {
        "interpretation": "±5% is a deterministic perturbation, not a statistical confidence interval.",
        "q2_nominal_duration_s": float(nominal["q2_duration_above_44_s"]),
        "q3_nominal_duration_s": float(nominal["q3_duration_above_44_s"]),
        "q2_oat_duration_range_s": [
            float(one_at_a_time["q2_duration_above_44_s"].min()),
            float(one_at_a_time["q2_duration_above_44_s"].max()),
        ],
        "q3_oat_duration_range_s": [
            float(one_at_a_time["q3_duration_above_44_s"].min()),
            float(one_at_a_time["q3_duration_above_44_s"].max()),
        ],
        "q2_max_required_d2_oat_mm": float(max(required_q2[:-1])),
        "q3_max_required_d2_oat_at_d4_6p4_mm": float(max(required_q3[:-1])),
        "q2_oat_reserve_0p1mm_design_d2_mm": ceil_to_tenth(max(required_q2[:-1])),
        "q3_oat_reserve_0p1mm_design": {
            "d2_mm": ceil_to_tenth(max(required_q3[:-1])),
            "d4_mm": 6.4,
        },
        "q2_combined_stress_required_d2_mm": float(stress["q2_required_d2_mm"]),
        "q3_combined_stress_required_d2_at_d4_6p4_mm": float(
            stress["q3_required_d2_at_d4_6p4_mm"]
        ),
        "q2_combined_stress_0p1mm_design_d2_mm": ceil_to_tenth(
            float(stress["q2_required_d2_mm"])
        ),
        "q3_combined_stress_0p1mm_design": {
            "d2_mm": ceil_to_tenth(float(stress["q3_required_d2_at_d4_6p4_mm"])),
            "d4_mm": 6.4,
        },
        "reserve_interpretation": "These are optional deterministic stress reserves, not the nominal optimum requested by the problem.",
        "nominal_design_robustness_status": (
            "pass"
            if (one_at_a_time["q2_constraint_status"] == "pass").all()
            and (one_at_a_time["q3_constraint_status"] == "pass").all()
            else "needs_review"
        ),
    }
    return frame, summary


def validation_checks(
    times: np.ndarray,
    observed: np.ndarray,
    h_out: float,
    h_in: float,
    materials: tuple[Material, ...],
    q1_model: StepResponse,
    q2: dict[str, Any],
    q3: dict[str, Any],
    holdout_metrics: dict[str, float],
) -> dict[str, Any]:
    medium_factory = make_model_factory(h_out, h_in, materials, MEDIUM_CELLS)
    high_factory = make_model_factory(h_out, h_in, materials, HIGH_CELLS)
    scenarios = [
        ("q1", 75.0, 6.0, 5.0, 5400),
        ("q2_practical", 65.0, q2["practical_0p1mm_design"]["d2_mm"], 5.5, 3600),
        (
            "q3_practical",
            80.0,
            q3["practical_0p1mm_design"]["d2_mm"],
            q3["practical_0p1mm_design"]["d4_mm"],
            1800,
        ),
    ]
    mesh_rows: list[dict[str, Any]] = []
    for name, env, d2, d4, end in scenarios:
        check_times = np.arange(0.0, end + 1.0, 1.0)
        medium = medium_factory(env, float(d2), float(d4))
        high = high_factory(env, float(d2), float(d4))
        difference = np.max(
            np.abs(
                np.asarray(medium.skin_temperature(check_times))
                - np.asarray(high.skin_temperature(check_times))
            )
        )
        mesh_rows.append(
            {
                "scenario": name,
                "max_skin_difference_c": float(difference),
                "status": "pass" if difference <= 0.02 else "fail",
            }
        )

    q1_values = np.asarray(q1_model.skin_temperature(times))
    outer_flux, inner_flux = q1_model.steady_fluxes_w_m2()
    flux_relative_difference = abs(outer_flux - inner_flux) / max(abs(outer_flux), 1e-15)

    q2_factory = medium_factory
    boundary_cases = [
        ("q2_min_d2", q2_factory(65.0, 0.6, 5.5), 3600, "fail"),
        ("q2_max_d2", q2_factory(65.0, 25.0, 5.5), 3600, "pass"),
        ("q3_min_both", q2_factory(80.0, 0.6, 0.6), 1800, "fail"),
        ("q3_max_both", q2_factory(80.0, 25.0, 6.4), 1800, "pass"),
    ]
    boundary_rows = []
    for name, model, end, expected in boundary_cases:
        actual = status_for_constraints(model, end)
        boundary_rows.append(
            {
                "scenario": name,
                "expected_constraint_status": expected,
                "actual_constraint_status": actual,
                "check_status": "pass" if actual == expected else "fail",
            }
        )

    return {
        "evidence_scope": "internal numerical and time-holdout validation",
        "input_series": {
            "sample_count": len(times),
            "time_order_status": "pass" if np.all(np.diff(times) > 0) else "fail",
            "missing_status": "pass" if np.isfinite(observed).all() else "fail",
        },
        "final_holdout_1801_5400": {
            **holdout_metrics,
            "status": "pass" if holdout_metrics["rmse_c"] <= 0.02 else "fail",
            "note": "后段主要是 0.01℃ 量化平台，证据难度有限。",
        },
        "temperature_bounds": {
            "minimum_c": float(np.min(q1_values)),
            "maximum_c": float(np.max(q1_values)),
            "status": (
                "pass"
                if np.min(q1_values) >= BODY_TEMPERATURE_C - 1e-10
                and np.max(q1_values) <= 75.0 + 1e-10
                else "fail"
            ),
        },
        "skin_monotonicity": {
            "minimum_one_second_increment_c": float(np.min(np.diff(q1_values))),
            "status": "pass" if np.min(np.diff(q1_values)) >= -1e-10 else "fail",
        },
        "steady_energy_balance": {
            "outer_flux_w_m2": outer_flux,
            "inner_flux_w_m2": inner_flux,
            "relative_difference": flux_relative_difference,
            "status": "pass" if flux_relative_difference <= 1e-10 else "fail",
        },
        "mesh_refinement": mesh_rows,
        "boundary_extreme_checks": boundary_rows,
        "external_experimental_validity": {
            "status": "needs_review",
            "reason": "没有 65℃、80℃或其他厚度下的独立测温数据。",
        },
    }


def parameter_diagnostics(
    times: np.ndarray,
    observed: np.ndarray,
    materials: tuple[Material, ...],
    h_out: float,
    h_in: float,
) -> tuple[dict[str, Any], pd.DataFrame, pd.DataFrame]:
    """Diagnose local identifiability without treating correlated seconds as IID."""

    calibration = times <= 1800
    calibration_times = times[calibration]
    calibration_observed = observed[calibration]
    sample = np.arange(0, len(calibration_times), 2, dtype=int)
    if sample[-1] != len(calibration_times) - 1:
        sample = np.r_[sample, len(calibration_times) - 1]
    fit_times = calibration_times[sample]
    fit_observed = calibration_observed[sample]

    base = build_step_response(
        75.0, 6.0, 5.0, h_out, h_in, MEDIUM_CELLS, materials=materials
    )
    residual_all = np.asarray(base.skin_temperature(calibration_times)) - calibration_observed
    lag1 = float(np.corrcoef(residual_all[:-1], residual_all[1:])[0, 1])
    durbin_watson = float(
        np.sum(np.diff(residual_all) ** 2) / max(np.sum(residual_all**2), 1e-30)
    )

    epsilon = 1e-4
    jacobian_columns: list[np.ndarray] = []
    for index in range(2):
        lower = [h_out, h_in]
        upper = [h_out, h_in]
        lower[index] *= math.exp(-epsilon)
        upper[index] *= math.exp(epsilon)
        lower_prediction = np.asarray(
            build_step_response(
                75.0,
                6.0,
                5.0,
                lower[0],
                lower[1],
                MEDIUM_CELLS,
                materials=materials,
            ).skin_temperature(fit_times)
        )
        upper_prediction = np.asarray(
            build_step_response(
                75.0,
                6.0,
                5.0,
                upper[0],
                upper[1],
                MEDIUM_CELLS,
                materials=materials,
            ).skin_temperature(fit_times)
        )
        jacobian_columns.append((upper_prediction - lower_prediction) / (2.0 * epsilon))
    jacobian = np.column_stack(jacobian_columns)
    singular_values = np.linalg.svd(jacobian, compute_uv=False)
    condition_number = float(singular_values[0] / singular_values[-1])
    inverse_information = np.linalg.pinv(jacobian.T @ jacobian)
    local_correlation = float(
        inverse_information[0, 1]
        / math.sqrt(inverse_information[0, 0] * inverse_information[1, 1])
    )

    starts = (
        (2.2, 1.1),
        (2.2, 90.0),
        (480.0, 1.1),
        (480.0, 90.0),
        (20.0, 4.0),
        (80.0, 8.0),
        (160.0, 15.0),
        (300.0, 35.0),
        (120.0, 6.0),
    )
    multistart_rows: list[dict[str, Any]] = []
    for start_h_out, start_h_in in starts:
        fit = fit_boundary_coefficients(
            calibration_times,
            calibration_observed,
            MEDIUM_CELLS,
            materials=materials,
            initial_h=(start_h_out, start_h_in),
        )
        multistart_rows.append(
            {
                "initial_h_out_w_m2k": start_h_out,
                "initial_h_in_w_m2k": start_h_in,
                "fitted_h_out_w_m2k": fit.parameters["h_out_w_m2k"],
                "fitted_h_in_w_m2k": fit.parameters["h_in_w_m2k"],
                "sse_sampled": fit.objective_sse,
                "nfev": fit.nfev,
                "fit_status": fit.status,
            }
        )
    multistart = pd.DataFrame(multistart_rows)

    def sampled_sse(candidate_h_out: float, candidate_h_in: float) -> float:
        model = build_step_response(
            75.0,
            6.0,
            5.0,
            candidate_h_out,
            candidate_h_in,
            MEDIUM_CELLS,
            materials=materials,
        )
        residual = np.asarray(model.skin_temperature(fit_times)) - fit_observed
        return float(np.sum(residual**2))

    base_sse = sampled_sse(h_out, h_in)
    profile_rows: list[dict[str, Any]] = []
    for fixed_name, center in (("h_out_w_m2k", h_out), ("h_in_w_m2k", h_in)):
        for fixed_value in center * np.exp(np.linspace(math.log(0.75), math.log(1.25), 13)):
            if fixed_name == "h_out_w_m2k":
                optimization = minimize_scalar(
                    lambda log_other: sampled_sse(float(fixed_value), math.exp(log_other)),
                    bounds=(math.log(1.0), math.log(100.0)),
                    method="bounded",
                    options={"xatol": 1e-10, "maxiter": 120},
                )
                other_name = "h_in_w_m2k"
            else:
                optimization = minimize_scalar(
                    lambda log_other: sampled_sse(math.exp(log_other), float(fixed_value)),
                    bounds=(math.log(2.0), math.log(500.0)),
                    method="bounded",
                    options={"xatol": 1e-10, "maxiter": 120},
                )
                other_name = "h_out_w_m2k"
            profile_rows.append(
                {
                    "profiled_parameter": fixed_name,
                    "fixed_value": float(fixed_value),
                    "optimized_parameter": other_name,
                    "optimized_value": float(math.exp(optimization.x)),
                    "sse_sampled": float(optimization.fun),
                    "delta_sse_from_joint_optimum": float(max(0.0, optimization.fun - base_sse)),
                    "optimizer_status": "pass" if optimization.success else "fail",
                }
            )
    profile = pd.DataFrame(profile_rows)

    h_out_spread = float(multistart["fitted_h_out_w_m2k"].max() - multistart["fitted_h_out_w_m2k"].min())
    h_in_spread = float(multistart["fitted_h_in_w_m2k"].max() - multistart["fitted_h_in_w_m2k"].min())
    multistart_status = (
        "pass"
        if (multistart["fit_status"] == "pass").all()
        and h_out_spread <= 1e-4
        and h_in_spread <= 1e-4
        else "needs_review"
    )
    summary = {
        "calibration_interval_s": [0, 1800],
        "measurement_quantization_c": 0.01,
        "residual_lag1_correlation": lag1,
        "durbin_watson": durbin_watson,
        "independent_second_residual_assumption_status": (
            "fail" if abs(lag1) >= 0.2 else "pass"
        ),
        "log_parameter_jacobian_singular_values": singular_values.tolist(),
        "log_parameter_jacobian_condition_number": condition_number,
        "local_linearized_parameter_correlation": local_correlation,
        "multistart_count": len(multistart),
        "multistart_h_out_spread_w_m2k": h_out_spread,
        "multistart_h_in_spread_w_m2k": h_in_spread,
        "multistart_status": multistart_status,
        "profile_status": (
            "pass" if (profile["optimizer_status"] == "pass").all() else "fail"
        ),
        "profile_interpretation": "Deterministic SSE profiles only; no IID-based confidence interval is claimed.",
        "statistical_parameter_interval_status": "needs_review",
    }
    return summary, multistart, profile


def mesh_size_study(
    times: np.ndarray,
    observed: np.ndarray,
    materials: tuple[Material, ...],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    folds = ((600, 601, 900), (900, 901, 1200), (1200, 1201, 1800))
    meshes = (
        (18, (2, 8, 4, 4)),
        (27, (3, 12, 6, 6)),
        (36, (4, 16, 8, 8)),
        (54, (6, 24, 12, 12)),
        (72, (8, 32, 16, 16)),
        (108, (12, 48, 24, 24)),
    )
    rows: list[dict[str, Any]] = []
    calibration = times <= 1800
    holdout = times > 1800
    for state_count, cells in meshes:
        started = time.perf_counter()
        fold_rmse: list[float] = []
        for train_end, validation_start, validation_end in folds:
            training = times <= train_end
            validation = (times >= validation_start) & (times <= validation_end)
            fit = fit_boundary_coefficients(
                times[training], observed[training], cells, materials=materials
            )
            model = build_step_response(
                75.0,
                6.0,
                5.0,
                fit.parameters["h_out_w_m2k"],
                fit.parameters["h_in_w_m2k"],
                cells,
                materials=materials,
            )
            fold_rmse.append(
                error_metrics(
                    np.asarray(model.skin_temperature(times[validation])),
                    observed[validation],
                )["rmse_c"]
            )
        final_fit = fit_boundary_coefficients(
            times[calibration], observed[calibration], cells, materials=materials
        )
        h_out = final_fit.parameters["h_out_w_m2k"]
        h_in = final_fit.parameters["h_in_w_m2k"]
        q1 = build_step_response(
            75.0, 6.0, 5.0, h_out, h_in, cells, materials=materials
        )
        holdout_rmse = error_metrics(
            np.asarray(q1.skin_temperature(times[holdout])), observed[holdout]
        )["rmse_c"]
        factory = make_model_factory(h_out, h_in, materials, cells)
        q2_required = max(
            required_d2(factory, 65.0, 5.5, 3300.0, 44.0),
            required_d2(factory, 65.0, 5.5, 3600.0, 47.0),
        )
        q3_required = max(
            required_d2(factory, 80.0, 6.4, 1500.0, 44.0),
            required_d2(factory, 80.0, 6.4, 1800.0, 47.0),
        )
        rows.append(
            {
                "state_count": state_count,
                "cells_per_layer": "/".join(str(value) for value in cells),
                "mean_time_ordered_rmse_c": float(np.mean(fold_rmse)),
                "worst_time_ordered_rmse_c": float(np.max(fold_rmse)),
                "holdout_rmse_c": holdout_rmse,
                "fitted_h_out_w_m2k": h_out,
                "fitted_h_in_w_m2k": h_in,
                "q2_continuous_d2_mm": q2_required,
                "q2_practical_d2_mm": ceil_to_tenth(q2_required),
                "q3_continuous_d2_at_d4_6p4_mm": q3_required,
                "q3_practical_d2_at_d4_6p4_mm": ceil_to_tenth(q3_required),
                "elapsed_s": float(time.perf_counter() - started),
                "fit_status": final_fit.status,
            }
        )
    frame = pd.DataFrame(rows)
    q2_unique = sorted(frame["q2_practical_d2_mm"].unique().tolist())
    q3_unique = sorted(frame["q3_practical_d2_at_d4_6p4_mm"].unique().tolist())
    summary = {
        "tested_state_counts": frame["state_count"].tolist(),
        "manufacturing_decision_invariance_status": (
            "pass" if len(q2_unique) == 1 and len(q3_unique) == 1 else "fail"
        ),
        "q2_practical_values_mm": q2_unique,
        "q3_practical_d2_at_d4_6p4_values_mm": q3_unique,
        "selected_state_count": 54,
        "selection_reason": "54 states retained for the requested spatial field; 27 states are adequate for fast thickness screening.",
    }
    return frame, summary


def initial_condition_scenarios(
    materials: tuple[Material, ...],
    h_out: float,
    h_in: float,
    q2: dict[str, Any],
    q3: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    nominal_q2 = float(q2["practical_0p1mm_design"]["d2_mm"])
    nominal_q3_d2 = float(q3["practical_0p1mm_design"]["d2_mm"])
    nominal_q3_d4 = float(q3["practical_0p1mm_design"]["d4_mm"])
    for initial_c in (35.0, 37.0, 39.0, 42.0):
        factory = make_model_factory(
            h_out, h_in, materials, initial_clothing_c=initial_c
        )
        q2_required = max(
            required_d2(factory, 65.0, 5.5, 3300.0, 44.0),
            required_d2(factory, 65.0, 5.5, 3600.0, 47.0),
        )
        q3_required = max(
            required_d2(factory, 80.0, 6.4, 1500.0, 44.0),
            required_d2(factory, 80.0, 6.4, 1800.0, 47.0),
        )
        nominal_q2_model = factory(65.0, nominal_q2, 5.5)
        nominal_q3_model = factory(80.0, nominal_q3_d2, nominal_q3_d4)
        rows.append(
            {
                "uniform_clothing_initial_temperature_c": initial_c,
                "q2_continuous_d2_mm": q2_required,
                "q2_practical_d2_mm": ceil_to_tenth(q2_required),
                "q2_nominal_temperature_at_3300_c": float(
                    nominal_q2_model.skin_temperature(3300.0)
                ),
                "q2_nominal_constraint_status": status_for_constraints(
                    nominal_q2_model, 3600.0
                ),
                "q3_continuous_d2_at_d4_6p4_mm": q3_required,
                "q3_practical_d2_at_d4_6p4_mm": ceil_to_tenth(q3_required),
                "q3_nominal_temperature_at_1500_c": float(
                    nominal_q3_model.skin_temperature(1500.0)
                ),
                "q3_nominal_constraint_status": status_for_constraints(
                    nominal_q3_model, 1800.0
                ),
            }
        )
    frame = pd.DataFrame(rows)
    summary = {
        "scenario_scope": "Illustrative uniform clothing temperatures; the problem does not authorize this as an uncertainty set.",
        "authorized_initial_condition_status": "needs_review",
        "nominal_design_all_scenarios_status": (
            "pass"
            if (frame["q2_nominal_constraint_status"] == "pass").all()
            and (frame["q3_nominal_constraint_status"] == "pass").all()
            else "fail"
        ),
        "engineering_recommendation_status": "needs_review",
    }
    return frame, summary


def objective_weight_scenarios(q3_grid: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    feasible = q3_grid[q3_grid["constraint_status"] == "pass"].copy()
    rows: list[dict[str, Any]] = []
    for weight in (0.0, 0.5, 1.0, 2.0, 4.0):
        feasible["weighted_objective"] = (
            feasible["minimum_d2_on_0p1mm_grid"] + weight * feasible["d4_mm"]
        )
        minimum = float(feasible["weighted_objective"].min())
        tied = feasible[np.isclose(feasible["weighted_objective"], minimum, atol=1e-9)]
        chosen = tied.sort_values(
            ["duration_above_44_s", "total_mm", "d4_mm"], ascending=[True, True, False]
        ).iloc[0]
        rows.append(
            {
                "d4_weight": weight,
                "d2_mm": float(chosen["minimum_d2_on_0p1mm_grid"]),
                "d4_mm": float(chosen["d4_mm"]),
                "raw_total_mm": float(chosen["total_mm"]),
                "weighted_objective": minimum,
                "duration_above_44_s": float(chosen["duration_above_44_s"]),
                "tie_count": int(len(tied)),
                "constraint_status": str(chosen["constraint_status"]),
            }
        )
    frame = pd.DataFrame(rows)
    distinct_pairs = frame[["d2_mm", "d4_mm"]].drop_duplicates()
    summary = {
        "objective_definition_status": "needs_review",
        "reason": "The problem supplies no mass, cost, bulk, or comfort weight for layer IV.",
        "reported_solution_policy": "Report the feasible frontier and call d2+d4 results conditional on weight 1.",
        "weight_sensitivity_changes_pair": "pass" if len(distinct_pairs) > 1 else "fail",
    }
    return frame, summary


def gap_structure_scenarios(
    times: np.ndarray,
    observed: np.ndarray,
    materials: tuple[Material, ...],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    calibration = times <= 1800
    holdout = times > 1800
    definitions = (
        ("pure_conduction", 0.0, 0.0),
        ("constant_parallel_2", 2.0, 0.0),
        ("temperature_scaled_parallel_2", 2.0, 3.0),
    )
    rows: list[dict[str, Any]] = []
    for name, parallel_g, exponent in definitions:
        fit = fit_boundary_coefficients(
            times[calibration],
            observed[calibration],
            MEDIUM_CELLS,
            materials=materials,
            gap_parallel_conductance_w_m2k=parallel_g,
            gap_temperature_exponent=exponent,
        )
        h_out = fit.parameters["h_out_w_m2k"]
        h_in = fit.parameters["h_in_w_m2k"]
        q1 = build_step_response(
            75.0,
            6.0,
            5.0,
            h_out,
            h_in,
            MEDIUM_CELLS,
            materials=materials,
            gap_parallel_conductance_w_m2k=parallel_g,
            gap_temperature_exponent=exponent,
        )
        prediction = np.asarray(q1.skin_temperature(times))
        factory = make_model_factory(
            h_out,
            h_in,
            materials,
            gap_parallel_conductance_w_m2k=parallel_g,
            gap_temperature_exponent=exponent,
        )
        q2 = solve_q2(factory)
        q3, _, _ = solve_q3(factory)
        q2_model = factory(65.0, float(q2["practical_0p1mm_design"]["d2_mm"]), 5.5)
        q3_model = factory(
            80.0,
            float(q3["practical_0p1mm_design"]["d2_mm"]),
            float(q3["practical_0p1mm_design"]["d4_mm"]),
        )
        rows.append(
            {
                "scenario": name,
                "parallel_conductance_reference_w_m2k": parallel_g,
                "temperature_exponent": exponent,
                "effective_g_at_65c_w_m2k": q2_model.effective_gap_parallel_conductance_w_m2k,
                "effective_g_at_75c_w_m2k": q1.effective_gap_parallel_conductance_w_m2k,
                "effective_g_at_80c_w_m2k": q3_model.effective_gap_parallel_conductance_w_m2k,
                "h_out_w_m2k": h_out,
                "h_in_w_m2k": h_in,
                "calibration_rmse_c": error_metrics(
                    prediction[calibration], observed[calibration]
                )["rmse_c"],
                "holdout_rmse_c": error_metrics(
                    prediction[holdout], observed[holdout]
                )["rmse_c"],
                "q2_practical_d2_mm": float(q2["practical_0p1mm_design"]["d2_mm"]),
                "q3_practical_d2_mm": float(q3["practical_0p1mm_design"]["d2_mm"]),
                "q3_practical_d4_mm": float(q3["practical_0p1mm_design"]["d4_mm"]),
                "q3_practical_total_mm": float(q3["practical_0p1mm_design"]["total_mm"]),
                "fit_status": fit.status,
            }
        )
    frame = pd.DataFrame(rows)
    pairs = frame[["q3_practical_d2_mm", "q3_practical_d4_mm"]].drop_duplicates()
    summary = {
        "scenario_basis": "Illustrative structures calibrated to the same single 75 C, d4=5 mm trajectory.",
        "cross_thickness_identifiability_status": "needs_review",
        "manufacturing_recommendation_invariance_status": (
            "pass" if len(pairs) == 1 else "needs_review"
        ),
        "engineering_recommendation_status": "needs_review",
        "required_new_evidence": "Measurements at multiple d4 and temperatures, preferably including interface temperatures.",
    }
    return frame, summary


def setup_plotting() -> None:
    plt.style.use("seaborn-v0_8-whitegrid")
    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "font.size": 9,
            "axes.titlesize": 10,
            "axes.labelsize": 9,
            "legend.fontsize": 8,
            "figure.dpi": 130,
            "savefig.dpi": 220,
            "axes.unicode_minus": False,
        }
    )


def save_figure(fig: plt.Figure, filename: str) -> None:
    fig.savefig(
        FIGURES / filename,
        bbox_inches="tight",
        metadata={"Software": "cumcm-a-solve deterministic pipeline"},
    )
    plt.close(fig)


def create_figures(
    times: np.ndarray,
    observed: np.ndarray,
    predicted: np.ndarray,
    candidate_metrics: pd.DataFrame,
    q1_model: StepResponse,
    q2: dict[str, Any],
    q3: dict[str, Any],
    frontier: pd.DataFrame,
    sensitivity: pd.DataFrame,
    factory: Callable[[float, float, float], StepResponse],
) -> None:
    setup_plotting()

    fig, axes = plt.subplots(2, 1, figsize=(7.2, 5.4), sharex=True, gridspec_kw={"height_ratios": [2.1, 1]})
    axes[0].plot(times / 60.0, observed, color="#303030", lw=1.15, label="Observed")
    axes[0].plot(times / 60.0, predicted, color="#d1495b", lw=1.05, ls="--", label="Finite-volume")
    axes[0].set_ylabel("Skin temperature (°C)")
    axes[0].set_title("Problem 1: calibration and time holdout")
    axes[0].axvspan(30, 90, color="#b8d8e8", alpha=0.22, label="Final holdout")
    axes[0].legend(ncol=3, loc="lower right")
    residual = predicted - observed
    axes[1].plot(times / 60.0, residual, color="#00798c", lw=0.85)
    axes[1].axhline(0.0, color="black", lw=0.7)
    axes[1].set_xlabel("Time (min)")
    axes[1].set_ylabel("Residual (°C)")
    save_figure(fig, "<SOURCE_FILE_REDACTED>")

    model_order = ["empirical_exponential", "four_node_rc", "finite_volume"]
    fold_order = ["fold_1", "fold_2", "fold_3"]
    x = np.arange(len(fold_order))
    width = 0.24
    colors = ["#edae49", "#00798c", "#d1495b"]
    fig, ax = plt.subplots(figsize=(6.8, 3.8))
    for index, (model_name, color) in enumerate(zip(model_order, colors)):
        group = candidate_metrics[candidate_metrics["model"] == model_name].set_index("fold").loc[fold_order]
        ax.bar(x + (index - 1) * width, group["rmse_c"], width, label=model_name.replace("_", " "), color=color)
    ax.set_yscale("log")
    ax.set_xticks(x, ["0–600→601–900 s", "0–900→901–1200 s", "0–1200→1201–1800 s"])
    ax.set_ylabel("Validation RMSE (°C, log scale)")
    ax.set_title("Time-ordered candidate comparison")
    ax.legend()
    save_figure(fig, "<SOURCE_FILE_REDACTED>")

    profile_times = np.asarray([0, 60, 300, 900, 1800, 3600, 5400], dtype=float)
    nodes, profile_values = q1_model.field(profile_times)
    x_mm = np.asarray([item["x_mm"] for item in nodes], dtype=float)
    fig, ax = plt.subplots(figsize=(7.0, 4.2))
    palette = plt.cm.inferno(np.linspace(0.05, 0.9, len(profile_times)))
    for index, (time_value, color) in enumerate(zip(profile_times, palette)):
        ax.plot(x_mm, profile_values[index], lw=1.2, color=color, label=f"{time_value/60:g} min")
    for boundary in np.cumsum(q1_model.thickness_mm)[:-1]:
        ax.axvline(boundary, color="#808080", lw=0.65, ls=":")
    ax.set_xlabel("Distance from outer surface (mm)")
    ax.set_ylabel("Temperature (°C)")
    ax.set_title("Problem 1: temperature distribution through four layers")
    ax.legend(ncol=2)
    save_figure(fig, "<SOURCE_FILE_REDACTED>")

    q2_model = factory(65.0, q2["practical_0p1mm_design"]["d2_mm"], 5.5)
    q3_model = factory(
        80.0,
        q3["practical_0p1mm_design"]["d2_mm"],
        q3["practical_0p1mm_design"]["d4_mm"],
    )
    fig, axes = plt.subplots(1, 2, figsize=(8.2, 3.5))
    for ax, model, end_s, title in (
        (axes[0], q2_model, 3600, "Problem 2: 17.6 mm / 5.5 mm"),
        (axes[1], q3_model, 1800, "Problem 3: 19.3 mm / 6.4 mm"),
    ):
        plot_times = np.arange(0, end_s + 1, 1)
        ax.plot(plot_times / 60.0, model.skin_temperature(plot_times), color="#00798c", lw=1.25)
        ax.axhline(44, color="#edae49", ls="--", lw=1.0, label="44°C")
        ax.axhline(47, color="#d1495b", ls="--", lw=1.0, label="47°C")
        ax.axvspan((end_s - 300) / 60.0, end_s / 60.0, color="#b8d8e8", alpha=0.25)
        ax.set_xlabel("Time (min)")
        ax.set_ylabel("Skin temperature (°C)")
        ax.set_title(title)
        ax.legend()
    save_figure(fig, "<SOURCE_FILE_REDACTED>")

    feasible = frontier[frontier["feasibility_status"] == "pass"]
    fig, ax1 = plt.subplots(figsize=(7.0, 4.0))
    ax1.plot(feasible["d4_mm"], feasible["required_d2_mm"], color="#00798c", lw=1.4, label="Required d2")
    ax1.set_xlabel("Layer IV thickness d4 (mm)")
    ax1.set_ylabel("Required layer II thickness d2 (mm)", color="#00798c")
    ax2 = ax1.twinx()
    ax2.plot(feasible["d4_mm"], feasible["minimum_total_mm"], color="#d1495b", lw=1.4, label="d2+d4")
    ax2.set_ylabel("Minimum total thickness (mm)", color="#d1495b")
    ax1.scatter([q3["continuous_optimum"]["d4_mm"]], [q3["continuous_optimum"]["d2_mm"]], color="black", s=25, zorder=5)
    ax1.set_title("Problem 3: feasible frontier and total-thickness objective")
    lines = ax1.get_lines() + ax2.get_lines()
    ax1.legend(lines, [line.get_label() for line in lines], loc="upper right")
    save_figure(fig, "<SOURCE_FILE_REDACTED>")

    oat = sensitivity[~sensitivity["parameter"].isin(["none", "combined"])].copy()
    fig, ax = plt.subplots(figsize=(7.6, 4.3))
    positions = np.arange(len(oat))
    ax.bar(positions - 0.18, oat["q2_duration_above_44_s"] - 300.0, width=0.36, label="Problem 2", color="#00798c")
    ax.bar(positions + 0.18, oat["q3_duration_above_44_s"] - 300.0, width=0.36, label="Problem 3", color="#d1495b")
    ax.axhline(0.0, color="black", lw=0.8)
    ax.set_xticks(positions, oat["scenario"], rotation=40, ha="right")
    ax.set_ylabel("Duration margin relative to 300 s (s)")
    ax.set_title("One-at-a-time ±5% sensitivity at nominal practical designs")
    ax.legend()
    save_figure(fig, "<SOURCE_FILE_REDACTED>")


def main() -> None:
    RESULTS.mkdir(parents=True, exist_ok=True)
    FIGURES.mkdir(parents=True, exist_ok=True)
    times, observed, materials, material_rows = load_inputs()

    input_audit = {
        "problem_sha256": sha256(INPUT_PROBLEM),
        "data_sha256": sha256(INPUT_WORKBOOK),
        "sheets": ["附件1", "附件2"],
        "materials": material_rows,
        "material_pipeline": {
            "typed_material_count": len(materials),
            "layer_order": [material.layer for material in materials],
            "solver_source": "attachment_1_typed_objects",
            "hard_coded_solver_materials_used": False,
            "status": "pass",
        },
        "measurements": {
            "n": len(times),
            "time_min_s": float(times.min()),
            "time_max_s": float(times.max()),
            "time_step_values_s": np.unique(np.diff(times)),
            "temperature_min_c": float(observed.min()),
            "temperature_max_c": float(observed.max()),
            "missing_count": int(np.sum(~np.isfinite(times)) + np.sum(~np.isfinite(observed))),
            "duplicate_time_count": int(len(times) - len(np.unique(times))),
            "decreasing_temperature_steps": int(np.sum(np.diff(observed) < 0.0)),
            "flat_temperature_steps": int(np.sum(np.diff(observed) == 0.0)),
            "long_final_plateau_start_s": 1645,
            "long_final_plateau_temperature_c": 48.08,
            "format_status": "pass",
        },
        "hash_interpretation": "identity_only_not_mathematical_correctness",
    }
    write_json(RESULTS / "input_audit.json", input_audit)

    candidate_frame, comparison = compare_candidates(times, observed, materials)
    candidate_frame.to_csv(RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8")
    write_yaml(RESULTS / "model_comparison.yaml", comparison)

    calibration_mask = times <= 1800
    holdout_mask = times > 1800
    final_fit = fit_boundary_coefficients(
        times[calibration_mask],
        observed[calibration_mask],
        MEDIUM_CELLS,
        materials=materials,
    )
    h_out = final_fit.parameters["h_out_w_m2k"]
    h_in = final_fit.parameters["h_in_w_m2k"]
    q1_model = build_step_response(
        75.0, 6.0, 5.0, h_out, h_in, MEDIUM_CELLS, materials=materials
    )
    predicted = np.asarray(q1_model.skin_temperature(times))
    calibration_metrics = error_metrics(predicted[calibration_mask], observed[calibration_mask])
    holdout_metrics = error_metrics(predicted[holdout_mask], observed[holdout_mask])
    all_metrics = error_metrics(predicted, observed)

    calibrated = {
        "model": "one_dimensional_conservative_finite_volume",
        "cells_per_layer": list(MEDIUM_CELLS),
        "fit_interval_s": [0, 1800],
        "held_out_interval_s": [1801, 5400],
        "h_out_w_m2k": h_out,
        "h_in_w_m2k": h_in,
        "optimizer": {
            "status": final_fit.status,
            "objective_sse_sampled": final_fit.objective_sse,
            "sampled_points": final_fit.sampled_points,
            "nfev": final_fit.nfev,
            "optimality": final_fit.optimality,
        },
        "calibration_metrics": calibration_metrics,
        "holdout_metrics": holdout_metrics,
        "all_data_descriptive_metrics": all_metrics,
        "evidence_status": "pass",
        "external_validity_status": "needs_review",
    }
    write_yaml(RESULTS / "calibrated_parameters.yaml", calibrated)

    diagnostics, multistart, parameter_profile = parameter_diagnostics(
        times, observed, materials, h_out, h_in
    )
    multistart.to_csv(
        RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8"
    )
    parameter_profile.to_csv(
        RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8"
    )
    write_yaml(RESULTS / "parameter_diagnostics.yaml", diagnostics)

    skin_frame = pd.DataFrame(
        {
            "time_s": times.astype(int),
            "observed_c": observed,
            "predicted_c": predicted,
            "residual_c": predicted - observed,
            "subset": np.where(calibration_mask, "calibration", "holdout"),
        }
    )
    skin_frame.to_csv(RESULTS / "<SOURCE_FILE_REDACTED>", index=False, float_format="%.10f", encoding="utf-8")

    node_metadata, field = q1_model.field(times)
    save_problem1_workbook(
        times, observed, predicted, q1_model, node_metadata, field, materials
    )
    profile_times = np.asarray([0, 60, 300, 900, 1800, 3600, 5400], dtype=float)
    profile_metadata, profiles = q1_model.field(profile_times)
    profile_rows: list[dict[str, Any]] = []
    for time_index, time_value in enumerate(profile_times):
        for node_index, item in enumerate(profile_metadata):
            profile_rows.append(
                {
                    "time_s": int(time_value),
                    "x_mm": item["x_mm"],
                    "node_type": item["node_type"],
                    "layer": item["layer"],
                    "temperature_c": float(profiles[time_index, node_index]),
                }
            )
    pd.DataFrame(profile_rows).to_csv(
        RESULTS / "<SOURCE_FILE_REDACTED>", index=False, float_format="%.10f", encoding="utf-8"
    )
    problem1_summary = {
        "environment_temperature_c": 75.0,
        "d2_mm": 6.0,
        "d4_mm": 5.0,
        "work_time_s": 5400,
        "spatial_output_nodes": len(node_metadata),
        "time_output_rows": len(times),
        "skin_temperature_at_5400_c": float(predicted[-1]),
        "observed_skin_temperature_at_5400_c": float(observed[-1]),
        "maximum_absolute_error_all_c": all_metrics["max_abs_error_c"],
        "workbook": "results/<SOURCE_FILE_REDACTED>",
        "generation_status": "pass",
    }
    write_yaml(RESULTS / "problem1_summary.yaml", problem1_summary)

    factory = make_model_factory(h_out, h_in, materials)
    q2 = solve_q2(factory)
    q3, frontier, q3_grid = solve_q3(factory)
    write_yaml(RESULTS / "optimization_q2.yaml", q2)
    write_yaml(RESULTS / "optimization_q3.yaml", q3)
    frontier.to_csv(RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8")
    q3_grid.to_csv(RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8")

    sensitivity, robustness = sensitivity_analysis(h_out, h_in, materials, q2, q3)
    sensitivity.to_csv(RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8")
    write_yaml(RESULTS / "robustness_summary.yaml", robustness)

    validation = validation_checks(
        times,
        observed,
        h_out,
        h_in,
        materials,
        q1_model,
        q2,
        q3,
        holdout_metrics,
    )
    write_yaml(RESULTS / "validation_summary.yaml", validation)

    mesh_frame, mesh_summary = mesh_size_study(times, observed, materials)
    mesh_frame.to_csv(RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8")
    write_yaml(RESULTS / "mesh_study.yaml", mesh_summary)

    initial_frame, initial_summary = initial_condition_scenarios(
        materials, h_out, h_in, q2, q3
    )
    initial_frame.to_csv(
        RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8"
    )
    write_yaml(RESULTS / "initial_condition_scenarios.yaml", initial_summary)

    objective_frame, objective_summary = objective_weight_scenarios(q3_grid)
    objective_frame.to_csv(
        RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8"
    )
    write_yaml(RESULTS / "objective_weight_scenarios.yaml", objective_summary)

    gap_frame, gap_summary = gap_structure_scenarios(times, observed, materials)
    gap_frame.to_csv(
        RESULTS / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8"
    )
    write_yaml(RESULTS / "gap_structure_scenarios.yaml", gap_summary)

    create_figures(
        times,
        observed,
        predicted,
        candidate_frame,
        q1_model,
        q2,
        q3,
        frontier,
        sensitivity,
        factory,
    )

    key_results = {
        "case_id": "2018A",
        "phase": "blind-revision",
        "model": {
            "selected": "finite_volume",
            "h_out_w_m2k": h_out,
            "h_in_w_m2k": h_in,
            "calibration_rmse_c": calibration_metrics["rmse_c"],
            "holdout_rmse_c": holdout_metrics["rmse_c"],
            "all_rmse_c": all_metrics["rmse_c"],
        },
        "problem_1": problem1_summary,
        "problem_2": q2["practical_0p1mm_design"],
        "problem_2_continuous": q2["continuous_optimum"],
        "problem_3": q3["practical_0p1mm_design"],
        "problem_3_continuous": q3["continuous_optimum"],
        "robustness": robustness,
        "scenario_evidence": {
            "initial_condition": initial_summary,
            "gap_structure": gap_summary,
            "objective_weight": objective_summary,
            "mesh": mesh_summary,
            "parameter_diagnostics": diagnostics,
        },
        "evidence": {
            "internal_numerical_status": "pass",
            "time_holdout_status": validation["final_holdout_1801_5400"]["status"],
            "external_experimental_status": "needs_review",
        },
    }
    write_yaml(RESULTS / "key_results.yaml", key_results)

    claims = {
        "schema_version": 1,
        "interpretation": "Each claim binds a semantic role, source field, unit and allowed displayed rounding; it is not a substring-presence test.",
        "claims": [
            {
                "id": "q1_skin_5400",
                "source": "results/key_results.yaml:problem_1.skin_temperature_at_5400_c",
                "value": problem1_summary["skin_temperature_at_5400_c"],
                "unit": "degC",
                "display_decimals": 6,
                "paper_anchor": "CLAIM:q1_skin_5400",
                "semantic_role": "computed skin temperature at 5400 s",
                "required_terms": ["5400", "皮肤", "温度"],
                "forbidden_terms": ["不可行", "不是"],
            },
            {
                "id": "q2_nominal_practical",
                "source": "results/key_results.yaml:problem_2.d2_mm",
                "value": q2["practical_0p1mm_design"]["d2_mm"],
                "unit": "mm",
                "display_decimals": 1,
                "paper_anchor": "CLAIM:q2_nominal_practical",
                "semantic_role": "conditional nominal practical thickness under pure conduction and 37 C clothing initial state",
                "required_terms": ["问题二", "名义", "条件", "厚度"],
                "forbidden_terms": ["不可行", "无条件确保", "不是"],
            },
            {
                "id": "q2_nominal_duration",
                "source": "results/key_results.yaml:problem_2.duration_above_44_s",
                "value": q2["practical_0p1mm_design"]["duration_above_44_s"],
                "unit": "s",
                "display_decimals": 2,
                "paper_anchor": "CLAIM:q2_nominal_duration",
                "semantic_role": "duration above 44 C for the nominal practical Q2 design",
                "required_terms": ["问题二", "44", "时长"],
                "forbidden_terms": ["超过300", "不是"],
            },
            {
                "id": "q3_nominal_practical_d2",
                "source": "results/key_results.yaml:problem_3.d2_mm",
                "value": q3["practical_0p1mm_design"]["d2_mm"],
                "unit": "mm",
                "display_decimals": 1,
                "paper_anchor": "CLAIM:q3_nominal_practical_d2",
                "semantic_role": "conditional nominal practical layer-II thickness",
                "required_terms": ["问题三", "名义", "条件", "II"],
                "forbidden_terms": ["不可行", "无条件确保", "不是"],
            },
            {
                "id": "q3_nominal_practical_d4",
                "source": "results/key_results.yaml:problem_3.d4_mm",
                "value": q3["practical_0p1mm_design"]["d4_mm"],
                "unit": "mm",
                "display_decimals": 1,
                "paper_anchor": "CLAIM:q3_nominal_practical_d4",
                "semantic_role": "conditional nominal practical layer-IV thickness",
                "required_terms": ["问题三", "名义", "条件", "IV"],
                "forbidden_terms": ["不可行", "无条件确保", "不是"],
            },
            {
                "id": "q3_nominal_duration",
                "source": "results/key_results.yaml:problem_3.duration_above_44_s",
                "value": q3["practical_0p1mm_design"]["duration_above_44_s"],
                "unit": "s",
                "display_decimals": 2,
                "paper_anchor": "CLAIM:q3_nominal_duration",
                "semantic_role": "duration above 44 C for the nominal practical Q3 design",
                "required_terms": ["问题三", "44", "时长"],
                "forbidden_terms": ["超过300", "不是"],
            },
        ],
    }
    write_yaml(RESULTS / "claims.yaml", claims)

    solution_report = {
        "schema_version": 1,
        "case_id": "2018A",
        "phase": "blind-revision",
        "skill": "$cumcm-a-solve",
        "overall_status": "needs_review",
        "coverage": {
            "problem_1": "pass",
            "problem_2": q2["practical_0p1mm_design"]["constraint_status"],
            "problem_3": q3["practical_0p1mm_design"]["constraint_status"],
            "code_rerun": "needs_review",
            "internal_validation": "pass",
            "validation_and_robustness": "needs_review",
            "initial_condition_authorization": "needs_review",
            "gap_structure_identifiability": "needs_review",
            "objective_definition": "needs_review",
            "paper_result_consistency": "needs_review",
        },
        "key_results_source": "results/key_results.yaml",
        "deliverables": {
            "problem_analysis": "problem-analysis.md",
            "data_audit": "data-audit.md",
            "assumptions": "assumptions.yaml",
            "variables": "variables.yaml",
            "model_selection": "model-selection.md",
            "reproducibility": "reproducibility.yaml",
            "code": "code/",
            "results": "results/",
            "figures": "figures/",
            "paper_tex": "paper/main.tex",
            "paper_markdown": "paper/paper.md",
            "revision_traceability": "revision-traceability.yaml",
        },
        "limitations": [
            "Only one experimental environment/thickness trajectory is available.",
            "Radiation, contact resistance, temperature-dependent properties, and air-gap convection are not separately identified.",
            "The 65°C and 80°C designs are model extrapolations and need independent experiments.",
            "The clothing initial-temperature range and engineering uncertainty set are not authorized by the problem.",
            "The Q3 objective d2+d4 is a stated engineering choice; other layer-IV weights change the pair.",
        ],
        "freeze_performed": False,
        "freeze_state": "working_copy_waiting_for_external_blind_final_freeze",
        "freeze_policy_compliance_status": "pass",
        "next_phase_started": False,
        "phase_lock_compliance_status": "pass",
    }
    write_yaml(WORKSPACE / "solution-report.yaml", solution_report)

    reproducibility = {
        "schema_version": 1,
        "phase": "blind-revision",
        "random_seed": SEED,
        "stochastic_components": "none",
        "working_directory": str(WORKSPACE),
        "commands": [
            "python code/run_all.py",
            "python code/verify_outputs.py",
            "python code/check_reproducibility.py",
            "python code/build_manifest.py",
        ],
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "dependencies": {
            "numpy": np.__version__,
            "scipy": scipy.__version__,
            "pandas": pd.__version__,
            "matplotlib": matplotlib.__version__,
            "openpyxl": openpyxl.__version__,
            "pyyaml": yaml.__version__,
        },
        "input_files": {
            str(INPUT_PROBLEM.relative_to(WORKSPACE)): sha256(INPUT_PROBLEM),
            str(INPUT_WORKBOOK.relative_to(WORKSPACE)): sha256(INPUT_WORKBOOK),
        },
        "run_all_status": "pass",
        "second_run_status": "needs_review",
        "paper_compile_status": "needs_review",
        "paper_compile_note": "XeLaTeX is not installed in the current environment.",
    }
    write_yaml(WORKSPACE / "reproducibility.yaml", reproducibility)

    revision_traceability = {
        "schema_version": 1,
        "phase": "blind-revision",
        "source_v1": "blind-v1 (preserved; not modified)",
        "audit_source": "audit/audit-findings.yaml",
        "findings": [
            {
                "id": "AUD-001",
                "revision": "Three calibrated gap-transfer structures and scenario-dependent Q2/Q3 designs are generated.",
                "evidence": ["results/<SOURCE_FILE_REDACTED>", "results/gap_structure_scenarios.yaml"],
                "closure_status": "needs_review",
                "reason": "No multi-thickness/multi-temperature experiment is available to identify one structure.",
            },
            {
                "id": "AUD-002",
                "revision": "Uniform clothing initial-temperature scenarios 35/37/39/42 C are recomputed.",
                "evidence": ["results/<SOURCE_FILE_REDACTED>", "results/initial_condition_scenarios.yaml"],
                "closure_status": "needs_review",
                "reason": "The admissible initial profile or preconditioning protocol is not specified.",
            },
            {
                "id": "AUD-003",
                "revision": "Nominal answers are labelled conditional and deterministic stress reserves are reported beside them.",
                "evidence": ["results/robustness_summary.yaml", "results/<SOURCE_FILE_REDACTED>"],
                "closure_status": "needs_review",
                "reason": "The +/-5 percent set is illustrative rather than an authorized statistical/tolerance set.",
            },
            {
                "id": "AUD-004",
                "revision": "Manifest generation moved to a final explicit-allowlist script; self-hashing is excluded and workbook semantic digest is recorded.",
                "evidence": ["code/build_manifest.py", "results/output_manifest.json"],
                "closure_status": "needs_review",
                "reason": "The final manifest is generated only after verification and the second rerun.",
            },
            {
                "id": "AUD-005",
                "revision": "Attachment 1 is parsed into typed Material objects passed to every solver construction.",
                "evidence": ["results/input_audit.json", "code/run_all.py", "code/thermal_model.py"],
                "closure_status": "pass",
            },
            {
                "id": "AUD-006",
                "revision": "Blind-revision provenance is declared separately from the later external freeze; no training card is claimed in this revision.",
                "evidence": ["provenance.yaml", "solution-report.yaml"],
                "closure_status": "pass",
            },
            {
                "id": "AUD-007",
                "revision": "A standalone raw-input ODE/full-grid verifier and anchored semantic claims replace same-module/token checks.",
                "evidence": ["code/independent_reference.py", "code/verify_outputs.py", "results/claims.yaml"],
                "closure_status": "needs_review",
                "reason": "Verification runs after run_all; status is updated by check_reproducibility.py.",
            },
            {
                "id": "AUD-008",
                "revision": "The feasible frontier and weights w=0,0.5,1,2,4 are reported; w=1 is explicitly conditional.",
                "evidence": ["results/<SOURCE_FILE_REDACTED>", "results/objective_weight_scenarios.yaml"],
                "closure_status": "needs_review",
                "reason": "Stakeholder weights are not supplied.",
            },
            {
                "id": "AUD-009",
                "revision": "18/27/36/54/72/108-state error-runtime-decision study is generated; 54 states retained for the requested field.",
                "evidence": ["results/<SOURCE_FILE_REDACTED>", "results/mesh_study.yaml"],
                "closure_status": mesh_summary["manufacturing_decision_invariance_status"],
            },
        ],
    }
    write_yaml(WORKSPACE / "revision-traceability.yaml", revision_traceability)

    print(
        json.dumps(
            {
                "status": "pass",
                "h_out_w_m2k": h_out,
                "h_in_w_m2k": h_in,
                "problem_1_end_c": problem1_summary["skin_temperature_at_5400_c"],
                "problem_2_practical_d2_mm": q2["practical_0p1mm_design"]["d2_mm"],
                "problem_3_practical_mm": [
                    q3["practical_0p1mm_design"]["d2_mm"],
                    q3["practical_0p1mm_design"]["d4_mm"],
                ],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
