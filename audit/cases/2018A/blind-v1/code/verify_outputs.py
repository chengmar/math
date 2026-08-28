"""Independent semantic checks for generated solve-stage artifacts."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
import yaml

from thermal_model import HIGH_CELLS, MEDIUM_CELLS, build_step_response


WORKSPACE = Path(__file__).resolve().parents[1]
RESULTS = WORKSPACE / "results"
PAPER_MD = WORKSPACE / "paper" / "paper.md"
PAPER_TEX = WORKSPACE / "paper" / "main.tex"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle)


def main() -> None:
    checks: list[dict[str, Any]] = []

    def add(check_id: str, status: str, evidence: Any) -> None:
        if status not in {"pass", "fail", "needs_review"}:
            raise ValueError(f"invalid status {status}")
        checks.append({"id": check_id, "status": status, "evidence": evidence})

    required = [
        WORKSPACE / "problem-analysis.md",
        WORKSPACE / "data-audit.md",
        WORKSPACE / "assumptions.yaml",
        WORKSPACE / "variables.yaml",
        WORKSPACE / "model-selection.md",
        WORKSPACE / "solution-report.yaml",
        WORKSPACE / "reproducibility.yaml",
        WORKSPACE / "code" / "thermal_model.py",
        WORKSPACE / "code" / "run_all.py",
        RESULTS / "<SOURCE_FILE_REDACTED>",
        RESULTS / "key_results.yaml",
        PAPER_MD,
        PAPER_TEX,
    ]
    missing = [str(path.relative_to(WORKSPACE)) for path in required if not path.is_file()]
    add("required_deliverables", "pass" if not missing else "fail", {"missing": missing})

    audit = json.loads((RESULTS / "input_audit.json").read_text(encoding="utf-8"))
    actual_problem_hash = sha256(WORKSPACE / "input" / "problem" / "<SOURCE_FILE_REDACTED>")
    actual_data_hash = sha256(WORKSPACE / "input" / "data" / "<SOURCE_FILE_REDACTED>")
    hashes_match = (
        actual_problem_hash == audit["problem_sha256"]
        and actual_data_hash == audit["data_sha256"]
    )
    add(
        "input_identity",
        "pass" if hashes_match else "fail",
        {
            "problem_sha256": actual_problem_hash,
            "data_sha256": actual_data_hash,
            "interpretation": "identity only, not mathematical correctness",
        },
    )

    calibrated = load_yaml(RESULTS / "calibrated_parameters.yaml")
    h_out = float(calibrated["h_out_w_m2k"])
    h_in = float(calibrated["h_in_w_m2k"])
    model = build_step_response(75.0, 6.0, 5.0, h_out, h_in, MEDIUM_CELLS)
    skin = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>")
    recomputed = np.asarray(model.skin_temperature(skin["time_s"].to_numpy(dtype=float)))
    csv_difference = float(np.max(np.abs(recomputed - skin["predicted_c"].to_numpy())))
    add(
        "problem1_csv_recomputation",
        "pass" if csv_difference <= 1e-9 else "fail",
        {"max_absolute_difference_c": csv_difference},
    )

    workbook = openpyxl.load_workbook(RESULTS / "<SOURCE_FILE_REDACTED>", data_only=True, read_only=True)
    expected_sheets = {
        "temperature_distribution",
        "spatial_nodes",
        "skin_comparison",
        "parameters",
    }
    add(
        "problem1_workbook_sheets",
        "pass" if set(workbook.sheetnames) == expected_sheets else "fail",
        {"sheets": workbook.sheetnames},
    )
    distribution = workbook["temperature_distribution"]
    row_iterator = distribution.iter_rows(values_only=True)
    header = list(next(row_iterator))
    first_data = list(next(row_iterator))
    last_data = first_data
    row_count = 2
    for values in row_iterator:
        last_data = list(values)
        row_count += 1
    column_count = len(header)
    dimensions_ok = row_count == 5402 and column_count == 60
    initial_difference = float(np.max(np.abs(np.asarray(first_data[1:], dtype=float) - 37.0)))
    workbook_skin_difference = abs(float(last_data[-1]) - float(recomputed[-1]))
    workbook_ok = (
        dimensions_ok
        and float(first_data[0]) == 0.0
        and float(last_data[0]) == 5400.0
        and initial_difference <= 1e-9
        and workbook_skin_difference <= 1e-9
    )
    add(
        "problem1_workbook_semantics",
        "pass" if workbook_ok else "fail",
        {
            "rows": row_count,
            "columns": column_count,
            "initial_max_difference_from_37_c": initial_difference,
            "last_skin_difference_from_csv_c": workbook_skin_difference,
        },
    )

    comparison = load_yaml(RESULTS / "model_comparison.yaml")
    candidate = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>")
    means = candidate.groupby("model")["rmse_c"].mean()
    best_model = str(means.idxmin())
    add(
        "candidate_selection",
        "pass"
        if best_model == comparison["selected_model"] == "finite_volume"
        else "fail",
        {"mean_rmse_c": means.to_dict(), "selected": comparison["selected_model"]},
    )

    q2 = load_yaml(RESULTS / "optimization_q2.yaml")
    q2_design = q2["practical_0p1mm_design"]
    q2_model = build_step_response(65.0, float(q2_design["d2_mm"]), 5.5, h_out, h_in, MEDIUM_CELLS)
    q2_end = float(q2_model.skin_temperature(3600.0))
    q2_duration = q2_model.duration_above_s(44.0, 3600.0)
    q2_ok = q2_end <= 47.0 + 1e-9 and q2_duration <= 300.0 + 1e-7
    add(
        "problem2_constraints_recomputed",
        "pass" if q2_ok else "fail",
        {"d2_mm": q2_design["d2_mm"], "end_temperature_c": q2_end, "duration_above_44_s": q2_duration},
    )
    q2_lower = build_step_response(65.0, float(q2_design["d2_mm"]) - 0.1, 5.5, h_out, h_in, MEDIUM_CELLS)
    q2_lower_duration = q2_lower.duration_above_s(44.0, 3600.0)
    add(
        "problem2_grid_optimality_neighbor",
        "pass" if q2_lower_duration > 300.0 else "fail",
        {"lower_d2_mm": float(q2_design["d2_mm"]) - 0.1, "duration_above_44_s": q2_lower_duration},
    )

    q3 = load_yaml(RESULTS / "optimization_q3.yaml")
    q3_design = q3["practical_0p1mm_design"]
    q3_model = build_step_response(
        80.0,
        float(q3_design["d2_mm"]),
        float(q3_design["d4_mm"]),
        h_out,
        h_in,
        MEDIUM_CELLS,
    )
    q3_end = float(q3_model.skin_temperature(1800.0))
    q3_duration = q3_model.duration_above_s(44.0, 1800.0)
    q3_ok = q3_end <= 47.0 + 1e-9 and q3_duration <= 300.0 + 1e-7
    add(
        "problem3_constraints_recomputed",
        "pass" if q3_ok else "fail",
        {
            "d2_mm": q3_design["d2_mm"],
            "d4_mm": q3_design["d4_mm"],
            "end_temperature_c": q3_end,
            "duration_above_44_s": q3_duration,
        },
    )
    q3_grid = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>")
    feasible = q3_grid[q3_grid["constraint_status"] == "pass"]
    minimum_total = float(feasible["total_mm"].min())
    add(
        "problem3_grid_optimality",
        "pass" if abs(minimum_total - float(q3_design["total_mm"])) <= 1e-9 else "fail",
        {"recomputed_grid_minimum_total_mm": minimum_total},
    )

    mesh_times = np.arange(0.0, 1801.0, 1.0)
    high = build_step_response(
        80.0,
        float(q3_design["d2_mm"]),
        float(q3_design["d4_mm"]),
        h_out,
        h_in,
        HIGH_CELLS,
    )
    mesh_difference = float(
        np.max(
            np.abs(
                np.asarray(q3_model.skin_temperature(mesh_times))
                - np.asarray(high.skin_temperature(mesh_times))
            )
        )
    )
    add(
        "independent_mesh_refinement",
        "pass" if mesh_difference <= 0.02 else "fail",
        {"q3_max_skin_difference_c": mesh_difference},
    )

    flux_out, flux_in = model.steady_fluxes_w_m2()
    flux_rel = abs(flux_out - flux_in) / abs(flux_out)
    add(
        "steady_energy_balance",
        "pass" if flux_rel <= 1e-10 else "fail",
        {"outer_flux_w_m2": flux_out, "inner_flux_w_m2": flux_in, "relative_difference": flux_rel},
    )
    add(
        "stable_linear_system",
        "pass" if float(np.max(model.eigenvalues)) < 0.0 else "fail",
        {"largest_eigenvalue_per_s": float(np.max(model.eigenvalues))},
    )

    memory_usage = (WORKSPACE / "reports" / "training-memory-usage.md").read_text(encoding="utf-8")
    add(
        "training_memory_decisions",
        "pass" if "pending" not in memory_usage else "fail",
        {"pending_present": "pending" in memory_usage},
    )

    if PAPER_MD.is_file() and PAPER_TEX.is_file():
        paper_text = PAPER_MD.read_text(encoding="utf-8") + "\n" + PAPER_TEX.read_text(encoding="utf-8")
        expected_tokens = ["119.979", "8.360", "17.6", "19.3", "6.4", "48.084889", "294.25", "292.37"]
        missing_tokens = [token for token in expected_tokens if token not in paper_text]
        add(
            "paper_key_number_presence",
            "pass" if not missing_tokens else "fail",
            {"missing_tokens": missing_tokens},
        )
    else:
        add("paper_key_number_presence", "fail", {"reason": "paper source missing"})

    add(
        "external_experimental_validity",
        "needs_review",
        {"reason": "No independent measurements at 65°C, 80°C, or alternative thicknesses."},
    )
    add(
        "mathematical_correctness_claim",
        "needs_review",
        {"reason": "Automated invariants and recomputation do not constitute a proof of model truth."},
    )

    critical_ids = {
        "required_deliverables",
        "input_identity",
        "problem1_csv_recomputation",
        "problem1_workbook_sheets",
        "problem1_workbook_semantics",
        "candidate_selection",
        "problem2_constraints_recomputed",
        "problem2_grid_optimality_neighbor",
        "problem3_constraints_recomputed",
        "problem3_grid_optimality",
        "independent_mesh_refinement",
        "steady_energy_balance",
        "stable_linear_system",
        "training_memory_decisions",
        "paper_key_number_presence",
    }
    critical_failed = [item["id"] for item in checks if item["id"] in critical_ids and item["status"] != "pass"]
    report = {
        "schema_version": 1,
        "phase": "solve",
        "internal_verification_status": "pass" if not critical_failed else "fail",
        "overall_scientific_status": "needs_review",
        "critical_failures": critical_failed,
        "checks": checks,
    }
    (RESULTS / "verification.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps({"internal_verification_status": report["internal_verification_status"], "critical_failures": critical_failed}, ensure_ascii=False))
    if critical_failed:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
