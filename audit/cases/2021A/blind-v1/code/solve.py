"""Blind, deterministic solution pipeline for CUMCM 2021 A.

The program reads only the copied inputs in ``input/`` and writes all numerical
results, the required workbook, and figures into the current solve workspace.
No network access or reference solution is used.
"""

from __future__ import annotations

import hashlib
import json
import math
import platform
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import scipy
from matplotlib.patches import Circle
from scipy.optimize import LinearConstraint, minimize, minimize_scalar
from scipy.stats import qmc


ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "input"
DATA_DIR = INPUT_DIR / "data"
RESULTS_DIR = ROOT / "results"
FIGURES_DIR = ROOT / "figures"
PAPER_DIR = ROOT / "paper"

SEED = 20_210_828
ALPHA_DEG = 36.795
BETA_DEG = 78.169
FOCAL_RATIO = 0.466
APERTURE_RADIUS_M = 150.0
RECEIVER_RADIUS_M = 0.5
STROKE_LIMIT_M = 0.6
EDGE_STRAIN_LIMIT = 0.0007
# The exact length constraint is nonlinear.  The optimization uses this tighter
# first-order cap and then validates the exact nonlinear lengths independently.
LINEAR_STRAIN_CAP = 0.00060
OUTPUT_DECIMALS = 6


@dataclass
class InputData:
    nodes_path: Path
    actuators_path: Path
    panels_path: Path
    workbook_path: Path
    nodes_frame: pd.DataFrame
    actuators_frame: pd.DataFrame
    panels_frame: pd.DataFrame
    node_ids: np.ndarray
    nodes: np.ndarray
    actuator_top: np.ndarray
    actuator_bottom: np.ndarray
    panel_ids: np.ndarray
    panel_indices: np.ndarray
    edges: np.ndarray


def ensure_directories() -> None:
    for directory in (RESULTS_DIR, FIGURES_DIR, PAPER_DIR):
        directory.mkdir(parents=True, exist_ok=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_default(value):
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, Path):
        return str(value)
    raise TypeError(f"Unsupported JSON value: {type(value)!r}")


def write_json(path: Path, payload: dict | list) -> None:
    path.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
            default=json_default,
        )
        + "\n",
        encoding="utf-8",
    )


def trinary(condition: bool) -> str:
    return "pass" if bool(condition) else "fail"


def read_csv_gb18030(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, encoding="gb18030")


def discover_inputs() -> InputData:
    csv_paths = sorted(DATA_DIR.glob("*.csv"))
    workbook_paths = sorted(DATA_DIR.glob("*.xlsx"))
    if len(csv_paths) != 3 or len(workbook_paths) != 1:
        raise RuntimeError(
            f"Expected three CSV files and one XLSX file, found "
            f"{len(csv_paths)} and {len(workbook_paths)}."
        )

    frames = {path: read_csv_gb18030(path) for path in csv_paths}
    node_candidates = [
        (path, frame)
        for path, frame in frames.items()
        if frame.shape[1] == 4 and "节点编号" in str(frame.columns[0])
    ]
    actuator_candidates = [
        (path, frame)
        for path, frame in frames.items()
        if frame.shape[1] == 7 and "主索节点编号" in str(frame.columns[0])
    ]
    panel_candidates = [
        (path, frame)
        for path, frame in frames.items()
        if frame.shape[1] == 3 and all("主索节点" in str(c) for c in frame.columns)
    ]
    if not (
        len(node_candidates) == len(actuator_candidates) == len(panel_candidates) == 1
    ):
        raise RuntimeError("Could not identify the three attachments by their schemas.")

    nodes_path, nodes_frame = node_candidates[0]
    actuators_path, actuators_frame = actuator_candidates[0]
    panels_path, panels_frame = panel_candidates[0]
    node_ids = nodes_frame.iloc[:, 0].astype(str).to_numpy()
    nodes = nodes_frame.iloc[:, 1:4].to_numpy(dtype=float)
    actuator_bottom = actuators_frame.iloc[:, 1:4].to_numpy(dtype=float)
    actuator_top = actuators_frame.iloc[:, 4:7].to_numpy(dtype=float)
    panel_ids = panels_frame.astype(str).to_numpy()

    if len(set(node_ids.tolist())) != len(node_ids):
        raise RuntimeError("Node identifiers are not unique.")
    id_to_index = {node_id: index for index, node_id in enumerate(node_ids)}
    missing_panel_ids = sorted(set(panel_ids.ravel()) - set(node_ids.tolist()))
    if missing_panel_ids:
        raise RuntimeError(f"Panel table contains unknown node IDs: {missing_panel_ids[:5]}")
    panel_indices = np.vectorize(id_to_index.get)(panel_ids).astype(int)
    edges = np.unique(
        np.sort(
            np.vstack(
                [
                    panel_indices[:, [0, 1]],
                    panel_indices[:, [1, 2]],
                    panel_indices[:, [2, 0]],
                ]
            ),
            axis=1,
        ),
        axis=0,
    )

    return InputData(
        nodes_path=nodes_path,
        actuators_path=actuators_path,
        panels_path=panels_path,
        workbook_path=workbook_paths[0],
        nodes_frame=nodes_frame,
        actuators_frame=actuators_frame,
        panels_frame=panels_frame,
        node_ids=node_ids,
        nodes=nodes,
        actuator_top=actuator_top,
        actuator_bottom=actuator_bottom,
        panel_ids=panel_ids,
        panel_indices=panel_indices,
        edges=edges,
    )


def direction_vector(alpha_deg: float, beta_deg: float) -> np.ndarray:
    alpha = math.radians(alpha_deg)
    beta = math.radians(beta_deg)
    return np.array(
        [math.cos(beta) * math.cos(alpha), math.cos(beta) * math.sin(alpha), math.sin(beta)],
        dtype=float,
    )


def transverse_basis(axis: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    # Horizontal transverse direction followed by a right-handed meridional one.
    alpha = math.atan2(axis[1], axis[0])
    e1 = np.array([-math.sin(alpha), math.cos(alpha), 0.0])
    e2 = np.cross(axis, e1)
    e1 /= np.linalg.norm(e1)
    e2 /= np.linalg.norm(e2)
    return e1, e2


def active_mask(nodes: np.ndarray, axis: np.ndarray) -> np.ndarray:
    axial = nodes @ axis
    rho_squared = np.maximum(0.0, np.einsum("ij,ij->i", nodes, nodes) - axial**2)
    return (rho_squared <= APERTURE_RADIUS_M**2) & (axial < 0.0)


def radial_intersection_radii(
    directions: np.ndarray,
    sphere_radius: float,
    nominal_focal_length: float,
    paraboloid_focal_length: float,
    axis: np.ndarray,
) -> np.ndarray:
    """Positive radial intersection with the focus-constrained paraboloid.

    The stable form 2v/(sqrt(q^2+4av)-q) avoids cancellation near q=-1.
    """

    q_axis = directions @ axis
    vertex_radius = sphere_radius - nominal_focal_length + paraboloid_focal_length
    coefficient = (1.0 - q_axis**2) / (4.0 * paraboloid_focal_length)
    discriminant_root = np.sqrt(
        np.maximum(0.0, q_axis**2 + 4.0 * coefficient * vertex_radius)
    )
    result = np.empty_like(q_axis)
    regular = coefficient > 1e-15
    result[regular] = (
        2.0 * vertex_radius / (discriminant_root[regular] - q_axis[regular])
    )
    result[~regular] = -vertex_radius / q_axis[~regular]
    return result


def continuous_minimax_paraboloid(
    sphere_radius: float, nominal_focal_length: float
) -> tuple[float, dict]:
    rho = np.linspace(0.0, APERTURE_RADIUS_M, 100_001)
    q_axis = -np.sqrt(1.0 - (rho / sphere_radius) ** 2)

    def displacement(focal_length: float) -> np.ndarray:
        vertex_radius = sphere_radius - nominal_focal_length + focal_length
        coefficient = (1.0 - q_axis**2) / (4.0 * focal_length)
        root = np.sqrt(q_axis**2 + 4.0 * coefficient * vertex_radius)
        radius = np.empty_like(rho)
        regular = coefficient > 1e-15
        radius[regular] = 2.0 * vertex_radius / (root[regular] - q_axis[regular])
        radius[~regular] = vertex_radius
        return sphere_radius - radius

    optimization = minimize_scalar(
        lambda focal_length: float(np.max(np.abs(displacement(focal_length)))),
        bounds=(nominal_focal_length - STROKE_LIMIT_M, nominal_focal_length + STROKE_LIMIT_M),
        method="bounded",
        options={"xatol": 1e-11, "maxiter": 200},
    )
    focal_length = float(optimization.x)
    disp = displacement(focal_length)
    interior_index = int(np.argmax(disp))
    area_weighted_rms = math.sqrt(
        float(np.trapezoid(disp**2 * rho, rho) / (0.5 * APERTURE_RADIUS_M**2))
    )
    details = {
        "optimizer_status": trinary(optimization.success),
        "objective_max_radial_displacement_m": float(np.max(np.abs(disp))),
        "center_displacement_m": float(disp[0]),
        "interior_max_displacement_m": float(disp[interior_index]),
        "interior_max_radius_m": float(rho[interior_index]),
        "edge_displacement_m": float(disp[-1]),
        "area_weighted_rms_displacement_m": area_weighted_rms,
        "grid_points": int(len(rho)),
    }
    return focal_length, details


def apply_radial_displacements(
    nodes: np.ndarray, directions: np.ndarray, mask: np.ndarray, active_displacements: np.ndarray
) -> np.ndarray:
    adjusted = nodes.copy()
    adjusted[mask] = nodes[mask] - active_displacements[:, None] * directions[mask]
    return adjusted


def edge_strains(
    baseline_nodes: np.ndarray, adjusted_nodes: np.ndarray, edges: np.ndarray
) -> np.ndarray:
    baseline_lengths = np.linalg.norm(
        baseline_nodes[edges[:, 0]] - baseline_nodes[edges[:, 1]], axis=1
    )
    adjusted_lengths = np.linalg.norm(
        adjusted_nodes[edges[:, 0]] - adjusted_nodes[edges[:, 1]], axis=1
    )
    return adjusted_lengths / baseline_lengths - 1.0


def affected_edges(edges: np.ndarray, mask: np.ndarray) -> np.ndarray:
    return edges[mask[edges[:, 0]] | mask[edges[:, 1]]]


def linear_strain_matrix(
    nodes: np.ndarray, directions: np.ndarray, edges: np.ndarray, mask: np.ndarray
) -> tuple[np.ndarray, np.ndarray]:
    active_indices = np.flatnonzero(mask)
    local = np.full(len(nodes), -1, dtype=int)
    local[active_indices] = np.arange(len(active_indices))
    edge_vectors = nodes[edges[:, 0]] - nodes[edges[:, 1]]
    lengths_squared = np.einsum("ij,ij->i", edge_vectors, edge_vectors)
    matrix = np.zeros((len(edges), len(active_indices)), dtype=float)
    rows = np.arange(len(edges))
    first_active = mask[edges[:, 0]]
    second_active = mask[edges[:, 1]]
    matrix[rows[first_active], local[edges[first_active, 0]]] = -np.einsum(
        "ij,ij->i", edge_vectors[first_active], directions[edges[first_active, 0]]
    ) / lengths_squared[first_active]
    matrix[rows[second_active], local[edges[second_active, 1]]] = np.einsum(
        "ij,ij->i", edge_vectors[second_active], directions[edges[second_active, 1]]
    ) / lengths_squared[second_active]
    return matrix, active_indices


def constrained_displacements(
    target: np.ndarray,
    strain_matrix: np.ndarray,
    strain_cap: float,
) -> tuple[np.ndarray, dict]:
    constraint = LinearConstraint(strain_matrix, -strain_cap, strain_cap)
    initial = np.zeros_like(target)
    optimization = minimize(
        lambda displacement: 0.5 * float(np.sum((displacement - target) ** 2)),
        initial,
        jac=lambda displacement: displacement - target,
        bounds=[(-STROKE_LIMIT_M, STROKE_LIMIT_M)] * len(target),
        constraints=[constraint],
        method="SLSQP",
        options={"ftol": 1e-12, "maxiter": 300, "disp": False},
    )
    result = np.asarray(optimization.x, dtype=float)
    details = {
        "optimizer_status": trinary(optimization.success),
        "optimizer_message": str(optimization.message),
        "iterations": int(optimization.nit),
        "objective_half_squared_radial_error": float(optimization.fun),
        "linear_constraint_min": float(np.min(strain_matrix @ result)),
        "linear_constraint_max": float(np.max(strain_matrix @ result)),
        "target_residual_rms_m": float(np.sqrt(np.mean((result - target) ** 2))),
        "target_residual_max_abs_m": float(np.max(np.abs(result - target))),
    }
    return result, details


def actuator_strokes(
    baseline_nodes: np.ndarray,
    adjusted_nodes: np.ndarray,
    top: np.ndarray,
    bottom: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    axes = top - bottom
    axes /= np.linalg.norm(axes, axis=1)[:, None]
    cable_lengths = np.linalg.norm(baseline_nodes - top, axis=1)
    vector = adjusted_nodes - top
    projection = np.einsum("ij,ij->i", vector, axes)
    constant = np.einsum("ij,ij->i", vector, vector) - cable_lengths**2
    discriminant = projection**2 - constant
    if np.min(discriminant) < -1e-9:
        raise RuntimeError(f"Negative actuator discriminant: {np.min(discriminant)}")
    square_root = np.sqrt(np.maximum(discriminant, 0.0))
    root_1 = projection - square_root
    root_2 = projection + square_root
    strokes = np.where(np.abs(root_1) <= np.abs(root_2), root_1, root_2)
    moved_top = top + strokes[:, None] * axes
    cable_residuals = np.linalg.norm(adjusted_nodes - moved_top, axis=1) - cable_lengths
    return strokes, axes, cable_lengths, cable_residuals


def exact_paraboloid_distances(
    points: np.ndarray,
    axis: np.ndarray,
    vertex: np.ndarray,
    focal_length: float,
) -> np.ndarray:
    relative = points - vertex
    axial = relative @ axis
    transverse = relative - axial[:, None] * axis
    rho = np.linalg.norm(transverse, axis=1)
    nearest_rho = rho.copy()
    for _ in range(20):
        nearest_axial = nearest_rho**2 / (4.0 * focal_length)
        gradient = (nearest_rho - rho) + (
            nearest_axial - axial
        ) * nearest_rho / (2.0 * focal_length)
        hessian = (
            1.0
            + nearest_rho**2 / (4.0 * focal_length**2)
            + (nearest_axial - axial) / (2.0 * focal_length)
        )
        step = gradient / hessian
        nearest_rho = np.maximum(0.0, nearest_rho - step)
        if float(np.max(np.abs(step))) < 1e-13:
            break
    nearest_axial = nearest_rho**2 / (4.0 * focal_length)
    distance = np.sqrt((nearest_rho - rho) ** 2 + (nearest_axial - axial) ** 2)
    implicit = axial - rho**2 / (4.0 * focal_length)
    return np.sign(implicit) * distance


def candidate_metrics(
    name: str,
    focal_length: float,
    active_displacements: np.ndarray,
    data: InputData,
    directions: np.ndarray,
    mask: np.ndarray,
    relevant_edges: np.ndarray,
    sphere_radius: float,
    nominal_focal_length: float,
) -> tuple[dict, np.ndarray, np.ndarray]:
    adjusted = apply_radial_displacements(data.nodes, directions, mask, active_displacements)
    strains = edge_strains(data.nodes, adjusted, relevant_edges)
    strokes, _, _, cable_residuals = actuator_strokes(
        data.nodes, adjusted, data.actuator_top, data.actuator_bottom
    )
    vertex_radius = sphere_radius - nominal_focal_length + focal_length
    vertex = -vertex_radius * direction_vector(ALPHA_DEG, BETA_DEG)
    surface_distances = exact_paraboloid_distances(
        adjusted[mask], direction_vector(ALPHA_DEG, BETA_DEG), vertex, focal_length
    )
    max_stroke = float(np.max(np.abs(strokes[mask])))
    max_strain = float(np.max(np.abs(strains)))
    constraint_status = trinary(
        max_stroke <= STROKE_LIMIT_M + 1e-12
        and max_strain <= EDGE_STRAIN_LIMIT + 1e-12
        and np.max(np.abs(cable_residuals[mask])) <= 1e-9
    )
    metrics = {
        "candidate": name,
        "focal_length_m": float(focal_length),
        "min_stroke_m": float(np.min(strokes[mask])),
        "max_stroke_m": float(np.max(strokes[mask])),
        "max_abs_stroke_m": max_stroke,
        "max_abs_exact_edge_strain": max_strain,
        "edge_violation_count": int(np.sum(np.abs(strains) > EDGE_STRAIN_LIMIT)),
        "surface_distance_rms_m": float(np.sqrt(np.mean(surface_distances**2))),
        "surface_distance_max_abs_m": float(np.max(np.abs(surface_distances))),
        "cable_closure_max_abs_m": float(np.max(np.abs(cable_residuals[mask]))),
        "constraint_status": constraint_status,
    }
    return metrics, adjusted, strokes


def ideal_ray_test(
    axis: np.ndarray,
    vertex: np.ndarray,
    focus: np.ndarray,
    focal_length: float,
) -> dict:
    e1, e2 = transverse_basis(axis)
    rho = np.linspace(0.0, APERTURE_RADIUS_M, 41)
    theta = np.linspace(0.0, 2.0 * np.pi, 96, endpoint=False)
    rr, tt = np.meshgrid(rho, theta, indexing="ij")
    transverse = (
        rr[..., None] * np.cos(tt)[..., None] * e1
        + rr[..., None] * np.sin(tt)[..., None] * e2
    )
    points = vertex + transverse + (rr**2 / (4.0 * focal_length))[..., None] * axis
    normals = axis - transverse / (2.0 * focal_length)
    normals /= np.linalg.norm(normals, axis=-1)[..., None]
    incoming = -axis
    reflected = incoming - 2.0 * np.sum(incoming * normals, axis=-1)[..., None] * normals
    parameter = np.sum((focus - points) * axis, axis=-1) / np.sum(
        reflected * axis, axis=-1
    )
    hits = points + parameter[..., None] * reflected
    error = hits - focus
    transverse_error = error - np.sum(error * axis, axis=-1)[..., None] * axis
    max_error = float(np.max(np.linalg.norm(transverse_error, axis=-1)))
    return {
        "sample_count": int(points.shape[0] * points.shape[1]),
        "max_focal_plane_transverse_error_m": max_error,
        "status": trinary(max_error <= 1e-9),
    }


def ray_evaluate(
    vertices: np.ndarray,
    panels: np.ndarray,
    axis: np.ndarray,
    focus: np.ndarray,
    sample_power: int,
    receiver_radii: Iterable[float],
    seed: int,
    collect_hits: bool = False,
) -> dict:
    receiver_radii = np.asarray(list(receiver_radii), dtype=float)
    unit_square = qmc.Sobol(d=2, scramble=True, seed=seed).random_base2(sample_power)
    root_u = np.sqrt(unit_square[:, 0])
    barycentric = np.column_stack(
        [
            1.0 - root_u,
            root_u * (1.0 - unit_square[:, 1]),
            root_u * unit_square[:, 1],
        ]
    )
    incoming = -axis
    denominator = 0.0
    numerators = np.zeros(len(receiver_radii), dtype=float)
    contributing_panels = 0
    e1, e2 = transverse_basis(axis)
    hit_x_parts: list[np.ndarray] = []
    hit_y_parts: list[np.ndarray] = []

    for panel in panels:
        triangle = vertices[panel]
        cross = np.cross(triangle[1] - triangle[0], triangle[2] - triangle[0])
        twice_area = float(np.linalg.norm(cross))
        if twice_area <= 0.0:
            continue
        area = 0.5 * twice_area
        normal = cross / twice_area
        projected_area = area * abs(float(incoming @ normal))
        points = barycentric @ triangle
        axial = points @ axis
        rho_squared = np.einsum("ij,ij->i", points, points) - axial**2
        inside = (rho_squared <= APERTURE_RADIUS_M**2) & (axial < 0.0)
        if not np.any(inside):
            continue
        contributing_panels += 1
        denominator += projected_area * float(np.mean(inside))

        reflected = incoming - 2.0 * float(incoming @ normal) * normal
        reflected_axial = float(reflected @ axis)
        if reflected_axial <= 1e-12:
            continue
        parameter = ((focus - points) @ axis) / reflected_axial
        hits = points + parameter[:, None] * reflected
        hit_offset = hits - focus
        hit_rho_squared = np.einsum("ij,ij->i", hit_offset, hit_offset) - (
            hit_offset @ axis
        ) ** 2
        forward = parameter > 0.0
        for index, receiver_radius in enumerate(receiver_radii):
            received = inside & forward & (hit_rho_squared <= receiver_radius**2)
            numerators[index] += projected_area * float(np.mean(received))

        if collect_hits:
            selected = inside & forward & (hit_rho_squared <= 36.0)
            if np.any(selected):
                hit_x_parts.append(hit_offset[selected] @ e1)
                hit_y_parts.append(hit_offset[selected] @ e2)

    ratios = numerators / denominator
    result = {
        "sample_power": int(sample_power),
        "samples_per_panel": int(2**sample_power),
        "seed": int(seed),
        "receiver_radii_m": receiver_radii,
        "numerator_projected_areas_m2": numerators,
        "denominator_projected_area_m2": float(denominator),
        "ratios": ratios,
        "contributing_panel_count": int(contributing_panels),
        "aperture_area_relative_error": float(
            denominator / (math.pi * APERTURE_RADIUS_M**2) - 1.0
        ),
    }
    if collect_hits:
        result["hit_x_m"] = (
            np.concatenate(hit_x_parts) if hit_x_parts else np.empty(0, dtype=float)
        )
        result["hit_y_m"] = (
            np.concatenate(hit_y_parts) if hit_y_parts else np.empty(0, dtype=float)
        )
    return result


def audit_inputs(data: InputData) -> dict:
    radii = np.linalg.norm(data.nodes, axis=1)
    top_to_bottom = data.actuator_top - data.actuator_bottom
    axes = top_to_bottom / np.linalg.norm(top_to_bottom, axis=1)[:, None]
    inward = -data.actuator_top / np.linalg.norm(data.actuator_top, axis=1)[:, None]
    triangle_cross = np.cross(
        data.nodes[data.panel_indices[:, 1]] - data.nodes[data.panel_indices[:, 0]],
        data.nodes[data.panel_indices[:, 2]] - data.nodes[data.panel_indices[:, 0]],
    )
    triangle_areas = 0.5 * np.linalg.norm(triangle_cross, axis=1)
    file_records = []
    for path in sorted(INPUT_DIR.rglob("*")):
        if path.is_file():
            file_records.append(
                {
                    "relative_path": path.relative_to(ROOT).as_posix(),
                    "bytes": path.stat().st_size,
                    "sha256": sha256_file(path),
                }
            )
    actuator_ids = data.actuators_frame.iloc[:, 0].astype(str).to_numpy()
    return {
        "status": trinary(
            len(data.nodes) == 2226
            and len(data.panels_frame) == 4300
            and len(data.edges) == 6525
            and np.array_equal(data.node_ids, actuator_ids)
            and np.all(triangle_areas > 0.0)
        ),
        "encoding": "GB18030 for the three CSV attachments",
        "files": file_records,
        "counts": {
            "nodes": int(len(data.nodes)),
            "actuators": int(len(data.actuator_top)),
            "panels": int(len(data.panel_indices)),
            "unique_edges": int(len(data.edges)),
        },
        "missing_cells": {
            "nodes": int(data.nodes_frame.isna().sum().sum()),
            "actuators": int(data.actuators_frame.isna().sum().sum()),
            "panels": int(data.panels_frame.isna().sum().sum()),
        },
        "duplicate_rows": {
            "nodes": int(data.nodes_frame.duplicated().sum()),
            "actuators": int(data.actuators_frame.duplicated().sum()),
            "panels": int(data.panels_frame.duplicated().sum()),
        },
        "node_radius_m": {
            "mean": float(np.mean(radii)),
            "std": float(np.std(radii)),
            "min": float(np.min(radii)),
            "max": float(np.max(radii)),
            "max_abs_deviation_from_median": float(np.max(np.abs(radii - np.median(radii)))),
        },
        "actuator_axis_inward_alignment": {
            "min_cosine": float(np.min(np.einsum("ij,ij->i", axes, inward))),
            "status": trinary(np.min(np.einsum("ij,ij->i", axes, inward)) > 0.999999),
        },
        "triangle_area_m2": {
            "min": float(np.min(triangle_areas)),
            "median": float(np.median(triangle_areas)),
            "max": float(np.max(triangle_areas)),
        },
        "id_alignment_status": trinary(np.array_equal(data.node_ids, actuator_ids)),
        "template_sheets": openpyxl.load_workbook(
            data.workbook_path, read_only=True
        ).sheetnames,
    }


def save_workbook(
    template_path: Path,
    vertex: np.ndarray,
    node_ids: np.ndarray,
    adjusted: np.ndarray,
    strokes: np.ndarray,
    mask: np.ndarray,
) -> None:
    workbook = openpyxl.load_workbook(template_path)
    sheet_vertex = workbook["理想抛物面顶点坐标"]
    sheet_nodes = workbook["调整后主索节点编号及坐标"]
    sheet_strokes = workbook["促动器顶端伸缩量"]
    for column, value in enumerate(vertex, start=1):
        sheet_vertex.cell(row=2, column=column, value=round(float(value), OUTPUT_DECIMALS))
    active_indices = np.flatnonzero(mask)
    for output_row, index in enumerate(active_indices, start=2):
        sheet_nodes.cell(output_row, 1, str(node_ids[index]))
        for column, value in enumerate(adjusted[index], start=2):
            sheet_nodes.cell(output_row, column, round(float(value), OUTPUT_DECIMALS))
        sheet_strokes.cell(output_row, 1, str(node_ids[index]))
        sheet_strokes.cell(output_row, 2, round(float(strokes[index]), OUTPUT_DECIMALS))
    workbook.save(ROOT / "<SOURCE_FILE_REDACTED>")


def save_figures(
    data: InputData,
    sphere_radius: float,
    nominal_focal_length: float,
    ideal_focal_length: float,
    axis: np.ndarray,
    vertex: np.ndarray,
    mask: np.ndarray,
    adjusted: np.ndarray,
    strokes: np.ndarray,
    surface_distances: np.ndarray,
    strains: np.ndarray,
    convergence: pd.DataFrame,
    adjusted_hits: dict,
    baseline_hits: dict,
    model_comparison: pd.DataFrame,
) -> None:
    plt.rcParams.update({"font.size": 10, "axes.grid": True, "grid.alpha": 0.25})

    rho = np.linspace(0.0, APERTURE_RADIUS_M, 2001)
    sphere_axial = -np.sqrt(sphere_radius**2 - rho**2)
    vertex_radius = sphere_radius - nominal_focal_length + ideal_focal_length
    parabola_axial = -vertex_radius + rho**2 / (4.0 * ideal_focal_length)
    q_axis = -np.sqrt(1.0 - (rho / sphere_radius) ** 2)
    coefficient = (1.0 - q_axis**2) / (4.0 * ideal_focal_length)
    root = np.sqrt(q_axis**2 + 4.0 * coefficient * vertex_radius)
    radial_radius = np.empty_like(rho)
    regular = coefficient > 1e-15
    radial_radius[regular] = 2.0 * vertex_radius / (root[regular] - q_axis[regular])
    radial_radius[~regular] = vertex_radius
    radial_displacement = sphere_radius - radial_radius

    fig, axes_plot = plt.subplots(1, 2, figsize=(11, 4.2))
    axes_plot[0].plot(rho, sphere_axial, label="Reference sphere")
    axes_plot[0].plot(rho, parabola_axial, label="Ideal paraboloid")
    axes_plot[0].set_xlabel("Transverse radius (m)")
    axes_plot[0].set_ylabel("Axial coordinate (m)")
    axes_plot[0].legend()
    axes_plot[0].set_title("Meridional profiles")
    axes_plot[1].plot(rho, radial_displacement)
    axes_plot[1].axhline(STROKE_LIMIT_M, color="r", linestyle="--", linewidth=0.8)
    axes_plot[1].axhline(-STROKE_LIMIT_M, color="r", linestyle="--", linewidth=0.8)
    axes_plot[1].set_xlabel("Transverse radius (m)")
    axes_plot[1].set_ylabel("Radial displacement (m)")
    axes_plot[1].set_title("Minimax ideal displacement")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)

    e1, e2 = transverse_basis(axis)
    local_x = data.nodes[mask] @ e1
    local_y = data.nodes[mask] @ e2
    fig, axes_plot = plt.subplots(1, 2, figsize=(11, 4.7))
    scatter = axes_plot[0].scatter(
        local_x, local_y, c=strokes[mask], s=13, cmap="coolwarm", rasterized=True
    )
    fig.colorbar(scatter, ax=axes_plot[0], label="Actuator stroke (m)")
    axes_plot[0].set_title("Actuator field")
    scatter = axes_plot[1].scatter(
        local_x, local_y, c=surface_distances, s=13, cmap="PiYG", rasterized=True
    )
    fig.colorbar(scatter, ax=axes_plot[1], label="Signed surface error (m)")
    axes_plot[1].set_title("Deviation from ideal paraboloid")
    for axis_plot in axes_plot:
        axis_plot.set_aspect("equal")
        axis_plot.set_xlabel("Local transverse x (m)")
        axis_plot.set_ylabel("Local transverse y (m)")
        axis_plot.add_patch(Circle((0, 0), APERTURE_RADIUS_M, fill=False, color="k", linewidth=0.7))
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)

    fig, axes_plot = plt.subplots(1, 2, figsize=(11, 4.2))
    axes_plot[0].hist(100.0 * strains, bins=45, color="#4472C4", alpha=0.85)
    axes_plot[0].axvline(100.0 * EDGE_STRAIN_LIMIT, color="r", linestyle="--")
    axes_plot[0].axvline(-100.0 * EDGE_STRAIN_LIMIT, color="r", linestyle="--")
    axes_plot[0].set_xlabel("Exact edge-length change (%)")
    axes_plot[0].set_ylabel("Edge count")
    axes_plot[0].set_title("Affected-edge validation")
    axes_plot[1].hist(strokes[mask], bins=38, color="#70AD47", alpha=0.85)
    axes_plot[1].axvline(STROKE_LIMIT_M, color="r", linestyle="--")
    axes_plot[1].axvline(-STROKE_LIMIT_M, color="r", linestyle="--")
    axes_plot[1].set_xlabel("Actuator stroke (m)")
    axes_plot[1].set_ylabel("Node count")
    axes_plot[1].set_title("Stroke validation")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)

    fig, axis_plot = plt.subplots(figsize=(7.2, 4.4))
    axis_plot.plot(
        convergence["samples_per_panel"],
        100.0 * convergence["adjusted_ratio"],
        marker="o",
        label="Adjusted surface",
    )
    axis_plot.plot(
        convergence["samples_per_panel"],
        100.0 * convergence["baseline_ratio"],
        marker="s",
        label="Reference sphere",
    )
    axis_plot.set_xscale("log", base=2)
    axis_plot.set_xlabel("Sobol samples per panel")
    axis_plot.set_ylabel("Reception ratio (%)")
    axis_plot.set_title("Independent ray-integration convergence")
    axis_plot.legend()
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)

    fig, axes_plot = plt.subplots(1, 2, figsize=(10.5, 4.5), sharex=True, sharey=True)
    for axis_plot, hits, title in [
        (axes_plot[0], adjusted_hits, "Adjusted surface"),
        (axes_plot[1], baseline_hits, "Reference sphere"),
    ]:
        if len(hits["hit_x_m"]):
            axis_plot.hexbin(
                hits["hit_x_m"],
                hits["hit_y_m"],
                gridsize=75,
                extent=(-6, 6, -6, 6),
                bins="log",
                mincnt=1,
                cmap="viridis",
            )
        axis_plot.add_patch(Circle((0, 0), RECEIVER_RADIUS_M, fill=False, color="r", linewidth=1.2))
        axis_plot.set_title(title)
        axis_plot.set_aspect("equal")
        axis_plot.set_xlabel("Focal-plane x (m)")
    axes_plot[0].set_ylabel("Focal-plane y (m)")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)

    fig, axes_plot = plt.subplots(1, 2, figsize=(10.5, 4.2))
    labels = ["Fixed-F\nradial", "Minimax\nradial", "Constrained\nselected"]
    axes_plot[0].bar(labels, model_comparison["max_abs_stroke_m"], color=["#A5A5A5", "#5B9BD5", "#70AD47"])
    axes_plot[0].axhline(STROKE_LIMIT_M, color="r", linestyle="--")
    axes_plot[0].set_ylabel("Maximum actuator stroke (m)")
    axes_plot[0].set_title("Stroke comparison")
    axes_plot[1].bar(labels, 100.0 * model_comparison["max_abs_exact_edge_strain"], color=["#A5A5A5", "#5B9BD5", "#70AD47"])
    axes_plot[1].axhline(100.0 * EDGE_STRAIN_LIMIT, color="r", linestyle="--")
    axes_plot[1].set_ylabel("Maximum edge-length change (%)")
    axes_plot[1].set_title("Constraint comparison")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)


def main() -> int:
    np.random.seed(SEED)
    ensure_directories()
    data = discover_inputs()
    audit = audit_inputs(data)

    radii = np.linalg.norm(data.nodes, axis=1)
    sphere_radius = float(np.median(radii))
    nominal_focal_length = FOCAL_RATIO * sphere_radius
    axis = direction_vector(ALPHA_DEG, BETA_DEG)
    directions = data.nodes / radii[:, None]
    mask = active_mask(data.nodes, axis)
    relevant_edges = affected_edges(data.edges, mask)
    strain_matrix, active_indices = linear_strain_matrix(
        data.nodes, directions, relevant_edges, mask
    )
    if not np.array_equal(active_indices, np.flatnonzero(mask)):
        raise RuntimeError("Active-node indexing mismatch.")

    ideal_focal_length, minimax_details = continuous_minimax_paraboloid(
        sphere_radius, nominal_focal_length
    )
    focus_radius = sphere_radius - nominal_focal_length
    vertex_radius = focus_radius + ideal_focal_length
    focus = -focus_radius * axis
    vertex = -vertex_radius * axis

    fixed_radii = radial_intersection_radii(
        directions, sphere_radius, nominal_focal_length, nominal_focal_length, axis
    )
    fixed_target = (radii - fixed_radii)[mask]
    ideal_radii = radial_intersection_radii(
        directions, sphere_radius, nominal_focal_length, ideal_focal_length, axis
    )
    ideal_target = (radii - ideal_radii)[mask]
    selected_displacements, solver_details = constrained_displacements(
        ideal_target, strain_matrix, LINEAR_STRAIN_CAP
    )

    candidate_rows = []
    fixed_metrics, _, _ = candidate_metrics(
        "C0_fixed_nominal_focal_radial_baseline",
        nominal_focal_length,
        fixed_target,
        data,
        directions,
        mask,
        relevant_edges,
        sphere_radius,
        nominal_focal_length,
    )
    candidate_rows.append(fixed_metrics)
    minimax_metrics, _, _ = candidate_metrics(
        "C1_minimax_radial_intersection",
        ideal_focal_length,
        ideal_target,
        data,
        directions,
        mask,
        relevant_edges,
        sphere_radius,
        nominal_focal_length,
    )
    candidate_rows.append(minimax_metrics)
    selected_metrics, adjusted, strokes = candidate_metrics(
        "C2_strain_constrained_radial_least_squares",
        ideal_focal_length,
        selected_displacements,
        data,
        directions,
        mask,
        relevant_edges,
        sphere_radius,
        nominal_focal_length,
    )
    candidate_rows.append(selected_metrics)
    model_comparison = pd.DataFrame(candidate_rows)
    model_comparison.to_csv(RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig")

    strokes, actuator_axes, cable_lengths, cable_residuals = actuator_strokes(
        data.nodes, adjusted, data.actuator_top, data.actuator_bottom
    )
    strains = edge_strains(data.nodes, adjusted, relevant_edges)
    surface_distances = exact_paraboloid_distances(
        adjusted[mask], axis, vertex, ideal_focal_length
    )
    rounded_adjusted = np.round(adjusted, OUTPUT_DECIMALS)
    rounded_strokes = np.round(strokes, OUTPUT_DECIMALS)
    rounded_strains = edge_strains(data.nodes, rounded_adjusted, relevant_edges)
    rounded_top = data.actuator_top + rounded_strokes[:, None] * actuator_axes
    rounded_cable_residuals = (
        np.linalg.norm(rounded_adjusted - rounded_top, axis=1) - cable_lengths
    )

    ideal_check = ideal_ray_test(axis, vertex, focus, ideal_focal_length)

    convergence_rows = []
    final_adjusted_ray = None
    final_baseline_ray = None
    for sample_power in (10, 12, 14, 16):
        adjusted_ray = ray_evaluate(
            adjusted,
            data.panel_indices,
            axis,
            focus,
            sample_power,
            [RECEIVER_RADIUS_M],
            SEED,
        )
        baseline_ray = ray_evaluate(
            data.nodes,
            data.panel_indices,
            axis,
            focus,
            sample_power,
            [RECEIVER_RADIUS_M],
            SEED,
        )
        convergence_rows.append(
            {
                "sample_power": sample_power,
                "samples_per_panel": 2**sample_power,
                "adjusted_ratio": float(adjusted_ray["ratios"][0]),
                "baseline_ratio": float(baseline_ray["ratios"][0]),
                "adjusted_denominator_area_m2": adjusted_ray[
                    "denominator_projected_area_m2"
                ],
                "baseline_denominator_area_m2": baseline_ray[
                    "denominator_projected_area_m2"
                ],
            }
        )
        final_adjusted_ray = adjusted_ray
        final_baseline_ray = baseline_ray
    convergence = pd.DataFrame(convergence_rows)
    convergence.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    receiver_sensitivity_adjusted = ray_evaluate(
        adjusted,
        data.panel_indices,
        axis,
        focus,
        14,
        [0.45, 0.50, 0.55],
        SEED + 101,
    )
    receiver_sensitivity_baseline = ray_evaluate(
        data.nodes,
        data.panel_indices,
        axis,
        focus,
        14,
        [0.45, 0.50, 0.55],
        SEED + 101,
    )
    receiver_sensitivity = pd.DataFrame(
        {
            "receiver_radius_m": [0.45, 0.50, 0.55],
            "adjusted_ratio": receiver_sensitivity_adjusted["ratios"],
            "baseline_ratio": receiver_sensitivity_baseline["ratios"],
        }
    )
    receiver_sensitivity["relative_improvement"] = (
        receiver_sensitivity["adjusted_ratio"]
        / receiver_sensitivity["baseline_ratio"]
        - 1.0
    )
    receiver_sensitivity.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    seed_rows = []
    for ray_seed in (SEED + 201, SEED + 202, SEED + 203):
        adjusted_seed = ray_evaluate(
            adjusted,
            data.panel_indices,
            axis,
            focus,
            14,
            [RECEIVER_RADIUS_M],
            ray_seed,
        )
        baseline_seed = ray_evaluate(
            data.nodes,
            data.panel_indices,
            axis,
            focus,
            14,
            [RECEIVER_RADIUS_M],
            ray_seed,
        )
        seed_rows.append(
            {
                "seed": ray_seed,
                "adjusted_ratio": float(adjusted_seed["ratios"][0]),
                "baseline_ratio": float(baseline_seed["ratios"][0]),
            }
        )
    seed_sensitivity = pd.DataFrame(seed_rows)
    seed_sensitivity.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    strain_sensitivity_rows = []
    for cap in (0.00058, 0.00060, 0.00062):
        cap_displacements, cap_solver = constrained_displacements(
            ideal_target, strain_matrix, cap
        )
        cap_adjusted = apply_radial_displacements(
            data.nodes, directions, mask, cap_displacements
        )
        cap_strains = edge_strains(data.nodes, cap_adjusted, relevant_edges)
        cap_surface_distances = exact_paraboloid_distances(
            cap_adjusted[mask], axis, vertex, ideal_focal_length
        )
        cap_strokes, _, _, _ = actuator_strokes(
            data.nodes, cap_adjusted, data.actuator_top, data.actuator_bottom
        )
        strain_sensitivity_rows.append(
            {
                "linear_strain_cap": cap,
                "optimizer_status": cap_solver["optimizer_status"],
                "exact_max_abs_edge_strain": float(np.max(np.abs(cap_strains))),
                "surface_distance_rms_m": float(
                    np.sqrt(np.mean(cap_surface_distances**2))
                ),
                "surface_distance_max_abs_m": float(
                    np.max(np.abs(cap_surface_distances))
                ),
                "max_abs_actuator_stroke_m": float(np.max(np.abs(cap_strokes[mask]))),
                "constraint_status": trinary(
                    np.max(np.abs(cap_strains)) <= EDGE_STRAIN_LIMIT
                    and np.max(np.abs(cap_strokes[mask])) <= STROKE_LIMIT_M
                ),
            }
        )
    strain_sensitivity = pd.DataFrame(strain_sensitivity_rows)
    strain_sensitivity.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    adjusted_hits = ray_evaluate(
        adjusted,
        data.panel_indices,
        axis,
        focus,
        9,
        [RECEIVER_RADIUS_M],
        SEED + 301,
        collect_hits=True,
    )
    baseline_hits = ray_evaluate(
        data.nodes,
        data.panel_indices,
        axis,
        focus,
        9,
        [RECEIVER_RADIUS_M],
        SEED + 301,
        collect_hits=True,
    )

    adjusted_ratio = float(final_adjusted_ray["ratios"][0])
    baseline_ratio = float(final_baseline_ray["ratios"][0])
    relative_improvement = adjusted_ratio / baseline_ratio - 1.0
    convergence_adjusted = float(
        abs(convergence.iloc[-1]["adjusted_ratio"] - convergence.iloc[-2]["adjusted_ratio"])
    )
    convergence_baseline = float(
        abs(convergence.iloc[-1]["baseline_ratio"] - convergence.iloc[-2]["baseline_ratio"])
    )

    validation = {
        "active_node_count": int(np.sum(mask)),
        "affected_edge_count": int(len(relevant_edges)),
        "solver": solver_details,
        "stroke": {
            "min_m": float(np.min(strokes[mask])),
            "max_m": float(np.max(strokes[mask])),
            "max_abs_m": float(np.max(np.abs(strokes[mask]))),
            "limit_m": STROKE_LIMIT_M,
            "status": trinary(np.max(np.abs(strokes[mask])) <= STROKE_LIMIT_M),
        },
        "edge_strain": {
            "min": float(np.min(strains)),
            "max": float(np.max(strains)),
            "max_abs": float(np.max(np.abs(strains))),
            "limit": EDGE_STRAIN_LIMIT,
            "violation_count": int(np.sum(np.abs(strains) > EDGE_STRAIN_LIMIT)),
            "status": trinary(np.max(np.abs(strains)) <= EDGE_STRAIN_LIMIT),
        },
        "surface_distance": {
            "mean_signed_m": float(np.mean(surface_distances)),
            "rms_m": float(np.sqrt(np.mean(surface_distances**2))),
            "max_abs_m": float(np.max(np.abs(surface_distances))),
        },
        "cable_closure": {
            "max_abs_m": float(np.max(np.abs(cable_residuals[mask]))),
            "status": trinary(np.max(np.abs(cable_residuals[mask])) <= 1e-9),
        },
        "rounded_output": {
            "decimals": OUTPUT_DECIMALS,
            "max_abs_edge_strain": float(np.max(np.abs(rounded_strains))),
            "max_abs_cable_residual_m": float(
                np.max(np.abs(rounded_cable_residuals[mask]))
            ),
            "status": trinary(
                np.max(np.abs(rounded_strains)) <= EDGE_STRAIN_LIMIT
                and np.max(np.abs(rounded_strokes[mask])) <= STROKE_LIMIT_M
                and np.max(np.abs(rounded_cable_residuals[mask])) <= 2e-6
            ),
        },
        "ideal_ray_identity": ideal_check,
        "ray_integration": {
            "final_samples_per_panel": int(2**16),
            "adjusted_ratio": adjusted_ratio,
            "baseline_ratio": baseline_ratio,
            "relative_improvement": relative_improvement,
            "adjusted_last_level_absolute_change": convergence_adjusted,
            "baseline_last_level_absolute_change": convergence_baseline,
            "adjusted_aperture_area_relative_error": final_adjusted_ray[
                "aperture_area_relative_error"
            ],
            "baseline_aperture_area_relative_error": final_baseline_ray[
                "aperture_area_relative_error"
            ],
            "numerical_status": trinary(
                convergence_adjusted <= 1e-5
                and convergence_baseline <= 1e-5
                and abs(final_adjusted_ray["aperture_area_relative_error"]) <= 1e-4
                and abs(final_baseline_ray["aperture_area_relative_error"]) <= 1e-4
            ),
            "physical_external_validity_status": "needs_review",
        },
    }

    ideal_payload = {
        "sphere_radius_m": sphere_radius,
        "nominal_focal_ratio": FOCAL_RATIO,
        "nominal_focal_length_m": nominal_focal_length,
        "ideal_focal_length_m": ideal_focal_length,
        "ideal_focal_ratio": ideal_focal_length / sphere_radius,
        "focus_radius_m": focus_radius,
        "vertex_radius_m": vertex_radius,
        "question_1_vertical": {
            "axis": [0.0, 0.0, 1.0],
            "focus_m": [0.0, 0.0, -focus_radius],
            "vertex_m": [0.0, 0.0, -vertex_radius],
            "equation": (
                f"z = {-vertex_radius:.12f} + (x^2+y^2)/({4*ideal_focal_length:.12f})"
            ),
        },
        "question_2": {
            "alpha_deg": ALPHA_DEG,
            "beta_deg": BETA_DEG,
            "axis": axis,
            "focus_m": focus,
            "vertex_m": vertex,
            "vector_equation": "u·(x-V)=||(I-uu^T)(x-V)||^2/(4f)",
        },
        "minimax": minimax_details,
    }
    write_json(RESULTS_DIR / "input_audit.json", audit)
    write_json(RESULTS_DIR / "ideal_paraboloid.json", ideal_payload)
    write_json(RESULTS_DIR / "validation_metrics.json", validation)

    active_indices = np.flatnonzero(mask)
    adjusted_frame = pd.DataFrame(
        {
            "node_id": data.node_ids[active_indices],
            "x_m": rounded_adjusted[active_indices, 0],
            "y_m": rounded_adjusted[active_indices, 1],
            "z_m": rounded_adjusted[active_indices, 2],
            "radial_displacement_m": np.round(selected_displacements, OUTPUT_DECIMALS),
            "actuator_stroke_m": rounded_strokes[active_indices],
            "signed_surface_distance_m": np.round(surface_distances, OUTPUT_DECIMALS),
        }
    )
    adjusted_frame.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )
    pd.DataFrame(
        {
            "node_id": data.node_ids[active_indices],
            "actuator_stroke_m": rounded_strokes[active_indices],
        }
    ).to_csv(RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig")

    evidence_matrix = pd.DataFrame(
        [
            {
                "layer": "input_identity",
                "status": audit["status"],
                "evidence": "results/input_audit.json with actual SHA-256 values",
            },
            {
                "layer": "optimizer_execution",
                "status": solver_details["optimizer_status"],
                "evidence": "SLSQP exit status and deterministic objective",
            },
            {
                "layer": "internal_geometry_constraints",
                "status": trinary(
                    validation["stroke"]["status"] == "pass"
                    and validation["edge_strain"]["status"] == "pass"
                    and validation["cable_closure"]["status"] == "pass"
                ),
                "evidence": "exact nonlinear post-checks in validation_metrics.json",
            },
            {
                "layer": "numerical_ray_integration",
                "status": validation["ray_integration"]["numerical_status"],
                "evidence": "nested Sobol refinement and ideal-paraboloid identity",
            },
            {
                "layer": "physical_external_validity",
                "status": "needs_review",
                "evidence": "piecewise-planar, geometric-optics assumptions lack external calibration",
            },
        ]
    )
    evidence_matrix.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    summary = {
        "case_id": "2021A",
        "phase": "solve",
        "seed": SEED,
        "question_1": {
            "ideal_focal_length_m": ideal_focal_length,
            "ideal_focal_ratio": ideal_focal_length / sphere_radius,
            "vertical_vertex_m": [0.0, 0.0, -vertex_radius],
            "max_ideal_radial_displacement_m": minimax_details[
                "objective_max_radial_displacement_m"
            ],
        },
        "question_2": {
            "vertex_m": vertex,
            "active_node_count": int(np.sum(mask)),
            "stroke_range_m": [
                float(np.min(strokes[mask])),
                float(np.max(strokes[mask])),
            ],
            "max_abs_edge_strain": float(np.max(np.abs(strains))),
            "surface_distance_rms_m": float(np.sqrt(np.mean(surface_distances**2))),
            "surface_distance_max_abs_m": float(np.max(np.abs(surface_distances))),
            "internal_constraint_status": selected_metrics["constraint_status"],
        },
        "question_3": {
            "adjusted_reception_ratio": adjusted_ratio,
            "baseline_sphere_reception_ratio": baseline_ratio,
            "relative_improvement": relative_improvement,
            "numerical_status": validation["ray_integration"]["numerical_status"],
            "physical_external_validity_status": "needs_review",
        },
    }
    write_json(RESULTS_DIR / "summary.json", summary)

    save_workbook(
        data.workbook_path,
        vertex,
        data.node_ids,
        rounded_adjusted,
        rounded_strokes,
        mask,
    )
    save_figures(
        data,
        sphere_radius,
        nominal_focal_length,
        ideal_focal_length,
        axis,
        vertex,
        mask,
        adjusted,
        strokes,
        surface_distances,
        strains,
        convergence,
        adjusted_hits,
        baseline_hits,
        model_comparison,
    )

    generated_tex = "\n".join(
        [
            "% Generated by code/solve.py; do not edit numerical values by hand.",
            f"\\newcommand{{\\SphereRadius}}{{{sphere_radius:.4f}}}",
            f"\\newcommand{{\\NominalFocalLength}}{{{nominal_focal_length:.4f}}}",
            f"\\newcommand{{\\IdealFocalLength}}{{{ideal_focal_length:.6f}}}",
            f"\\newcommand{{\\IdealFocalRatio}}{{{ideal_focal_length/sphere_radius:.6f}}}",
            f"\\newcommand{{\\IdealMaxStroke}}{{{minimax_details['objective_max_radial_displacement_m']:.6f}}}",
            f"\\newcommand{{\\VertexX}}{{{vertex[0]:.6f}}}",
            f"\\newcommand{{\\VertexY}}{{{vertex[1]:.6f}}}",
            f"\\newcommand{{\\VertexZ}}{{{vertex[2]:.6f}}}",
            f"\\newcommand{{\\ActiveNodeCount}}{{{int(np.sum(mask))}}}",
            f"\\newcommand{{\\StrokeMin}}{{{np.min(strokes[mask]):.6f}}}",
            f"\\newcommand{{\\StrokeMax}}{{{np.max(strokes[mask]):.6f}}}",
            f"\\newcommand{{\\EdgeStrainPct}}{{{100*np.max(np.abs(strains)):.6f}}}",
            f"\\newcommand{{\\SurfaceRMS}}{{{np.sqrt(np.mean(surface_distances**2)):.6f}}}",
            f"\\newcommand{{\\SurfaceMax}}{{{np.max(np.abs(surface_distances)):.6f}}}",
            f"\\newcommand{{\\AdjustedReceptionPct}}{{{100*adjusted_ratio:.3f}}}",
            f"\\newcommand{{\\BaselineReceptionPct}}{{{100*baseline_ratio:.3f}}}",
            f"\\newcommand{{\\RelativeImprovementPct}}{{{100*relative_improvement:.1f}}}",
        ]
    )
    (PAPER_DIR / "generated_numbers.tex").write_text(generated_tex + "\n", encoding="utf-8")

    environment = {
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "numpy": np.__version__,
        "pandas": pd.__version__,
        "scipy": scipy.__version__,
        "matplotlib": matplotlib.__version__,
        "openpyxl": openpyxl.__version__,
        "seed": SEED,
        "command": "python code/solve.py",
    }
    write_json(RESULTS_DIR / "environment.json", environment)

    print(json.dumps(summary, ensure_ascii=False, indent=2, default=json_default))
    internal_ok = (
        audit["status"] == "pass"
        and selected_metrics["constraint_status"] == "pass"
        and validation["rounded_output"]["status"] == "pass"
        and validation["ideal_ray_identity"]["status"] == "pass"
        and validation["ray_integration"]["numerical_status"] == "pass"
    )
    print(f"INTERNAL_STATUS={trinary(internal_ok)}")
    return 0 if internal_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
