"""Rerun and consistency checks for the blind-solve deliverables."""

from __future__ import annotations

import hashlib
import json
import subprocess
import sys
import time
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "results"
PAPER = ROOT / "paper"
CHECK_PATH = RESULTS / "reproduction_check.json"

NUMERIC_FILES = [
    "results/input_audit.json",
    "results/ideal_paraboloid.json",
    "results/summary.json",
    "results/validation_metrics.json",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "paper/generated_numbers.tex",
]

REQUIRED_PATHS = [
    "problem-analysis.md",
    "data-audit.md",
    "assumptions.yaml",
    "variables.yaml",
    "model-selection.md",
    "solution-report.yaml",
    "reproducibility.yaml",
    "code/solve.py",
    "code/verify.py",
    "code/requirements.txt",
    "results/summary.json",
    "results/validation_metrics.json",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "paper/main.tex",
    "paper/paper.md",
    "<SOURCE_FILE_REDACTED>",
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def hashes() -> dict[str, str]:
    return {relative: sha256(ROOT / relative) for relative in NUMERIC_FILES}


def run_solver() -> dict:
    start = time.perf_counter()
    completed = subprocess.run(
        [sys.executable, str(ROOT / "code" / "solve.py")],
        cwd=ROOT,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=600,
        check=False,
    )
    elapsed = time.perf_counter() - start
    return {
        "return_code": completed.returncode,
        "elapsed_seconds": elapsed,
        "status": "pass" if completed.returncode == 0 else "fail",
        "last_stdout_line": completed.stdout.strip().splitlines()[-1]
        if completed.stdout.strip()
        else "",
        "stderr_tail": completed.stderr[-1000:],
    }


def workbook_checks(summary: dict) -> dict:
    workbook = openpyxl.load_workbook(ROOT / "<SOURCE_FILE_REDACTED>", read_only=True, data_only=True)
    vertex_sheet = workbook["理想抛物面顶点坐标"]
    node_sheet = workbook["调整后主索节点编号及坐标"]
    stroke_sheet = workbook["促动器顶端伸缩量"]
    vertex = np.array([vertex_sheet.cell(2, column).value for column in range(1, 4)], dtype=float)
    expected_vertex = np.array(summary["question_2"]["vertex_m"], dtype=float)

    node_csv = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>", encoding="utf-8-sig")
    stroke_csv = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>", encoding="utf-8-sig")
    workbook_node_ids = [
        str(node_sheet.cell(row, 1).value) for row in range(2, node_sheet.max_row + 1)
    ]
    workbook_nodes = np.array(
        [
            [node_sheet.cell(row, column).value for column in range(2, 5)]
            for row in range(2, node_sheet.max_row + 1)
        ],
        dtype=float,
    )
    workbook_stroke_ids = [
        str(stroke_sheet.cell(row, 1).value) for row in range(2, stroke_sheet.max_row + 1)
    ]
    workbook_strokes = np.array(
        [stroke_sheet.cell(row, 2).value for row in range(2, stroke_sheet.max_row + 1)],
        dtype=float,
    )
    node_values = node_csv[["x_m", "y_m", "z_m"]].to_numpy(float)
    stroke_values = stroke_csv["actuator_stroke_m"].to_numpy(float)
    checks = {
        "vertex_max_abs_difference_m": float(np.max(np.abs(vertex - expected_vertex))),
        "node_row_count": int(len(workbook_node_ids)),
        "expected_node_row_count": int(summary["question_2"]["active_node_count"]),
        "node_id_match": workbook_node_ids == node_csv["node_id"].astype(str).tolist(),
        "node_coordinate_max_abs_difference_m": float(
            np.max(np.abs(workbook_nodes - node_values))
        ),
        "stroke_id_match": workbook_stroke_ids
        == stroke_csv["node_id"].astype(str).tolist(),
        "stroke_max_abs_difference_m": float(
            np.max(np.abs(workbook_strokes - stroke_values))
        ),
    }
    checks["status"] = (
        "pass"
        if checks["vertex_max_abs_difference_m"] <= 5.1e-7
        and checks["node_row_count"] == checks["expected_node_row_count"]
        and checks["node_id_match"]
        and checks["node_coordinate_max_abs_difference_m"] <= 1e-12
        and checks["stroke_id_match"]
        and checks["stroke_max_abs_difference_m"] <= 1e-12
        else "fail"
    )
    return checks


def paper_checks(summary: dict) -> dict:
    markdown = (PAPER / "paper.md").read_text(encoding="utf-8")
    tex = (PAPER / "main.tex").read_text(encoding="utf-8")
    generated = (PAPER / "generated_numbers.tex").read_text(encoding="utf-8")
    expected_markdown_claims = [
        f"{summary['question_1']['ideal_focal_length_m']:.6f}",
        f"{summary['question_2']['vertex_m'][0]:.6f}",
        f"{summary['question_2']['vertex_m'][1]:.6f}",
        f"{summary['question_2']['vertex_m'][2]:.6f}",
        str(summary["question_2"]["active_node_count"]),
        f"{100*summary['question_2']['max_abs_edge_strain']:.6f}%",
        f"{100*summary['question_3']['adjusted_reception_ratio']:.3f}%",
        f"{100*summary['question_3']['baseline_sphere_reception_ratio']:.3f}%",
        f"{100*summary['question_3']['relative_improvement']:.1f}%",
    ]
    missing_markdown = [claim for claim in expected_markdown_claims if claim not in markdown]
    macro_claims = [
        f"{summary['question_1']['ideal_focal_length_m']:.6f}",
        f"{summary['question_2']['vertex_m'][0]:.6f}",
        f"{100*summary['question_3']['adjusted_reception_ratio']:.3f}",
        f"{100*summary['question_3']['baseline_sphere_reception_ratio']:.3f}",
        f"{100*summary['question_3']['relative_improvement']:.1f}",
    ]
    missing_macros = [claim for claim in macro_claims if claim not in generated]
    uses_generated = "\\input{generated_numbers}" in tex
    return {
        "markdown_missing_claims": missing_markdown,
        "tex_missing_generated_claims": missing_macros,
        "tex_uses_generated_numbers": uses_generated,
        "status": "pass"
        if not missing_markdown and not missing_macros and uses_generated
        else "fail",
    }


def main() -> int:
    first_run = run_solver()
    first_hashes = hashes() if first_run["status"] == "pass" else {}
    second_run = run_solver()
    second_hashes = hashes() if second_run["status"] == "pass" else {}
    hash_equal = bool(first_hashes) and first_hashes == second_hashes

    summary = json.loads((RESULTS / "summary.json").read_text(encoding="utf-8"))
    validation = json.loads(
        (RESULTS / "validation_metrics.json").read_text(encoding="utf-8")
    )
    workbook = workbook_checks(summary)
    paper = paper_checks(summary)
    missing_paths = [relative for relative in REQUIRED_PATHS if not (ROOT / relative).exists()]
    required_status = "pass" if not missing_paths else "fail"
    internal_status = (
        "pass"
        if validation["stroke"]["status"] == "pass"
        and validation["edge_strain"]["status"] == "pass"
        and validation["cable_closure"]["status"] == "pass"
        and validation["rounded_output"]["status"] == "pass"
        and validation["ideal_ray_identity"]["status"] == "pass"
        and validation["ray_integration"]["numerical_status"] == "pass"
        else "fail"
    )
    overall = (
        "pass"
        if first_run["status"] == "pass"
        and second_run["status"] == "pass"
        and hash_equal
        and workbook["status"] == "pass"
        and paper["status"] == "pass"
        and required_status == "pass"
        and internal_status == "pass"
        else "fail"
    )
    payload = {
        "overall_status": overall,
        "runs": [first_run, second_run],
        "numeric_hash_reproduction": {
            "status": "pass" if hash_equal else "fail",
            "first_run_sha256": first_hashes,
            "second_run_sha256": second_hashes,
        },
        "workbook_consistency": workbook,
        "paper_result_consistency": paper,
        "required_deliverables": {
            "status": required_status,
            "missing_paths": missing_paths,
        },
        "internal_constraint_and_numerical_status": internal_status,
        "mathematical_correctness_status": "needs_review",
        "physical_external_validity_status": "needs_review",
        "notes": [
            "Binary XLSX hashes are not compared because ZIP member timestamps may vary; cell values are compared instead.",
            "A deterministic rerun establishes execution reproducibility, not a formal proof of mathematical correctness or external physical validity.",
        ],
    }
    CHECK_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if overall == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
