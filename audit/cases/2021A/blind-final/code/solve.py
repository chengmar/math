"""Blind, deterministic solution pipeline for CUMCM 2021 A.

The program reads only the copied inputs in ``input/`` and writes all numerical
results, the required workbook, and figures into the current solve workspace.
No network access or reference solution is used.
"""

from __future__ import annotations

import hashlib
import json
import math
import platform
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import scipy
from matplotlib.colors import LogNorm
from matplotlib.patches import Circle
from scipy.optimize import LinearConstraint, minimize, minimize_scalar, nnls
from scipy.stats import qmc


ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "input"
DATA_DIR = INPUT_DIR / "data"
RESULTS_DIR = ROOT / "results"
FIGURES_DIR = ROOT / "figures"
PAPER_DIR = ROOT / "paper"

SEED = 20_210_828
ALPHA_DEG = 36.795
BETA_DEG = 78.169
FOCAL_RATIO = 0.466
APERTURE_RADIUS_M = 150.0
RECEIVER_RADIUS_M = 0.5
STROKE_LIMIT_M = 0.6
EDGE_STRAIN_LIMIT = 0.0007
# The convex inner problem uses a linearized strain cap.  The revised value is
# the largest 0.001%-grid cap that leaves at least 4e-6 exact/rounded strain
# reserve near the joint focal optimum.  Exact nonlinear and six-decimal
# constraints remain independent acceptance tests.
LINEAR_STRAIN_CAP = 0.00062
DESIGN_EDGE_RESERVE = 4.0e-6
JOINT_FOCAL_OFFSET_BOUNDS_M = (-STROKE_LIMIT_M, STROKE_LIMIT_M)
OUTPUT_DECIMALS = 6


@dataclass
class InputData:
    nodes_path: Path
    actuators_path: Path
    panels_path: Path
    workbook_path: Path
    nodes_frame: pd.DataFrame
    actuators_frame: pd.DataFrame
    panels_frame: pd.DataFrame
    node_ids: np.ndarray
    nodes: np.ndarray
    actuator_top: np.ndarray
    actuator_bottom: np.ndarray
    panel_ids: np.ndarray
    panel_indices: np.ndarray
    edges: np.ndarray


def ensure_directories() -> None:
    for directory in (RESULTS_DIR, FIGURES_DIR, PAPER_DIR):
        directory.mkdir(parents=True, exist_ok=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_default(value):
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, Path):
        return str(value)
    raise TypeError(f"Unsupported JSON value: {type(value)!r}")


def write_json(path: Path, payload: dict | list) -> None:
    path.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
            default=json_default,
        )
        + "\n",
        encoding="utf-8",
    )


def trinary(condition: bool) -> str:
    return "pass" if bool(condition) else "fail"


def read_csv_gb18030(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, encoding="gb18030")


def discover_inputs() -> InputData:
    csv_paths = sorted(DATA_DIR.glob("*.csv"))
    workbook_paths = sorted(DATA_DIR.glob("*.xlsx"))
    if len(csv_paths) != 3 or len(workbook_paths) != 1:
        raise RuntimeError(
            f"Expected three CSV files and one XLSX file, found "
            f"{len(csv_paths)} and {len(workbook_paths)}."
        )

    frames = {path: read_csv_gb18030(path) for path in csv_paths}
    node_candidates = [
        (path, frame)
        for path, frame in frames.items()
        if frame.shape[1] == 4 and "节点编号" in str(frame.columns[0])
    ]
    actuator_candidates = [
        (path, frame)
        for path, frame in frames.items()
        if frame.shape[1] == 7 and "主索节点编号" in str(frame.columns[0])
    ]
    panel_candidates = [
        (path, frame)
        for path, frame in frames.items()
        if frame.shape[1] == 3 and all("主索节点" in str(c) for c in frame.columns)
    ]
    if not (
        len(node_candidates) == len(actuator_candidates) == len(panel_candidates) == 1
    ):
        raise RuntimeError("Could not identify the three attachments by their schemas.")

    nodes_path, nodes_frame = node_candidates[0]
    actuators_path, actuators_frame = actuator_candidates[0]
    panels_path, panels_frame = panel_candidates[0]
    node_ids = nodes_frame.iloc[:, 0].astype(str).to_numpy()
    nodes = nodes_frame.iloc[:, 1:4].to_numpy(dtype=float)
    actuator_bottom = actuators_frame.iloc[:, 1:4].to_numpy(dtype=float)
    actuator_top = actuators_frame.iloc[:, 4:7].to_numpy(dtype=float)
    panel_ids = panels_frame.astype(str).to_numpy()

    if len(set(node_ids.tolist())) != len(node_ids):
        raise RuntimeError("Node identifiers are not unique.")
    id_to_index = {node_id: index for index, node_id in enumerate(node_ids)}
    missing_panel_ids = sorted(set(panel_ids.ravel()) - set(node_ids.tolist()))
    if missing_panel_ids:
        raise RuntimeError(f"Panel table contains unknown node IDs: {missing_panel_ids[:5]}")
    panel_indices = np.vectorize(id_to_index.get)(panel_ids).astype(int)
    edges = np.unique(
        np.sort(
            np.vstack(
                [
                    panel_indices[:, [0, 1]],
                    panel_indices[:, [1, 2]],
                    panel_indices[:, [2, 0]],
                ]
            ),
            axis=1,
        ),
        axis=0,
    )

    return InputData(
        nodes_path=nodes_path,
        actuators_path=actuators_path,
        panels_path=panels_path,
        workbook_path=workbook_paths[0],
        nodes_frame=nodes_frame,
        actuators_frame=actuators_frame,
        panels_frame=panels_frame,
        node_ids=node_ids,
        nodes=nodes,
        actuator_top=actuator_top,
        actuator_bottom=actuator_bottom,
        panel_ids=panel_ids,
        panel_indices=panel_indices,
        edges=edges,
    )


def direction_vector(alpha_deg: float, beta_deg: float) -> np.ndarray:
    alpha = math.radians(alpha_deg)
    beta = math.radians(beta_deg)
    return np.array(
        [math.cos(beta) * math.cos(alpha), math.cos(beta) * math.sin(alpha), math.sin(beta)],
        dtype=float,
    )


def transverse_basis(axis: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    # Horizontal transverse direction followed by a right-handed meridional one.
    alpha = math.atan2(axis[1], axis[0])
    e1 = np.array([-math.sin(alpha), math.cos(alpha), 0.0])
    e2 = np.cross(axis, e1)
    e1 /= np.linalg.norm(e1)
    e2 /= np.linalg.norm(e2)
    return e1, e2


def active_mask(nodes: np.ndarray, axis: np.ndarray) -> np.ndarray:
    axial = nodes @ axis
    rho_squared = np.maximum(0.0, np.einsum("ij,ij->i", nodes, nodes) - axial**2)
    return (rho_squared <= APERTURE_RADIUS_M**2) & (axial < 0.0)


def radial_intersection_radii(
    directions: np.ndarray,
    sphere_radius: float,
    nominal_focal_length: float,
    paraboloid_focal_length: float,
    axis: np.ndarray,
) -> np.ndarray:
    """Positive radial intersection with the focus-constrained paraboloid.

    The stable form 2v/(sqrt(q^2+4av)-q) avoids cancellation near q=-1.
    """

    q_axis = directions @ axis
    vertex_radius = sphere_radius - nominal_focal_length + paraboloid_focal_length
    coefficient = (1.0 - q_axis**2) / (4.0 * paraboloid_focal_length)
    discriminant_root = np.sqrt(
        np.maximum(0.0, q_axis**2 + 4.0 * coefficient * vertex_radius)
    )
    result = np.empty_like(q_axis)
    regular = coefficient > 1e-15
    result[regular] = (
        2.0 * vertex_radius / (discriminant_root[regular] - q_axis[regular])
    )
    result[~regular] = -vertex_radius / q_axis[~regular]
    return result


def continuous_minimax_paraboloid(
    sphere_radius: float, nominal_focal_length: float
) -> tuple[float, dict]:
    rho = np.linspace(0.0, APERTURE_RADIUS_M, 100_001)
    q_axis = -np.sqrt(1.0 - (rho / sphere_radius) ** 2)

    def displacement(focal_length: float) -> np.ndarray:
        vertex_radius = sphere_radius - nominal_focal_length + focal_length
        coefficient = (1.0 - q_axis**2) / (4.0 * focal_length)
        root = np.sqrt(q_axis**2 + 4.0 * coefficient * vertex_radius)
        radius = np.empty_like(rho)
        regular = coefficient > 1e-15
        radius[regular] = 2.0 * vertex_radius / (root[regular] - q_axis[regular])
        radius[~regular] = vertex_radius
        return sphere_radius - radius

    optimization = minimize_scalar(
        lambda focal_length: float(np.max(np.abs(displacement(focal_length)))),
        bounds=(nominal_focal_length - STROKE_LIMIT_M, nominal_focal_length + STROKE_LIMIT_M),
        method="bounded",
        options={"xatol": 1e-11, "maxiter": 200},
    )
    focal_length = float(optimization.x)
    disp = displacement(focal_length)
    interior_index = int(np.argmax(disp))
    area_weighted_rms = math.sqrt(
        float(np.trapezoid(disp**2 * rho, rho) / (0.5 * APERTURE_RADIUS_M**2))
    )
    details = {
        "optimizer_status": trinary(optimization.success),
        "objective_max_radial_displacement_m": float(np.max(np.abs(disp))),
        "center_displacement_m": float(disp[0]),
        "interior_max_displacement_m": float(disp[interior_index]),
        "interior_max_radius_m": float(rho[interior_index]),
        "edge_displacement_m": float(disp[-1]),
        "area_weighted_rms_displacement_m": area_weighted_rms,
        "grid_points": int(len(rho)),
    }
    return focal_length, details


def apply_radial_displacements(
    nodes: np.ndarray, directions: np.ndarray, mask: np.ndarray, active_displacements: np.ndarray
) -> np.ndarray:
    adjusted = nodes.copy()
    adjusted[mask] = nodes[mask] - active_displacements[:, None] * directions[mask]
    return adjusted


def edge_strains(
    baseline_nodes: np.ndarray, adjusted_nodes: np.ndarray, edges: np.ndarray
) -> np.ndarray:
    baseline_lengths = np.linalg.norm(
        baseline_nodes[edges[:, 0]] - baseline_nodes[edges[:, 1]], axis=1
    )
    adjusted_lengths = np.linalg.norm(
        adjusted_nodes[edges[:, 0]] - adjusted_nodes[edges[:, 1]], axis=1
    )
    return adjusted_lengths / baseline_lengths - 1.0


def affected_edges(edges: np.ndarray, mask: np.ndarray) -> np.ndarray:
    return edges[mask[edges[:, 0]] | mask[edges[:, 1]]]


def linear_strain_matrix(
    nodes: np.ndarray, directions: np.ndarray, edges: np.ndarray, mask: np.ndarray
) -> tuple[np.ndarray, np.ndarray]:
    active_indices = np.flatnonzero(mask)
    local = np.full(len(nodes), -1, dtype=int)
    local[active_indices] = np.arange(len(active_indices))
    edge_vectors = nodes[edges[:, 0]] - nodes[edges[:, 1]]
    lengths_squared = np.einsum("ij,ij->i", edge_vectors, edge_vectors)
    matrix = np.zeros((len(edges), len(active_indices)), dtype=float)
    rows = np.arange(len(edges))
    first_active = mask[edges[:, 0]]
    second_active = mask[edges[:, 1]]
    matrix[rows[first_active], local[edges[first_active, 0]]] = -np.einsum(
        "ij,ij->i", edge_vectors[first_active], directions[edges[first_active, 0]]
    ) / lengths_squared[first_active]
    matrix[rows[second_active], local[edges[second_active, 1]]] = np.einsum(
        "ij,ij->i", edge_vectors[second_active], directions[edges[second_active, 1]]
    ) / lengths_squared[second_active]
    return matrix, active_indices


def constrained_displacements(
    target: np.ndarray,
    strain_matrix: np.ndarray,
    strain_cap: float,
    initial: np.ndarray | None = None,
) -> tuple[np.ndarray, dict]:
    constraint = LinearConstraint(strain_matrix, -strain_cap, strain_cap)
    if initial is None:
        initial = np.zeros_like(target)
    else:
        initial = np.asarray(initial, dtype=float).copy()
        linear_values = strain_matrix @ initial
        maximum = float(np.max(np.abs(linear_values)))
        if maximum > strain_cap:
            initial *= strain_cap / maximum
    optimization = minimize(
        lambda displacement: 0.5 * float(np.sum((displacement - target) ** 2)),
        initial,
        jac=lambda displacement: displacement - target,
        bounds=[(-STROKE_LIMIT_M, STROKE_LIMIT_M)] * len(target),
        constraints=[constraint],
        method="SLSQP",
        options={"ftol": 1e-12, "maxiter": 300, "disp": False},
    )
    result = np.asarray(optimization.x, dtype=float)
    linear_values = strain_matrix @ result
    active_upper = np.flatnonzero(linear_values >= strain_cap - 2e-8)
    active_lower = np.flatnonzero(linear_values <= -strain_cap + 2e-8)
    active_normals = np.vstack(
        [strain_matrix[active_upper], -strain_matrix[active_lower]]
    )
    gradient = result - target
    if len(active_normals):
        multipliers, _ = nnls(active_normals.T, -gradient, maxiter=20_000)
        stationarity = gradient + active_normals.T @ multipliers
        dual_minimum = float(np.min(multipliers))
    else:
        stationarity = gradient
        dual_minimum = 0.0
    details = {
        "optimizer_status": trinary(optimization.success),
        "optimizer_message": str(optimization.message),
        "iterations": int(optimization.nit),
        "objective_half_squared_radial_error": float(optimization.fun),
        "radial_residual_rms_m": float(np.sqrt(np.mean((result - target) ** 2))),
        "radial_residual_max_abs_m": float(np.max(np.abs(result - target))),
        "linear_constraint_min": float(np.min(linear_values)),
        "linear_constraint_max": float(np.max(linear_values)),
        "active_linear_constraint_count": int(len(active_normals)),
        "kkt_stationarity_l2": float(np.linalg.norm(stationarity)),
        "kkt_stationarity_linf": float(np.max(np.abs(stationarity))),
        "kkt_dual_minimum": dual_minimum,
        "kkt_status": trinary(
            optimization.success
            and np.max(np.abs(stationarity)) <= 1e-7
            and dual_minimum >= -1e-12
        ),
        "target_residual_rms_m": float(np.sqrt(np.mean((result - target) ** 2))),
        "target_residual_max_abs_m": float(np.max(np.abs(result - target))),
    }
    return result, details


def rounded_constraint_metrics(
    adjusted: np.ndarray,
    strokes: np.ndarray,
    data: InputData,
    relevant_edges: np.ndarray,
) -> dict:
    rounded_adjusted = np.round(adjusted, OUTPUT_DECIMALS)
    rounded_strokes = np.round(strokes, OUTPUT_DECIMALS)
    rounded_strains = edge_strains(data.nodes, rounded_adjusted, relevant_edges)
    actuator_axes = data.actuator_top - data.actuator_bottom
    actuator_axes /= np.linalg.norm(actuator_axes, axis=1)[:, None]
    cable_lengths = np.linalg.norm(data.nodes - data.actuator_top, axis=1)
    rounded_top = data.actuator_top + rounded_strokes[:, None] * actuator_axes
    rounded_closure = (
        np.linalg.norm(rounded_adjusted - rounded_top, axis=1) - cable_lengths
    )
    return {
        "adjusted": rounded_adjusted,
        "strokes": rounded_strokes,
        "strains": rounded_strains,
        "closure": rounded_closure,
        "max_abs_edge_strain": float(np.max(np.abs(rounded_strains))),
        "max_abs_stroke_m": float(np.max(np.abs(rounded_strokes))),
    }


def joint_focal_search(
    data: InputData,
    radii: np.ndarray,
    sphere_radius: float,
    nominal_focal_length: float,
    axis: np.ndarray,
    directions: np.ndarray,
    mask: np.ndarray,
    relevant_edges: np.ndarray,
    strain_matrix: np.ndarray,
) -> tuple[float, np.ndarray, dict, pd.DataFrame]:
    """Jointly select focal length and the feasible nodal projection.

    The outer scalar variable is focal length.  For each value the inner strict
    convex QP minimizes nodal radial squared error.  Exact geometry and rounded
    output constraints gate every accepted outer evaluation.
    """

    focus_radius = sphere_radius - nominal_focal_length
    cache: dict[float, tuple[np.ndarray, dict]] = {}
    evaluation_rows: list[dict] = []

    def evaluate(focal_length: float, method: str) -> tuple[float, np.ndarray, dict]:
        key = round(float(focal_length), 12)
        if key in cache:
            displacement, payload = cache[key]
            return payload["objective_radial_rms_m"], displacement, payload
        target_radii = radial_intersection_radii(
            directions,
            sphere_radius,
            nominal_focal_length,
            float(focal_length),
            axis,
        )
        target = (radii - target_radii)[mask]
        initial = None
        if cache:
            nearest_key = min(cache, key=lambda value: abs(value - key))
            initial = cache[nearest_key][0]
        displacement, solver = constrained_displacements(
            target, strain_matrix, LINEAR_STRAIN_CAP, initial=initial
        )
        adjusted = apply_radial_displacements(
            data.nodes, directions, mask, displacement
        )
        strains = edge_strains(data.nodes, adjusted, relevant_edges)
        strokes, _, _, cable_residuals = actuator_strokes(
            data.nodes, adjusted, data.actuator_top, data.actuator_bottom
        )
        rounded = rounded_constraint_metrics(adjusted, strokes, data, relevant_edges)
        vertex = -(focus_radius + focal_length) * axis
        surface_distances = exact_paraboloid_distances(
            adjusted[mask], axis, vertex, float(focal_length)
        )
        exact_max_strain = float(np.max(np.abs(strains)))
        exact_max_stroke = float(np.max(np.abs(strokes[mask])))
        exact_closure = float(np.max(np.abs(cable_residuals[mask])))
        exact_status = trinary(
            exact_max_strain <= EDGE_STRAIN_LIMIT
            and exact_max_stroke <= STROKE_LIMIT_M
            and exact_closure <= 1e-9
        )
        rounded_status = trinary(
            rounded["max_abs_edge_strain"] <= EDGE_STRAIN_LIMIT
            and float(np.max(np.abs(rounded["strokes"][mask]))) <= STROKE_LIMIT_M
            and float(np.max(np.abs(rounded["closure"][mask]))) <= 2e-6
        )
        reserve_status = trinary(
            max(exact_max_strain, rounded["max_abs_edge_strain"])
            <= EDGE_STRAIN_LIMIT - DESIGN_EDGE_RESERVE
        )
        radial_rms = float(np.sqrt(np.mean((displacement - target) ** 2)))
        accepted = (
            solver["optimizer_status"] == "pass"
            and solver["kkt_status"] == "pass"
            and exact_status == "pass"
            and rounded_status == "pass"
            and reserve_status == "pass"
        )
        objective = radial_rms if accepted else 1_000.0 + radial_rms
        payload = {
            "method": method,
            "focal_length_m": float(focal_length),
            "focal_offset_from_nominal_m": float(focal_length - nominal_focal_length),
            "linear_strain_cap": LINEAR_STRAIN_CAP,
            "objective_radial_rms_m": radial_rms,
            "orthogonal_surface_rms_m": float(
                np.sqrt(np.mean(surface_distances**2))
            ),
            "orthogonal_surface_max_abs_m": float(
                np.max(np.abs(surface_distances))
            ),
            "exact_max_abs_edge_strain": exact_max_strain,
            "rounded_max_abs_edge_strain": rounded["max_abs_edge_strain"],
            "exact_max_abs_stroke_m": exact_max_stroke,
            "exact_cable_closure_max_abs_m": exact_closure,
            "exact_constraint_status": exact_status,
            "rounded_constraint_status": rounded_status,
            "design_reserve_status": reserve_status,
            "accepted_status": trinary(accepted),
            "solver": solver,
            "target": target,
            "adjusted": adjusted,
            "strokes": strokes,
        }
        cache[key] = (displacement, payload)
        evaluation_rows.append(
            {k: v for k, v in payload.items() if k not in {"solver", "target", "adjusted", "strokes"}}
        )
        return objective, displacement, payload

    lower = nominal_focal_length + JOINT_FOCAL_OFFSET_BOUNDS_M[0]
    upper = nominal_focal_length + JOINT_FOCAL_OFFSET_BOUNDS_M[1]
    grid = np.linspace(lower, upper, 13)
    grid_values = [evaluate(value, "coarse_grid")[0] for value in grid]
    best_index = int(np.argmin(grid_values))
    bracket_lower = grid[max(0, best_index - 1)]
    bracket_upper = grid[min(len(grid) - 1, best_index + 1)]
    refinement = minimize_scalar(
        lambda value: evaluate(float(value), "bounded_refinement")[0],
        bounds=(float(bracket_lower), float(bracket_upper)),
        method="bounded",
        options={"xatol": 2e-6, "maxiter": 30},
    )
    final_objective, final_displacement, final_payload = evaluate(
        float(refinement.x), "selected"
    )
    known_objective, _, known_payload = evaluate(
        nominal_focal_length + 0.4, "audit_counterexample_anchor"
    )
    envelope_best = min(row["objective_radial_rms_m"] for row in evaluation_rows if row["accepted_status"] == "pass")
    search_details = {
        "outer_optimizer_status": trinary(refinement.success),
        "outer_optimizer_message": str(refinement.message),
        "outer_function_evaluations": int(refinement.nfev),
        "coarse_grid_points": int(len(grid)),
        "focal_bounds_m": [float(lower), float(upper)],
        "selected_objective_radial_rms_m": float(final_objective),
        "evaluated_envelope_best_radial_rms_m": float(envelope_best),
        "audit_anchor_radial_rms_m": float(known_objective),
        "audit_anchor_accepted_status": known_payload["accepted_status"],
        "known_anchor_not_better_status": trinary(final_objective <= known_objective + 1e-9),
        "status": trinary(
            refinement.success
            and final_payload["accepted_status"] == "pass"
            and final_objective <= known_objective + 1e-9
        ),
    }
    final_payload["joint_search"] = search_details
    search_frame = pd.DataFrame(evaluation_rows).sort_values(
        ["focal_length_m", "method"], kind="stable"
    )
    return float(refinement.x), final_displacement, final_payload, search_frame


def maximum_feasible_uniform_scale(
    target: np.ndarray,
    data: InputData,
    directions: np.ndarray,
    mask: np.ndarray,
    relevant_edges: np.ndarray,
) -> tuple[float, np.ndarray, dict]:
    """One-parameter feasible baseline d=lambda*d* with rounded checks."""

    def check(scale: float) -> tuple[bool, dict]:
        displacement = scale * target
        adjusted = apply_radial_displacements(data.nodes, directions, mask, displacement)
        strains = edge_strains(data.nodes, adjusted, relevant_edges)
        strokes, _, _, cable_residuals = actuator_strokes(
            data.nodes, adjusted, data.actuator_top, data.actuator_bottom
        )
        rounded = rounded_constraint_metrics(adjusted, strokes, data, relevant_edges)
        feasible = (
            np.max(np.abs(strains)) <= EDGE_STRAIN_LIMIT
            and np.max(np.abs(strokes[mask])) <= STROKE_LIMIT_M
            and np.max(np.abs(cable_residuals[mask])) <= 1e-9
            and rounded["max_abs_edge_strain"] <= EDGE_STRAIN_LIMIT
            and np.max(np.abs(rounded["strokes"][mask])) <= STROKE_LIMIT_M
            and np.max(np.abs(rounded["closure"][mask])) <= 2e-6
        )
        return bool(feasible), {
            "adjusted": adjusted,
            "strokes": strokes,
            "strains": strains,
            "rounded": rounded,
        }

    lower, upper = 0.0, 1.0
    for _ in range(50):
        middle = 0.5 * (lower + upper)
        feasible, _ = check(middle)
        if feasible:
            lower = middle
        else:
            upper = middle
    scale = max(0.0, lower - 1e-10)
    feasible, details = check(scale)
    details["status"] = trinary(feasible)
    return scale, scale * target, details


def actuator_strokes(
    baseline_nodes: np.ndarray,
    adjusted_nodes: np.ndarray,
    top: np.ndarray,
    bottom: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    axes = top - bottom
    axes /= np.linalg.norm(axes, axis=1)[:, None]
    cable_lengths = np.linalg.norm(baseline_nodes - top, axis=1)
    vector = adjusted_nodes - top
    projection = np.einsum("ij,ij->i", vector, axes)
    constant = np.einsum("ij,ij->i", vector, vector) - cable_lengths**2
    discriminant = projection**2 - constant
    if np.min(discriminant) < -1e-9:
        raise RuntimeError(f"Negative actuator discriminant: {np.min(discriminant)}")
    square_root = np.sqrt(np.maximum(discriminant, 0.0))
    root_1 = projection - square_root
    root_2 = projection + square_root
    strokes = np.where(np.abs(root_1) <= np.abs(root_2), root_1, root_2)
    moved_top = top + strokes[:, None] * axes
    cable_residuals = np.linalg.norm(adjusted_nodes - moved_top, axis=1) - cable_lengths
    return strokes, axes, cable_lengths, cable_residuals


def exact_paraboloid_distances(
    points: np.ndarray,
    axis: np.ndarray,
    vertex: np.ndarray,
    focal_length: float,
) -> np.ndarray:
    relative = points - vertex
    axial = relative @ axis
    transverse = relative - axial[:, None] * axis
    rho = np.linalg.norm(transverse, axis=1)
    nearest_rho = rho.copy()
    for _ in range(20):
        nearest_axial = nearest_rho**2 / (4.0 * focal_length)
        gradient = (nearest_rho - rho) + (
            nearest_axial - axial
        ) * nearest_rho / (2.0 * focal_length)
        hessian = (
            1.0
            + nearest_rho**2 / (4.0 * focal_length**2)
            + (nearest_axial - axial) / (2.0 * focal_length)
        )
        step = gradient / hessian
        nearest_rho = np.maximum(0.0, nearest_rho - step)
        if float(np.max(np.abs(step))) < 1e-13:
            break
    nearest_axial = nearest_rho**2 / (4.0 * focal_length)
    distance = np.sqrt((nearest_rho - rho) ** 2 + (nearest_axial - axial) ** 2)
    implicit = axial - rho**2 / (4.0 * focal_length)
    return np.sign(implicit) * distance


def candidate_metrics(
    name: str,
    focal_length: float,
    active_displacements: np.ndarray,
    data: InputData,
    directions: np.ndarray,
    mask: np.ndarray,
    relevant_edges: np.ndarray,
    sphere_radius: float,
    nominal_focal_length: float,
    target: np.ndarray | None = None,
    parameter_count: int | None = None,
) -> tuple[dict, np.ndarray, np.ndarray]:
    adjusted = apply_radial_displacements(data.nodes, directions, mask, active_displacements)
    strains = edge_strains(data.nodes, adjusted, relevant_edges)
    strokes, _, _, cable_residuals = actuator_strokes(
        data.nodes, adjusted, data.actuator_top, data.actuator_bottom
    )
    vertex_radius = sphere_radius - nominal_focal_length + focal_length
    vertex = -vertex_radius * direction_vector(ALPHA_DEG, BETA_DEG)
    surface_distances = exact_paraboloid_distances(
        adjusted[mask], direction_vector(ALPHA_DEG, BETA_DEG), vertex, focal_length
    )
    max_stroke = float(np.max(np.abs(strokes[mask])))
    max_strain = float(np.max(np.abs(strains)))
    constraint_status = trinary(
        max_stroke <= STROKE_LIMIT_M + 1e-12
        and max_strain <= EDGE_STRAIN_LIMIT + 1e-12
        and np.max(np.abs(cable_residuals[mask])) <= 1e-9
    )
    metrics = {
        "candidate": name,
        "parameter_count": int(parameter_count if parameter_count is not None else len(active_displacements)),
        "focal_length_m": float(focal_length),
        "min_stroke_m": float(np.min(strokes[mask])),
        "max_stroke_m": float(np.max(strokes[mask])),
        "max_abs_stroke_m": max_stroke,
        "max_abs_exact_edge_strain": max_strain,
        "edge_violation_count": int(np.sum(np.abs(strains) > EDGE_STRAIN_LIMIT)),
        "surface_distance_rms_m": float(np.sqrt(np.mean(surface_distances**2))),
        "surface_distance_max_abs_m": float(np.max(np.abs(surface_distances))),
        "radial_residual_rms_m": float(
            np.sqrt(np.mean((active_displacements - target) ** 2))
        )
        if target is not None
        else 0.0,
        "radial_residual_max_abs_m": float(
            np.max(np.abs(active_displacements - target))
        )
        if target is not None
        else 0.0,
        "cable_closure_max_abs_m": float(np.max(np.abs(cable_residuals[mask]))),
        "constraint_status": constraint_status,
    }
    return metrics, adjusted, strokes


def ideal_ray_test(
    axis: np.ndarray,
    vertex: np.ndarray,
    focus: np.ndarray,
    focal_length: float,
) -> dict:
    e1, e2 = transverse_basis(axis)
    rho = np.linspace(0.0, APERTURE_RADIUS_M, 41)
    theta = np.linspace(0.0, 2.0 * np.pi, 96, endpoint=False)
    rr, tt = np.meshgrid(rho, theta, indexing="ij")
    transverse = (
        rr[..., None] * np.cos(tt)[..., None] * e1
        + rr[..., None] * np.sin(tt)[..., None] * e2
    )
    points = vertex + transverse + (rr**2 / (4.0 * focal_length))[..., None] * axis
    normals = axis - transverse / (2.0 * focal_length)
    normals /= np.linalg.norm(normals, axis=-1)[..., None]
    incoming = -axis
    reflected = incoming - 2.0 * np.sum(incoming * normals, axis=-1)[..., None] * normals
    parameter = np.sum((focus - points) * axis, axis=-1) / np.sum(
        reflected * axis, axis=-1
    )
    hits = points + parameter[..., None] * reflected
    error = hits - focus
    transverse_error = error - np.sum(error * axis, axis=-1)[..., None] * axis
    max_error = float(np.max(np.linalg.norm(transverse_error, axis=-1)))
    return {
        "sample_count": int(points.shape[0] * points.shape[1]),
        "max_focal_plane_transverse_error_m": max_error,
        "status": trinary(max_error <= 1e-9),
    }


def ray_evaluate(
    vertices: np.ndarray,
    panels: np.ndarray,
    axis: np.ndarray,
    focus: np.ndarray,
    sample_power: int,
    receiver_radii: Iterable[float],
    seed: int,
    collect_hits: bool = False,
    panel_model: str = "planar_facets",
    reference_vertices: np.ndarray | None = None,
    sphere_radius: float | None = None,
) -> dict:
    """Projected-power ray integration under one of two panel envelopes.

    ``planar_facets`` uses the three current vertices as a flat triangular
    mirror.  ``inherited_spherical_curvature`` starts from the reference
    spherical patch and adds the barycentric interpolation of its three nodal
    displacements; differentiating that map supplies a pointwise normal.
    """

    if panel_model not in {"planar_facets", "inherited_spherical_curvature"}:
        raise ValueError(f"Unknown panel model: {panel_model}")
    if panel_model == "inherited_spherical_curvature":
        if reference_vertices is None or sphere_radius is None:
            raise ValueError("Curved panels require reference vertices and sphere radius.")
        reference_vertices = np.asarray(reference_vertices, dtype=float)

    receiver_radii = np.asarray(list(receiver_radii), dtype=float)
    unit_square = qmc.Sobol(d=2, scramble=True, seed=seed).random_base2(sample_power)
    root_u = np.sqrt(unit_square[:, 0])
    barycentric = np.column_stack(
        [
            1.0 - root_u,
            root_u * (1.0 - unit_square[:, 1]),
            root_u * unit_square[:, 1],
        ]
    )
    incoming = -axis
    denominator = 0.0
    numerators = np.zeros(len(receiver_radii), dtype=float)
    contributing_panels = 0
    e1, e2 = transverse_basis(axis)
    hit_x_parts: list[np.ndarray] = []
    hit_y_parts: list[np.ndarray] = []
    hit_weight_parts: list[np.ndarray] = []
    samples_per_panel = len(barycentric)

    for panel in panels:
        triangle = vertices[panel]
        if panel_model == "planar_facets":
            cross_vector = np.cross(
                triangle[1] - triangle[0], triangle[2] - triangle[0]
            )
            cross_norm = float(np.linalg.norm(cross_vector))
            if cross_norm <= 0.0:
                continue
            normal = cross_vector / cross_norm
            normals = np.broadcast_to(normal, (samples_per_panel, 3))
            points = barycentric @ triangle
            projected_weights = np.full(
                samples_per_panel,
                0.5 * abs(float(incoming @ cross_vector)),
                dtype=float,
            )
        else:
            reference_triangle = reference_vertices[panel]
            displacement_triangle = triangle - reference_triangle
            chord_points = barycentric @ reference_triangle
            chord_norms = np.linalg.norm(chord_points, axis=1)
            unit_radial = chord_points / chord_norms[:, None]
            spherical_points = float(sphere_radius) * unit_radial
            tangent_u = reference_triangle[1] - reference_triangle[0]
            tangent_v = reference_triangle[2] - reference_triangle[0]
            derivative_u = float(sphere_radius) / chord_norms[:, None] * (
                tangent_u - np.einsum("ij,j->i", unit_radial, tangent_u)[:, None] * unit_radial
            )
            derivative_v = float(sphere_radius) / chord_norms[:, None] * (
                tangent_v - np.einsum("ij,j->i", unit_radial, tangent_v)[:, None] * unit_radial
            )
            derivative_u += displacement_triangle[1] - displacement_triangle[0]
            derivative_v += displacement_triangle[2] - displacement_triangle[0]
            cross_vectors = np.cross(derivative_u, derivative_v)
            cross_norms = np.linalg.norm(cross_vectors, axis=1)
            valid_geometry = cross_norms > 0.0
            if not np.any(valid_geometry):
                continue
            normals = np.zeros_like(cross_vectors)
            normals[valid_geometry] = (
                cross_vectors[valid_geometry] / cross_norms[valid_geometry, None]
            )
            points = spherical_points + barycentric @ displacement_triangle
            projected_weights = 0.5 * np.abs(cross_vectors @ incoming)

        axial = points @ axis
        rho_squared = np.einsum("ij,ij->i", points, points) - axial**2
        inside = (rho_squared <= APERTURE_RADIUS_M**2) & (axial < 0.0)
        if not np.any(inside):
            continue
        contributing_panels += 1
        denominator += float(np.mean(projected_weights * inside))

        incident_dot = normals @ incoming
        reflected = incoming - 2.0 * incident_dot[:, None] * normals
        reflected_axial = reflected @ axis
        parameter = np.zeros(samples_per_panel, dtype=float)
        toward_focal_plane = reflected_axial > 1e-12
        parameter[toward_focal_plane] = (
            ((focus - points[toward_focal_plane]) @ axis)
            / reflected_axial[toward_focal_plane]
        )
        hits = points + parameter[:, None] * reflected
        hit_offset = hits - focus
        hit_rho_squared = np.einsum("ij,ij->i", hit_offset, hit_offset) - (
            hit_offset @ axis
        ) ** 2
        forward = toward_focal_plane & (parameter > 0.0)
        for index, receiver_radius in enumerate(receiver_radii):
            received = inside & forward & (hit_rho_squared <= receiver_radius**2)
            numerators[index] += float(np.mean(projected_weights * received))

        if collect_hits:
            selected = inside & forward
            if np.any(selected):
                hit_x_parts.append(hit_offset[selected] @ e1)
                hit_y_parts.append(hit_offset[selected] @ e2)
                hit_weight_parts.append(projected_weights[selected] / samples_per_panel)

    ratios = numerators / denominator
    result = {
        "sample_power": int(sample_power),
        "samples_per_panel": int(2**sample_power),
        "seed": int(seed),
        "panel_model": panel_model,
        "receiver_radii_m": receiver_radii,
        "numerator_projected_areas_m2": numerators,
        "denominator_projected_area_m2": float(denominator),
        "ratios": ratios,
        "contributing_panel_count": int(contributing_panels),
        "aperture_area_relative_error": float(
            denominator / (math.pi * APERTURE_RADIUS_M**2) - 1.0
        ),
    }
    if collect_hits:
        hit_x = np.concatenate(hit_x_parts) if hit_x_parts else np.empty(0, dtype=float)
        hit_y = np.concatenate(hit_y_parts) if hit_y_parts else np.empty(0, dtype=float)
        hit_weight = (
            np.concatenate(hit_weight_parts)
            if hit_weight_parts
            else np.empty(0, dtype=float)
        )
        hit_radius = np.hypot(hit_x, hit_y)
        total_forward_weight = float(np.sum(hit_weight))
        result["hit_x_m"] = hit_x
        result["hit_y_m"] = hit_y
        result["hit_weight_m2"] = hit_weight
        result["forward_projected_area_m2"] = total_forward_weight
        result["max_forward_hit_radius_m"] = (
            float(np.max(hit_radius)) if len(hit_radius) else 0.0
        )
        result["power_fraction_within_6m"] = (
            float(np.sum(hit_weight[hit_radius <= 6.0]) / total_forward_weight)
            if total_forward_weight > 0.0
            else 0.0
        )
    return result


def audit_inputs(data: InputData) -> dict:
    radii = np.linalg.norm(data.nodes, axis=1)
    top_to_bottom = data.actuator_top - data.actuator_bottom
    axes = top_to_bottom / np.linalg.norm(top_to_bottom, axis=1)[:, None]
    inward = -data.actuator_top / np.linalg.norm(data.actuator_top, axis=1)[:, None]
    triangle_cross = np.cross(
        data.nodes[data.panel_indices[:, 1]] - data.nodes[data.panel_indices[:, 0]],
        data.nodes[data.panel_indices[:, 2]] - data.nodes[data.panel_indices[:, 0]],
    )
    triangle_areas = 0.5 * np.linalg.norm(triangle_cross, axis=1)
    file_records = []
    for path in sorted(INPUT_DIR.rglob("*")):
        if path.is_file():
            file_records.append(
                {
                    "relative_path": path.relative_to(ROOT).as_posix(),
                    "bytes": path.stat().st_size,
                    "sha256": sha256_file(path),
                }
            )
    actuator_ids = data.actuators_frame.iloc[:, 0].astype(str).to_numpy()
    return {
        "status": trinary(
            len(data.nodes) == 2226
            and len(data.panels_frame) == 4300
            and len(data.edges) == 6525
            and np.array_equal(data.node_ids, actuator_ids)
            and np.all(triangle_areas > 0.0)
        ),
        "encoding": "GB18030 for the three CSV attachments",
        "files": file_records,
        "counts": {
            "nodes": int(len(data.nodes)),
            "actuators": int(len(data.actuator_top)),
            "panels": int(len(data.panel_indices)),
            "unique_edges": int(len(data.edges)),
        },
        "missing_cells": {
            "nodes": int(data.nodes_frame.isna().sum().sum()),
            "actuators": int(data.actuators_frame.isna().sum().sum()),
            "panels": int(data.panels_frame.isna().sum().sum()),
        },
        "duplicate_rows": {
            "nodes": int(data.nodes_frame.duplicated().sum()),
            "actuators": int(data.actuators_frame.duplicated().sum()),
            "panels": int(data.panels_frame.duplicated().sum()),
        },
        "node_radius_m": {
            "mean": float(np.mean(radii)),
            "std": float(np.std(radii)),
            "min": float(np.min(radii)),
            "max": float(np.max(radii)),
            "max_abs_deviation_from_median": float(np.max(np.abs(radii - np.median(radii)))),
        },
        "actuator_axis_inward_alignment": {
            "min_cosine": float(np.min(np.einsum("ij,ij->i", axes, inward))),
            "status": trinary(np.min(np.einsum("ij,ij->i", axes, inward)) > 0.999999),
        },
        "triangle_area_m2": {
            "min": float(np.min(triangle_areas)),
            "median": float(np.median(triangle_areas)),
            "max": float(np.max(triangle_areas)),
        },
        "id_alignment_status": trinary(np.array_equal(data.node_ids, actuator_ids)),
        "template_sheets": openpyxl.load_workbook(
            data.workbook_path, read_only=True
        ).sheetnames,
    }


def save_workbook(
    template_path: Path,
    vertex: np.ndarray,
    node_ids: np.ndarray,
    adjusted: np.ndarray,
    strokes: np.ndarray,
    mask: np.ndarray,
) -> None:
    workbook = openpyxl.load_workbook(template_path)
    sheet_vertex = workbook["理想抛物面顶点坐标"]
    sheet_nodes = workbook["调整后主索节点编号及坐标"]
    sheet_strokes = workbook["促动器顶端伸缩量"]
    for column, value in enumerate(vertex, start=1):
        sheet_vertex.cell(row=2, column=column, value=round(float(value), OUTPUT_DECIMALS))
    active_indices = np.flatnonzero(mask)
    for output_row, index in enumerate(active_indices, start=2):
        sheet_nodes.cell(output_row, 1, str(node_ids[index]))
        for column, value in enumerate(adjusted[index], start=2):
            sheet_nodes.cell(output_row, column, round(float(value), OUTPUT_DECIMALS))
        sheet_strokes.cell(output_row, 1, str(node_ids[index]))
        sheet_strokes.cell(output_row, 2, round(float(strokes[index]), OUTPUT_DECIMALS))
    workbook.save(ROOT / "<SOURCE_FILE_REDACTED>")


def save_figures(
    data: InputData,
    sphere_radius: float,
    nominal_focal_length: float,
    ideal_focal_length: float,
    axis: np.ndarray,
    vertex: np.ndarray,
    mask: np.ndarray,
    adjusted: np.ndarray,
    strokes: np.ndarray,
    surface_distances: np.ndarray,
    strains: np.ndarray,
    convergence: pd.DataFrame,
    hit_sets: dict[str, dict],
    model_comparison: pd.DataFrame,
) -> None:
    plt.rcParams.update({"font.size": 10, "axes.grid": True, "grid.alpha": 0.25})

    rho = np.linspace(0.0, APERTURE_RADIUS_M, 2001)
    sphere_axial = -np.sqrt(sphere_radius**2 - rho**2)
    vertex_radius = sphere_radius - nominal_focal_length + ideal_focal_length
    parabola_axial = -vertex_radius + rho**2 / (4.0 * ideal_focal_length)
    q_axis = -np.sqrt(1.0 - (rho / sphere_radius) ** 2)
    coefficient = (1.0 - q_axis**2) / (4.0 * ideal_focal_length)
    root = np.sqrt(q_axis**2 + 4.0 * coefficient * vertex_radius)
    radial_radius = np.empty_like(rho)
    regular = coefficient > 1e-15
    radial_radius[regular] = 2.0 * vertex_radius / (root[regular] - q_axis[regular])
    radial_radius[~regular] = vertex_radius
    radial_displacement = sphere_radius - radial_radius

    fig, axes_plot = plt.subplots(1, 2, figsize=(11, 4.2))
    axes_plot[0].plot(rho, sphere_axial, label="Reference sphere")
    axes_plot[0].plot(rho, parabola_axial, label="Ideal paraboloid")
    axes_plot[0].set_xlabel("Transverse radius (m)")
    axes_plot[0].set_ylabel("Axial coordinate (m)")
    axes_plot[0].legend()
    axes_plot[0].set_title("Meridional profiles")
    axes_plot[1].plot(rho, radial_displacement)
    axes_plot[1].axhline(STROKE_LIMIT_M, color="r", linestyle="--", linewidth=0.8)
    axes_plot[1].axhline(-STROKE_LIMIT_M, color="r", linestyle="--", linewidth=0.8)
    axes_plot[1].set_xlabel("Transverse radius (m)")
    axes_plot[1].set_ylabel("Radial displacement (m)")
    axes_plot[1].set_title("Joint-design ideal displacement")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)

    e1, e2 = transverse_basis(axis)
    local_x = data.nodes[mask] @ e1
    local_y = data.nodes[mask] @ e2
    fig, axes_plot = plt.subplots(1, 2, figsize=(11, 4.7))
    scatter = axes_plot[0].scatter(
        local_x, local_y, c=strokes[mask], s=13, cmap="coolwarm", rasterized=True
    )
    fig.colorbar(scatter, ax=axes_plot[0], label="Actuator stroke (m)")
    axes_plot[0].set_title("Actuator field")
    scatter = axes_plot[1].scatter(
        local_x, local_y, c=surface_distances, s=13, cmap="PiYG", rasterized=True
    )
    fig.colorbar(scatter, ax=axes_plot[1], label="Signed surface error (m)")
    axes_plot[1].set_title("Deviation from ideal paraboloid")
    for axis_plot in axes_plot:
        axis_plot.set_aspect("equal")
        axis_plot.set_xlabel("Local transverse x (m)")
        axis_plot.set_ylabel("Local transverse y (m)")
        axis_plot.add_patch(Circle((0, 0), APERTURE_RADIUS_M, fill=False, color="k", linewidth=0.7))
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)

    fig, axes_plot = plt.subplots(1, 2, figsize=(11, 4.2))
    axes_plot[0].hist(100.0 * strains, bins=45, color="#4472C4", alpha=0.85)
    axes_plot[0].axvline(100.0 * EDGE_STRAIN_LIMIT, color="r", linestyle="--")
    axes_plot[0].axvline(-100.0 * EDGE_STRAIN_LIMIT, color="r", linestyle="--")
    axes_plot[0].set_xlabel("Exact edge-length change (%)")
    axes_plot[0].set_ylabel("Edge count")
    axes_plot[0].set_title("Affected-edge validation")
    axes_plot[1].hist(strokes[mask], bins=38, color="#70AD47", alpha=0.85)
    axes_plot[1].axvline(STROKE_LIMIT_M, color="r", linestyle="--")
    axes_plot[1].axvline(-STROKE_LIMIT_M, color="r", linestyle="--")
    axes_plot[1].set_xlabel("Actuator stroke (m)")
    axes_plot[1].set_ylabel("Node count")
    axes_plot[1].set_title("Stroke validation")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)

    fig, axis_plot = plt.subplots(figsize=(7.6, 4.8))
    styles = {
        "planar_facets": ("--", "Planar facets"),
        "inherited_spherical_curvature": ("-", "Inherited curvature"),
    }
    for panel_model, group in convergence.groupby("panel_model", sort=False):
        linestyle, model_label = styles[panel_model]
        axis_plot.plot(
            group["samples_per_panel"],
            100.0 * group["adjusted_ratio"],
            marker="o",
            linestyle=linestyle,
            label=f"{model_label}: adjusted",
        )
        axis_plot.plot(
            group["samples_per_panel"],
            100.0 * group["baseline_ratio"],
            marker="s",
            linestyle=linestyle,
            label=f"{model_label}: sphere",
        )
    axis_plot.set_xscale("log", base=2)
    axis_plot.set_xlabel("Sobol samples per panel")
    axis_plot.set_ylabel("Reception ratio (%)")
    axis_plot.set_title("Ray-integration convergence by panel model")
    axis_plot.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)

    hit_order = [
        ("curved_adjusted", "Inherited curvature: adjusted"),
        ("curved_baseline", "Inherited curvature: sphere"),
        ("planar_adjusted", "Planar facets: adjusted"),
        ("planar_baseline", "Planar facets: sphere"),
    ]

    def weighted_spot_figure(limit: float, output_name: str, local_zoom: bool) -> None:
        histograms = []
        for key, _ in hit_order:
            hits = hit_sets[key]
            histogram, _, _ = np.histogram2d(
                hits["hit_x_m"],
                hits["hit_y_m"],
                bins=150,
                range=[[-limit, limit], [-limit, limit]],
                weights=hits["hit_weight_m2"],
            )
            histograms.append(histogram.T)
        positive = np.concatenate([histogram[histogram > 0.0] for histogram in histograms])
        maximum = float(np.max(positive)) if len(positive) else 1.0
        minimum = max(float(np.min(positive)) if len(positive) else 1e-12, maximum * 1e-5)
        norm = LogNorm(vmin=minimum, vmax=maximum)
        fig, axes_plot = plt.subplots(2, 2, figsize=(11.5, 9.0), sharex=True, sharey=True)
        plotted = None
        for axis_plot, histogram, (key, title) in zip(axes_plot.ravel(), histograms, hit_order):
            plotted = axis_plot.imshow(
                histogram,
                origin="lower",
                extent=(-limit, limit, -limit, limit),
                cmap="viridis",
                norm=norm,
                interpolation="nearest",
            )
            axis_plot.add_patch(
                Circle((0, 0), RECEIVER_RADIUS_M, fill=False, color="r", linewidth=1.0)
            )
            hits = hit_sets[key]
            annotation = (
                f"max r={hits['max_forward_hit_radius_m']:.1f} m\n"
                f"power within 6 m={100*hits['power_fraction_within_6m']:.1f}%"
            )
            axis_plot.text(
                0.02,
                0.98,
                annotation,
                transform=axis_plot.transAxes,
                va="top",
                fontsize=8,
                color="white",
                bbox={"facecolor": "black", "alpha": 0.45, "edgecolor": "none"},
            )
            axis_plot.set_title(title)
            axis_plot.set_aspect("equal")
            axis_plot.set_xlabel("Focal-plane x (m)")
            axis_plot.set_ylabel("Focal-plane y (m)")
        color_axis = fig.add_axes([0.88, 0.15, 0.025, 0.70])
        fig.colorbar(plotted, cax=color_axis, label="Projected power per bin (m²)")
        fig.suptitle(
            "Projected-power focal distributions"
            + (" — 6 m conditional zoom" if local_zoom else " — all forward hits")
        )
        fig.subplots_adjust(left=0.07, right=0.84, bottom=0.07, top=0.92, wspace=0.20, hspace=0.20)
        fig.savefig(FIGURES_DIR / output_name, dpi=180)
        plt.close(fig)

    full_limit = max(
        6.0,
        math.ceil(
            max(
                max(np.max(np.abs(hit_sets[key]["hit_x_m"])), np.max(np.abs(hit_sets[key]["hit_y_m"])))
                if len(hit_sets[key]["hit_x_m"])
                else 0.0
                for key, _ in hit_order
            )
        ),
    )
    weighted_spot_figure(float(full_limit), "<SOURCE_FILE_REDACTED>", local_zoom=False)
    weighted_spot_figure(6.0, "<SOURCE_FILE_REDACTED>", local_zoom=True)

    labels = model_comparison["display_label"].tolist()
    colors = ["#A5A5A5", "#5B9BD5", "#70AD47"]
    fig, axes_plot = plt.subplots(2, 2, figsize=(11.0, 8.0))
    axes_plot[0, 0].bar(labels, model_comparison["radial_residual_rms_m"], color=colors)
    axes_plot[0, 0].set_ylabel("Primary radial RMS (m)")
    axes_plot[0, 0].set_title("Declared objective")
    axes_plot[0, 1].bar(labels, model_comparison["surface_distance_max_abs_m"], color=colors)
    axes_plot[0, 1].set_ylabel("Maximum orthogonal error (m)")
    axes_plot[0, 1].set_title("Pareto companion metric")
    axes_plot[1, 0].bar(labels, model_comparison["max_abs_stroke_m"], color=colors)
    axes_plot[1, 0].axhline(STROKE_LIMIT_M, color="r", linestyle="--")
    axes_plot[1, 0].set_ylabel("Maximum actuator stroke (m)")
    axes_plot[1, 0].set_title("Stroke comparison")
    axes_plot[1, 1].bar(
        labels,
        100.0 * model_comparison["max_abs_exact_edge_strain"],
        color=colors,
    )
    axes_plot[1, 1].axhline(100.0 * EDGE_STRAIN_LIMIT, color="r", linestyle="--")
    axes_plot[1, 1].set_ylabel("Maximum edge-length change (%)")
    axes_plot[1, 1].set_title("Exact constraint comparison")
    for axis_plot in axes_plot.ravel():
        axis_plot.tick_params(axis="x", labelrotation=8)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "<SOURCE_FILE_REDACTED>", dpi=180)
    plt.close(fig)


def main() -> int:
    np.random.seed(SEED)
    ensure_directories()
    data = discover_inputs()
    audit = audit_inputs(data)

    radii = np.linalg.norm(data.nodes, axis=1)
    sphere_radius = float(np.median(radii))
    nominal_focal_length = FOCAL_RATIO * sphere_radius
    focus_radius = sphere_radius - nominal_focal_length
    axis = direction_vector(ALPHA_DEG, BETA_DEG)
    directions = data.nodes / radii[:, None]
    mask = active_mask(data.nodes, axis)
    relevant_edges = affected_edges(data.edges, mask)
    strain_matrix, active_indices = linear_strain_matrix(
        data.nodes, directions, relevant_edges, mask
    )
    if not np.array_equal(active_indices, np.flatnonzero(mask)):
        raise RuntimeError("Active-node indexing mismatch.")

    question_one_focal, minimax_details = continuous_minimax_paraboloid(
        sphere_radius, nominal_focal_length
    )
    question_one_vertex_radius = focus_radius + question_one_focal
    focus = -focus_radius * axis

    selected_focal, selected_displacements, joint_payload, focal_search = joint_focal_search(
        data,
        radii,
        sphere_radius,
        nominal_focal_length,
        axis,
        directions,
        mask,
        relevant_edges,
        strain_matrix,
    )
    focal_search.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )
    selected_target = joint_payload["target"]
    selected_vertex_radius = focus_radius + selected_focal
    vertex = -selected_vertex_radius * axis
    solver_details = joint_payload["solver"]

    scale, scaled_displacements, scale_details = maximum_feasible_uniform_scale(
        selected_target, data, directions, mask, relevant_edges
    )
    scaled_metrics, scaled_adjusted, scaled_strokes = candidate_metrics(
        "B0_uniform_scale",
        selected_focal,
        scaled_displacements,
        data,
        directions,
        mask,
        relevant_edges,
        sphere_radius,
        nominal_focal_length,
        target=selected_target,
        parameter_count=1,
    )

    v1_target_radii = radial_intersection_radii(
        directions,
        sphere_radius,
        nominal_focal_length,
        question_one_focal,
        axis,
    )
    v1_target = (radii - v1_target_radii)[mask]
    v1_displacements, v1_solver = constrained_displacements(
        v1_target, strain_matrix, 0.00060
    )
    v1_metrics, v1_adjusted, v1_strokes = candidate_metrics(
        "B1_frozen_v1_strategy",
        question_one_focal,
        v1_displacements,
        data,
        directions,
        mask,
        relevant_edges,
        sphere_radius,
        nominal_focal_length,
        target=v1_target,
        parameter_count=len(v1_displacements),
    )

    selected_metrics, adjusted, strokes = candidate_metrics(
        "J2_joint_focal_projection",
        selected_focal,
        selected_displacements,
        data,
        directions,
        mask,
        relevant_edges,
        sphere_radius,
        nominal_focal_length,
        target=selected_target,
        parameter_count=len(selected_displacements) + 1,
    )

    candidate_payloads = [
        (scaled_metrics, scaled_adjusted, scaled_strokes, "1-param scale"),
        (v1_metrics, v1_adjusted, v1_strokes, "Frozen V1"),
        (selected_metrics, adjusted, strokes, "Joint selected"),
    ]
    candidate_rows = []
    for metrics, candidate_adjusted, candidate_strokes, label in candidate_payloads:
        rounded_candidate = rounded_constraint_metrics(
            candidate_adjusted, candidate_strokes, data, relevant_edges
        )
        row = dict(metrics)
        row["display_label"] = label
        row["rounded_max_abs_edge_strain"] = rounded_candidate[
            "max_abs_edge_strain"
        ]
        row["rounded_constraint_status"] = trinary(
            rounded_candidate["max_abs_edge_strain"] <= EDGE_STRAIN_LIMIT
            and np.max(np.abs(rounded_candidate["strokes"][mask])) <= STROKE_LIMIT_M
            and np.max(np.abs(rounded_candidate["closure"][mask])) <= 2e-6
        )
        candidate_rows.append(row)
    model_comparison = pd.DataFrame(candidate_rows)
    model_comparison.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    audit_anchor_focal = nominal_focal_length + 0.4
    audit_anchor_radii = radial_intersection_radii(
        directions,
        sphere_radius,
        nominal_focal_length,
        audit_anchor_focal,
        axis,
    )
    audit_anchor_target = (radii - audit_anchor_radii)[mask]
    audit_anchor_displacements, audit_anchor_solver = constrained_displacements(
        audit_anchor_target, strain_matrix, 0.00060
    )
    audit_anchor_metrics, audit_anchor_adjusted, audit_anchor_strokes = candidate_metrics(
        "AUD_f_offset_0p4_cap_0p060pct",
        audit_anchor_focal,
        audit_anchor_displacements,
        data,
        directions,
        mask,
        relevant_edges,
        sphere_radius,
        nominal_focal_length,
        target=audit_anchor_target,
        parameter_count=len(audit_anchor_displacements) + 1,
    )
    audit_anchor_rounded = rounded_constraint_metrics(
        audit_anchor_adjusted, audit_anchor_strokes, data, relevant_edges
    )
    anchor_status = trinary(
        audit_anchor_metrics["constraint_status"] == "pass"
        and audit_anchor_rounded["max_abs_edge_strain"] <= EDGE_STRAIN_LIMIT
    )
    selected_not_dominated = trinary(
        selected_metrics["radial_residual_rms_m"]
        <= audit_anchor_metrics["radial_residual_rms_m"] + 1e-9
        and selected_metrics["surface_distance_rms_m"]
        <= audit_anchor_metrics["surface_distance_rms_m"] + 1e-9
    )

    strokes, actuator_axes, cable_lengths, cable_residuals = actuator_strokes(
        data.nodes, adjusted, data.actuator_top, data.actuator_bottom
    )
    strains = edge_strains(data.nodes, adjusted, relevant_edges)
    surface_distances = exact_paraboloid_distances(
        adjusted[mask], axis, vertex, selected_focal
    )
    rounded = rounded_constraint_metrics(adjusted, strokes, data, relevant_edges)
    rounded_adjusted = rounded["adjusted"]
    rounded_strokes = rounded["strokes"]
    rounded_strains = rounded["strains"]
    rounded_cable_residuals = rounded["closure"]

    ideal_check = ideal_ray_test(axis, vertex, focus, selected_focal)

    panel_models = ("planar_facets", "inherited_spherical_curvature")
    convergence_rows = []
    final_rays: dict[tuple[str, str], dict] = {}
    for panel_model in panel_models:
        extra = (
            {
                "reference_vertices": data.nodes,
                "sphere_radius": sphere_radius,
            }
            if panel_model == "inherited_spherical_curvature"
            else {}
        )
        for sample_power in (8, 10, 12, 14):
            adjusted_ray = ray_evaluate(
                adjusted,
                data.panel_indices,
                axis,
                focus,
                sample_power,
                [RECEIVER_RADIUS_M],
                SEED,
                panel_model=panel_model,
                **extra,
            )
            baseline_ray = ray_evaluate(
                data.nodes,
                data.panel_indices,
                axis,
                focus,
                sample_power,
                [RECEIVER_RADIUS_M],
                SEED,
                panel_model=panel_model,
                **extra,
            )
            convergence_rows.append(
                {
                    "panel_model": panel_model,
                    "sample_power": sample_power,
                    "samples_per_panel": 2**sample_power,
                    "adjusted_ratio": float(adjusted_ray["ratios"][0]),
                    "baseline_ratio": float(baseline_ray["ratios"][0]),
                    "relative_improvement": float(
                        adjusted_ray["ratios"][0] / baseline_ray["ratios"][0] - 1.0
                    ),
                    "adjusted_denominator_area_m2": adjusted_ray[
                        "denominator_projected_area_m2"
                    ],
                    "baseline_denominator_area_m2": baseline_ray[
                        "denominator_projected_area_m2"
                    ],
                    "adjusted_aperture_area_relative_error": adjusted_ray[
                        "aperture_area_relative_error"
                    ],
                    "baseline_aperture_area_relative_error": baseline_ray[
                        "aperture_area_relative_error"
                    ],
                }
            )
            final_rays[(panel_model, "adjusted")] = adjusted_ray
            final_rays[(panel_model, "baseline")] = baseline_ray
    convergence = pd.DataFrame(convergence_rows)
    convergence.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    receiver_rows = []
    for panel_model in panel_models:
        extra = (
            {"reference_vertices": data.nodes, "sphere_radius": sphere_radius}
            if panel_model == "inherited_spherical_curvature"
            else {}
        )
        adjusted_receiver = ray_evaluate(
            adjusted,
            data.panel_indices,
            axis,
            focus,
            12,
            [0.45, 0.50, 0.55],
            SEED + 101,
            panel_model=panel_model,
            **extra,
        )
        baseline_receiver = ray_evaluate(
            data.nodes,
            data.panel_indices,
            axis,
            focus,
            12,
            [0.45, 0.50, 0.55],
            SEED + 101,
            panel_model=panel_model,
            **extra,
        )
        for radius_value, adjusted_ratio, baseline_ratio in zip(
            [0.45, 0.50, 0.55],
            adjusted_receiver["ratios"],
            baseline_receiver["ratios"],
        ):
            receiver_rows.append(
                {
                    "panel_model": panel_model,
                    "receiver_radius_m": radius_value,
                    "adjusted_ratio": float(adjusted_ratio),
                    "baseline_ratio": float(baseline_ratio),
                    "relative_improvement": float(adjusted_ratio / baseline_ratio - 1.0),
                }
            )
    receiver_sensitivity = pd.DataFrame(receiver_rows)
    receiver_sensitivity.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    seed_rows = []
    for panel_model in panel_models:
        extra = (
            {"reference_vertices": data.nodes, "sphere_radius": sphere_radius}
            if panel_model == "inherited_spherical_curvature"
            else {}
        )
        for ray_seed in (SEED + 201, SEED + 202, SEED + 203):
            adjusted_seed = ray_evaluate(
                adjusted,
                data.panel_indices,
                axis,
                focus,
                11,
                [RECEIVER_RADIUS_M],
                ray_seed,
                panel_model=panel_model,
                **extra,
            )
            baseline_seed = ray_evaluate(
                data.nodes,
                data.panel_indices,
                axis,
                focus,
                11,
                [RECEIVER_RADIUS_M],
                ray_seed,
                panel_model=panel_model,
                **extra,
            )
            seed_rows.append(
                {
                    "panel_model": panel_model,
                    "seed": ray_seed,
                    "samples_per_panel": 2**11,
                    "adjusted_ratio": float(adjusted_seed["ratios"][0]),
                    "baseline_ratio": float(baseline_seed["ratios"][0]),
                    "relative_improvement": float(
                        adjusted_seed["ratios"][0] / baseline_seed["ratios"][0] - 1.0
                    ),
                }
            )
    seed_sensitivity = pd.DataFrame(seed_rows)
    seed_sensitivity.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    strain_sensitivity_rows = []
    initial_cap_solution = None
    for cap in (0.00060, 0.00061, 0.00062, 0.00063):
        cap_displacements, cap_solver = constrained_displacements(
            selected_target, strain_matrix, cap, initial=initial_cap_solution
        )
        initial_cap_solution = cap_displacements
        cap_adjusted = apply_radial_displacements(
            data.nodes, directions, mask, cap_displacements
        )
        cap_strains = edge_strains(data.nodes, cap_adjusted, relevant_edges)
        cap_strokes, _, _, _ = actuator_strokes(
            data.nodes, cap_adjusted, data.actuator_top, data.actuator_bottom
        )
        cap_rounded = rounded_constraint_metrics(
            cap_adjusted, cap_strokes, data, relevant_edges
        )
        cap_distances = exact_paraboloid_distances(
            cap_adjusted[mask], axis, vertex, selected_focal
        )
        cap_exact_max = float(np.max(np.abs(cap_strains)))
        cap_rounded_max = cap_rounded["max_abs_edge_strain"]
        strain_sensitivity_rows.append(
            {
                "linear_strain_cap": cap,
                "optimizer_status": cap_solver["optimizer_status"],
                "kkt_status": cap_solver["kkt_status"],
                "exact_max_abs_edge_strain": cap_exact_max,
                "rounded_max_abs_edge_strain": cap_rounded_max,
                "radial_residual_rms_m": cap_solver["radial_residual_rms_m"],
                "surface_distance_rms_m": float(np.sqrt(np.mean(cap_distances**2))),
                "surface_distance_max_abs_m": float(np.max(np.abs(cap_distances))),
                "max_abs_actuator_stroke_m": float(np.max(np.abs(cap_strokes[mask]))),
                "exact_constraint_status": trinary(
                    cap_exact_max <= EDGE_STRAIN_LIMIT
                    and np.max(np.abs(cap_strokes[mask])) <= STROKE_LIMIT_M
                ),
                "rounded_constraint_status": trinary(
                    cap_rounded_max <= EDGE_STRAIN_LIMIT
                    and np.max(np.abs(cap_rounded["strokes"][mask])) <= STROKE_LIMIT_M
                    and np.max(np.abs(cap_rounded["closure"][mask])) <= 2e-6
                ),
                "design_reserve_status": trinary(
                    max(cap_exact_max, cap_rounded_max)
                    <= EDGE_STRAIN_LIMIT - DESIGN_EDGE_RESERVE
                ),
            }
        )
    strain_sensitivity = pd.DataFrame(strain_sensitivity_rows)
    strain_sensitivity.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    hit_sets = {}
    for panel_model, prefix in [
        ("inherited_spherical_curvature", "curved"),
        ("planar_facets", "planar"),
    ]:
        extra = (
            {"reference_vertices": data.nodes, "sphere_radius": sphere_radius}
            if panel_model == "inherited_spherical_curvature"
            else {}
        )
        hit_sets[f"{prefix}_adjusted"] = ray_evaluate(
            adjusted,
            data.panel_indices,
            axis,
            focus,
            7,
            [RECEIVER_RADIUS_M],
            SEED + 301,
            collect_hits=True,
            panel_model=panel_model,
            **extra,
        )
        hit_sets[f"{prefix}_baseline"] = ray_evaluate(
            data.nodes,
            data.panel_indices,
            axis,
            focus,
            7,
            [RECEIVER_RADIUS_M],
            SEED + 301,
            collect_hits=True,
            panel_model=panel_model,
            **extra,
        )
    spot_metrics = {
        key: {
            "panel_model": value["panel_model"],
            "samples_per_panel": value["samples_per_panel"],
            "denominator_projected_area_m2": value[
                "denominator_projected_area_m2"
            ],
            "forward_projected_area_m2": value["forward_projected_area_m2"],
            "forward_power_closure_relative_error": (
                value["forward_projected_area_m2"]
                / value["denominator_projected_area_m2"]
                - 1.0
            ),
            "max_forward_hit_radius_m": value["max_forward_hit_radius_m"],
            "power_fraction_within_6m": value["power_fraction_within_6m"],
        }
        for key, value in hit_sets.items()
    }
    write_json(RESULTS_DIR / "focal_spot_metrics.json", spot_metrics)

    ray_model_validation = {}
    for panel_model in panel_models:
        group = convergence[convergence["panel_model"] == panel_model].reset_index(drop=True)
        final_adjusted = final_rays[(panel_model, "adjusted")]
        final_baseline = final_rays[(panel_model, "baseline")]
        adjusted_ratio = float(final_adjusted["ratios"][0])
        baseline_ratio = float(final_baseline["ratios"][0])
        convergence_adjusted = float(
            abs(group.iloc[-1]["adjusted_ratio"] - group.iloc[-2]["adjusted_ratio"])
        )
        convergence_baseline = float(
            abs(group.iloc[-1]["baseline_ratio"] - group.iloc[-2]["baseline_ratio"])
        )
        seed_group = seed_sensitivity[seed_sensitivity["panel_model"] == panel_model]
        adjusted_seed_range = float(
            seed_group["adjusted_ratio"].max() - seed_group["adjusted_ratio"].min()
        )
        baseline_seed_range = float(
            seed_group["baseline_ratio"].max() - seed_group["baseline_ratio"].min()
        )
        numerical_status = trinary(
            convergence_adjusted <= 5e-5
            and convergence_baseline <= 5e-5
            and adjusted_seed_range <= 5e-4
            and baseline_seed_range <= 5e-4
            and abs(final_adjusted["aperture_area_relative_error"]) <= 2e-4
            and abs(final_baseline["aperture_area_relative_error"]) <= 2e-4
        )
        ray_model_validation[panel_model] = {
            "final_samples_per_panel": int(2**14),
            "adjusted_ratio": adjusted_ratio,
            "baseline_ratio": baseline_ratio,
            "relative_improvement": adjusted_ratio / baseline_ratio - 1.0,
            "adjusted_last_level_absolute_change": convergence_adjusted,
            "baseline_last_level_absolute_change": convergence_baseline,
            "adjusted_seed_range": adjusted_seed_range,
            "baseline_seed_range": baseline_seed_range,
            "adjusted_aperture_area_relative_error": final_adjusted[
                "aperture_area_relative_error"
            ],
            "baseline_aperture_area_relative_error": final_baseline[
                "aperture_area_relative_error"
            ],
            "numerical_status": numerical_status,
        }

    primary_ray = ray_model_validation["inherited_spherical_curvature"]
    planar_ray = ray_model_validation["planar_facets"]
    overall_ray_numerical_status = trinary(
        all(item["numerical_status"] == "pass" for item in ray_model_validation.values())
    )

    rounded_output_status = trinary(
        np.max(np.abs(rounded_strains)) <= EDGE_STRAIN_LIMIT
        and np.max(np.abs(rounded_strokes[mask])) <= STROKE_LIMIT_M
        and np.max(np.abs(rounded_cable_residuals[mask])) <= 2e-6
    )
    validation = {
        "active_node_count": int(np.sum(mask)),
        "affected_edge_count": int(len(relevant_edges)),
        "question_1_minimax": minimax_details,
        "joint_focal_search": joint_payload["joint_search"],
        "solver": solver_details,
        "audit_counterexample_anchor": {
            "focal_length_m": audit_anchor_focal,
            "linear_strain_cap": 0.00060,
            "radial_residual_rms_m": audit_anchor_metrics["radial_residual_rms_m"],
            "orthogonal_surface_rms_m": audit_anchor_metrics["surface_distance_rms_m"],
            "exact_max_abs_edge_strain": audit_anchor_metrics[
                "max_abs_exact_edge_strain"
            ],
            "rounded_max_abs_edge_strain": audit_anchor_rounded[
                "max_abs_edge_strain"
            ],
            "solver_status": audit_anchor_solver["optimizer_status"],
            "constraint_status": anchor_status,
            "selected_not_dominated_status": selected_not_dominated,
        },
        "uniform_scale_baseline": {
            "scale": scale,
            "status": scale_details["status"],
            "radial_residual_rms_m": scaled_metrics["radial_residual_rms_m"],
            "orthogonal_surface_rms_m": scaled_metrics["surface_distance_rms_m"],
            "orthogonal_surface_max_abs_m": scaled_metrics[
                "surface_distance_max_abs_m"
            ],
        },
        "stroke": {
            "min_m": float(np.min(strokes[mask])),
            "max_m": float(np.max(strokes[mask])),
            "max_abs_m": float(np.max(np.abs(strokes[mask]))),
            "limit_m": STROKE_LIMIT_M,
            "status": trinary(np.max(np.abs(strokes[mask])) <= STROKE_LIMIT_M),
        },
        "edge_strain": {
            "min": float(np.min(strains)),
            "max": float(np.max(strains)),
            "max_abs": float(np.max(np.abs(strains))),
            "limit": EDGE_STRAIN_LIMIT,
            "design_reserve": DESIGN_EDGE_RESERVE,
            "violation_count": int(np.sum(np.abs(strains) > EDGE_STRAIN_LIMIT)),
            "status": trinary(np.max(np.abs(strains)) <= EDGE_STRAIN_LIMIT),
            "design_reserve_status": trinary(
                np.max(np.abs(strains)) <= EDGE_STRAIN_LIMIT - DESIGN_EDGE_RESERVE
            ),
        },
        "surface_distance": {
            "primary_objective_radial_rms_m": selected_metrics[
                "radial_residual_rms_m"
            ],
            "mean_signed_orthogonal_m": float(np.mean(surface_distances)),
            "orthogonal_rms_m": float(np.sqrt(np.mean(surface_distances**2))),
            "orthogonal_max_abs_m": float(np.max(np.abs(surface_distances))),
        },
        "cable_closure": {
            "max_abs_m": float(np.max(np.abs(cable_residuals[mask]))),
            "status": trinary(np.max(np.abs(cable_residuals[mask])) <= 1e-9),
        },
        "rounded_output": {
            "decimals": OUTPUT_DECIMALS,
            "max_abs_edge_strain": float(np.max(np.abs(rounded_strains))),
            "max_abs_stroke_m": float(np.max(np.abs(rounded_strokes[mask]))),
            "max_abs_cable_residual_m": float(
                np.max(np.abs(rounded_cable_residuals[mask]))
            ),
            "status": rounded_output_status,
        },
        "ideal_ray_identity": ideal_check,
        "ray_integration": {
            "models": ray_model_validation,
            "numerical_status": overall_ray_numerical_status,
            "panel_geometry_interpretation_status": "needs_review",
            "physical_external_validity_status": "needs_review",
            "note": (
                "The inherited-curvature result is primary because the statement says each panel is part of the reference sphere; "
                "the adjusted-patch displacement interpolation remains a modeling assumption, so both envelopes are reported."
            ),
        },
    }

    ideal_payload = {
        "sphere_radius_m": sphere_radius,
        "nominal_focal_ratio": FOCAL_RATIO,
        "nominal_focal_length_m": nominal_focal_length,
        "focus_radius_m": focus_radius,
        "question_1_vertical": {
            "selection_criterion": "continuous minimax radial displacement",
            "criterion_uniqueness_status": "needs_review",
            "focal_length_m": question_one_focal,
            "focal_ratio": question_one_focal / sphere_radius,
            "axis": [0.0, 0.0, 1.0],
            "focus_m": [0.0, 0.0, -focus_radius],
            "vertex_m": [0.0, 0.0, -question_one_vertex_radius],
            "equation": (
                f"z = {-question_one_vertex_radius:.12f} + (x^2+y^2)/({4*question_one_focal:.12f})"
            ),
            "minimax": minimax_details,
        },
        "question_2_joint_design": {
            "selection_criterion": "minimum active-node radial RMS after feasible projection",
            "focal_length_m": selected_focal,
            "focal_ratio": selected_focal / sphere_radius,
            "focal_offset_from_nominal_m": selected_focal - nominal_focal_length,
            "alpha_deg": ALPHA_DEG,
            "beta_deg": BETA_DEG,
            "axis": axis,
            "focus_m": focus,
            "vertex_m": vertex,
            "vector_equation": "u·(x-V)=||(I-uu^T)(x-V)||^2/(4f)",
            "joint_search": joint_payload["joint_search"],
        },
    }
    write_json(RESULTS_DIR / "input_audit.json", audit)
    write_json(RESULTS_DIR / "ideal_paraboloid.json", ideal_payload)
    write_json(RESULTS_DIR / "validation_metrics.json", validation)

    active_indices = np.flatnonzero(mask)
    adjusted_frame = pd.DataFrame(
        {
            "node_id": data.node_ids[active_indices],
            "x_m": rounded_adjusted[active_indices, 0],
            "y_m": rounded_adjusted[active_indices, 1],
            "z_m": rounded_adjusted[active_indices, 2],
            "radial_displacement_m": np.round(selected_displacements, OUTPUT_DECIMALS),
            "actuator_stroke_m": rounded_strokes[active_indices],
            "signed_surface_distance_m": np.round(surface_distances, OUTPUT_DECIMALS),
        }
    )
    adjusted_frame.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )
    pd.DataFrame(
        {
            "node_id": data.node_ids[active_indices],
            "actuator_stroke_m": rounded_strokes[active_indices],
        }
    ).to_csv(RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig")

    evidence_matrix = pd.DataFrame(
        [
            {
                "layer": "input_identity",
                "status": audit["status"],
                "evidence": "results/input_audit.json with actual SHA-256 values",
            },
            {
                "layer": "joint_model_selection",
                "status": trinary(
                    joint_payload["joint_search"]["status"] == "pass"
                    and selected_not_dominated == "pass"
                ),
                "evidence": "full focal grid, bounded refinement, KKT residual, and audit anchor",
            },
            {
                "layer": "internal_geometry_constraints",
                "status": trinary(
                    validation["stroke"]["status"] == "pass"
                    and validation["edge_strain"]["status"] == "pass"
                    and validation["cable_closure"]["status"] == "pass"
                    and validation["rounded_output"]["status"] == "pass"
                ),
                "evidence": "exact nonlinear and six-decimal checks in validation_metrics.json",
            },
            {
                "layer": "numerical_ray_integration",
                "status": overall_ray_numerical_status,
                "evidence": "two panel envelopes, nested Sobol refinement, three seeds, and ideal identity",
            },
            {
                "layer": "panel_geometry_interpretation",
                "status": "needs_review",
                "evidence": "planar and inherited-curvature envelopes materially differ",
            },
            {
                "layer": "physical_external_validity",
                "status": "needs_review",
                "evidence": "geometric optics omits gaps, blockage, feed pattern, and measured panel pose",
            },
        ]
    )
    evidence_matrix.to_csv(
        RESULTS_DIR / "<SOURCE_FILE_REDACTED>", index=False, encoding="utf-8-sig"
    )

    summary = {
        "case_id": "2021A",
        "phase": "blind-revision",
        "seed": SEED,
        "question_1": {
            "selection_criterion": "continuous minimax radial displacement",
            "criterion_uniqueness_status": "needs_review",
            "ideal_focal_length_m": question_one_focal,
            "ideal_focal_ratio": question_one_focal / sphere_radius,
            "vertical_vertex_m": [0.0, 0.0, -question_one_vertex_radius],
            "max_ideal_radial_displacement_m": minimax_details[
                "objective_max_radial_displacement_m"
            ],
        },
        "question_2": {
            "selection_criterion": "joint focal and nodal radial-RMS minimization",
            "focal_length_m": selected_focal,
            "focal_offset_from_nominal_m": selected_focal - nominal_focal_length,
            "vertex_m": vertex,
            "active_node_count": int(np.sum(mask)),
            "stroke_range_m": [
                float(np.min(strokes[mask])),
                float(np.max(strokes[mask])),
            ],
            "max_abs_edge_strain": float(np.max(np.abs(strains))),
            "radial_residual_rms_m": selected_metrics["radial_residual_rms_m"],
            "surface_distance_rms_m": float(np.sqrt(np.mean(surface_distances**2))),
            "surface_distance_max_abs_m": float(np.max(np.abs(surface_distances))),
            "audit_anchor_not_better_status": selected_not_dominated,
            "internal_constraint_status": selected_metrics["constraint_status"],
            "rounded_output_status": rounded_output_status,
        },
        "question_3": {
            "primary_panel_model": "inherited_spherical_curvature",
            "primary_adjusted_reception_ratio": primary_ray["adjusted_ratio"],
            "primary_baseline_reception_ratio": primary_ray["baseline_ratio"],
            "primary_relative_improvement": primary_ray["relative_improvement"],
            "planar_adjusted_reception_ratio": planar_ray["adjusted_ratio"],
            "planar_baseline_reception_ratio": planar_ray["baseline_ratio"],
            "planar_relative_improvement": planar_ray["relative_improvement"],
            "numerical_status": overall_ray_numerical_status,
            "panel_geometry_interpretation_status": "needs_review",
            "physical_external_validity_status": "needs_review",
        },
    }
    write_json(RESULTS_DIR / "summary.json", summary)

    save_workbook(
        data.workbook_path,
        vertex,
        data.node_ids,
        rounded_adjusted,
        rounded_strokes,
        mask,
    )
    save_figures(
        data,
        sphere_radius,
        nominal_focal_length,
        selected_focal,
        axis,
        vertex,
        mask,
        adjusted,
        strokes,
        surface_distances,
        strains,
        convergence,
        hit_sets,
        model_comparison,
    )

    generated_tex = "\n".join(
        [
            "% Generated by code/solve.py; do not edit numerical values by hand.",
            f"\\newcommand{{\\SphereRadius}}{{{sphere_radius:.4f}}}",
            f"\\newcommand{{\\NominalFocalLength}}{{{nominal_focal_length:.4f}}}",
            f"\\newcommand{{\\QuestionOneFocalLength}}{{{question_one_focal:.6f}}}",
            f"\\newcommand{{\\QuestionOneFocalRatio}}{{{question_one_focal/sphere_radius:.6f}}}",
            f"\\newcommand{{\\QuestionOneMaxDisplacement}}{{{minimax_details['objective_max_radial_displacement_m']:.6f}}}",
            f"\\newcommand{{\\IdealFocalLength}}{{{selected_focal:.6f}}}",
            f"\\newcommand{{\\IdealFocalRatio}}{{{selected_focal/sphere_radius:.6f}}}",
            f"\\newcommand{{\\SelectedFocalOffset}}{{{selected_focal-nominal_focal_length:.6f}}}",
            f"\\newcommand{{\\VertexX}}{{{vertex[0]:.6f}}}",
            f"\\newcommand{{\\VertexY}}{{{vertex[1]:.6f}}}",
            f"\\newcommand{{\\VertexZ}}{{{vertex[2]:.6f}}}",
            f"\\newcommand{{\\ActiveNodeCount}}{{{int(np.sum(mask))}}}",
            f"\\newcommand{{\\StrokeMin}}{{{np.min(strokes[mask]):.6f}}}",
            f"\\newcommand{{\\StrokeMax}}{{{np.max(strokes[mask]):.6f}}}",
            f"\\newcommand{{\\EdgeStrainPct}}{{{100*np.max(np.abs(strains)):.6f}}}",
            f"\\newcommand{{\\RadialRMS}}{{{selected_metrics['radial_residual_rms_m']:.6f}}}",
            f"\\newcommand{{\\SurfaceRMS}}{{{np.sqrt(np.mean(surface_distances**2)):.6f}}}",
            f"\\newcommand{{\\SurfaceMax}}{{{np.max(np.abs(surface_distances)):.6f}}}",
            f"\\newcommand{{\\CurvedAdjustedReceptionPct}}{{{100*primary_ray['adjusted_ratio']:.3f}}}",
            f"\\newcommand{{\\CurvedBaselineReceptionPct}}{{{100*primary_ray['baseline_ratio']:.3f}}}",
            f"\\newcommand{{\\CurvedRelativeImprovementPct}}{{{100*primary_ray['relative_improvement']:.1f}}}",
            f"\\newcommand{{\\PlanarAdjustedReceptionPct}}{{{100*planar_ray['adjusted_ratio']:.3f}}}",
            f"\\newcommand{{\\PlanarBaselineReceptionPct}}{{{100*planar_ray['baseline_ratio']:.3f}}}",
            f"\\newcommand{{\\PlanarRelativeImprovementPct}}{{{100*planar_ray['relative_improvement']:.1f}}}",
            f"\\newcommand{{\\AdjustedReceptionPct}}{{{100*primary_ray['adjusted_ratio']:.3f}}}",
            f"\\newcommand{{\\BaselineReceptionPct}}{{{100*primary_ray['baseline_ratio']:.3f}}}",
            f"\\newcommand{{\\RelativeImprovementPct}}{{{100*primary_ray['relative_improvement']:.1f}}}",
        ]
    )
    (PAPER_DIR / "generated_numbers.tex").write_text(
        generated_tex + "\n", encoding="utf-8"
    )

    environment = {
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "numpy": np.__version__,
        "pandas": pd.__version__,
        "scipy": scipy.__version__,
        "matplotlib": matplotlib.__version__,
        "openpyxl": openpyxl.__version__,
        "seed": SEED,
        "command": "python code/solve.py",
    }
    write_json(RESULTS_DIR / "environment.json", environment)

    print(json.dumps(summary, ensure_ascii=False, indent=2, default=json_default))
    internal_ok = (
        audit["status"] == "pass"
        and joint_payload["joint_search"]["status"] == "pass"
        and solver_details["kkt_status"] == "pass"
        and selected_not_dominated == "pass"
        and selected_metrics["constraint_status"] == "pass"
        and validation["edge_strain"]["design_reserve_status"] == "pass"
        and rounded_output_status == "pass"
        and ideal_check["status"] == "pass"
        and overall_ray_numerical_status == "pass"
    )
    print(f"INTERNAL_STATUS={trinary(internal_ok)}")
    return 0 if internal_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
