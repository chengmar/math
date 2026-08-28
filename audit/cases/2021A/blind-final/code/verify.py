"""Independent rerun, constraint, document, and consistency checks."""

from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
import sys
import time
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "results"
PAPER = ROOT / "paper"
CHECK_PATH = RESULTS / "reproduction_check.json"

DETERMINISTIC_FILES = [
    "results/input_audit.json",
    "results/ideal_paraboloid.json",
    "results/summary.json",
    "results/validation_metrics.json",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/<SOURCE_FILE_REDACTED>",
    "results/focal_spot_metrics.json",
    "paper/generated_numbers.tex",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
]

REQUIRED_PATHS = [
    "problem-analysis.md",
    "data-audit.md",
    "assumptions.yaml",
    "variables.yaml",
    "model-selection.md",
    "solution-report.yaml",
    "reproducibility.yaml",
    "code/solve.py",
    "code/verify.py",
    "code/requirements.txt",
    "results/summary.json",
    "results/validation_metrics.json",
    "results/<SOURCE_FILE_REDACTED>",
    "results/focal_spot_metrics.json",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "figures/<SOURCE_FILE_REDACTED>",
    "paper/main.tex",
    "paper/paper.md",
    "paper/generated_numbers.tex",
    "<SOURCE_FILE_REDACTED>",
]


def trinary(condition: bool) -> str:
    return "pass" if bool(condition) else "fail"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def hashes() -> dict[str, str]:
    return {relative: sha256(ROOT / relative) for relative in DETERMINISTIC_FILES}


def run_solver() -> dict:
    start = time.perf_counter()
    completed = subprocess.run(
        [sys.executable, str(ROOT / "code" / "solve.py")],
        cwd=ROOT,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=900,
        check=False,
    )
    elapsed = time.perf_counter() - start
    return {
        "return_code": completed.returncode,
        "elapsed_seconds": elapsed,
        "status": trinary(completed.returncode == 0),
        "last_stdout_line": (
            completed.stdout.strip().splitlines()[-1]
            if completed.stdout.strip()
            else ""
        ),
        "stderr_tail": completed.stderr[-2000:],
    }


def workbook_checks(summary: dict) -> dict:
    workbook = openpyxl.load_workbook(ROOT / "<SOURCE_FILE_REDACTED>", read_only=True, data_only=True)
    vertex_sheet = workbook["理想抛物面顶点坐标"]
    node_sheet = workbook["调整后主索节点编号及坐标"]
    stroke_sheet = workbook["促动器顶端伸缩量"]
    vertex = np.array(
        [vertex_sheet.cell(2, column).value for column in range(1, 4)], dtype=float
    )
    expected_vertex = np.array(summary["question_2"]["vertex_m"], dtype=float)
    node_csv = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>", encoding="utf-8-sig")
    stroke_csv = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>", encoding="utf-8-sig")
    workbook_node_ids = [
        str(node_sheet.cell(row, 1).value)
        for row in range(2, node_sheet.max_row + 1)
    ]
    workbook_nodes = np.array(
        [
            [node_sheet.cell(row, column).value for column in range(2, 5)]
            for row in range(2, node_sheet.max_row + 1)
        ],
        dtype=float,
    )
    workbook_stroke_ids = [
        str(stroke_sheet.cell(row, 1).value)
        for row in range(2, stroke_sheet.max_row + 1)
    ]
    workbook_strokes = np.array(
        [stroke_sheet.cell(row, 2).value for row in range(2, stroke_sheet.max_row + 1)],
        dtype=float,
    )
    node_values = node_csv[["x_m", "y_m", "z_m"]].to_numpy(float)
    stroke_values = stroke_csv["actuator_stroke_m"].to_numpy(float)
    checks = {
        "vertex_max_abs_difference_m": float(np.max(np.abs(vertex - expected_vertex))),
        "node_row_count": int(len(workbook_node_ids)),
        "expected_node_row_count": int(summary["question_2"]["active_node_count"]),
        "node_id_match": workbook_node_ids == node_csv["node_id"].astype(str).tolist(),
        "node_coordinate_max_abs_difference_m": float(
            np.max(np.abs(workbook_nodes - node_values))
        ),
        "stroke_id_match": workbook_stroke_ids
        == stroke_csv["node_id"].astype(str).tolist(),
        "stroke_max_abs_difference_m": float(
            np.max(np.abs(workbook_strokes - stroke_values))
        ),
    }
    checks["status"] = trinary(
        checks["vertex_max_abs_difference_m"] <= 5.1e-7
        and checks["node_row_count"] == checks["expected_node_row_count"]
        and checks["node_id_match"]
        and checks["node_coordinate_max_abs_difference_m"] <= 1e-12
        and checks["stroke_id_match"]
        and checks["stroke_max_abs_difference_m"] <= 1e-12
    )
    return checks


def document_integrity_checks() -> dict:
    markdown_paths = sorted(ROOT.glob("*.md")) + sorted(PAPER.glob("*.md"))
    decode_failures = []
    control_locations = []
    for path in markdown_paths:
        try:
            text = path.read_bytes().decode("utf-8", errors="strict")
        except UnicodeDecodeError as exc:
            decode_failures.append(
                {"path": path.relative_to(ROOT).as_posix(), "error": str(exc)}
            )
            continue
        line = 1
        column = 0
        for index, character in enumerate(text):
            if character == "\n":
                line += 1
                column = 0
                continue
            column += 1
            if character == "\r":
                invalid = not (
                    index + 1 < len(text) and text[index + 1] == "\n"
                )
            else:
                invalid = ord(character) < 32 and character != "\t"
            if invalid:
                control_locations.append(
                    {
                        "path": path.relative_to(ROOT).as_posix(),
                        "line": line,
                        "column": column,
                        "codepoint": ord(character),
                    }
                )
    paper_markdown = (PAPER / "paper.md").read_text(encoding="utf-8")
    required_formula_tokens = [
        r"\boldsymbol",
        r"\varepsilon",
        r"\rho",
        r"\beta",
        r"\lambda",
    ]
    missing_formula_tokens = [
        token for token in required_formula_tokens if token not in paper_markdown
    ]
    return {
        "utf8_decode_failures": decode_failures,
        "unexpected_control_characters": control_locations,
        "unexpected_control_character_count": len(control_locations),
        "missing_formula_tokens": missing_formula_tokens,
        "status": trinary(
            not decode_failures and not control_locations and not missing_formula_tokens
        ),
    }


def paper_checks(summary: dict) -> dict:
    markdown = (PAPER / "paper.md").read_text(encoding="utf-8")
    tex = (PAPER / "main.tex").read_text(encoding="utf-8")
    generated = (PAPER / "generated_numbers.tex").read_text(encoding="utf-8")
    q2 = summary["question_2"]
    q3 = summary["question_3"]
    expected_markdown_claims = [
        f"{summary['question_1']['ideal_focal_length_m']:.6f}",
        f"{q2['focal_length_m']:.6f}",
        f"{q2['vertex_m'][0]:.6f}",
        f"{q2['vertex_m'][1]:.6f}",
        f"{q2['vertex_m'][2]:.6f}",
        str(q2["active_node_count"]),
        f"{100*q2['max_abs_edge_strain']:.6f}%",
        f"{100*q3['primary_adjusted_reception_ratio']:.3f}%",
        f"{100*q3['primary_baseline_reception_ratio']:.3f}%",
        f"{100*q3['primary_relative_improvement']:.1f}%",
        f"{100*q3['planar_adjusted_reception_ratio']:.3f}%",
        f"{100*q3['planar_baseline_reception_ratio']:.3f}%",
        f"{100*q3['planar_relative_improvement']:.1f}%",
    ]
    missing_markdown = [
        claim for claim in expected_markdown_claims if claim not in markdown
    ]
    macro_claims = [
        f"{summary['question_1']['ideal_focal_length_m']:.6f}",
        f"{q2['focal_length_m']:.6f}",
        f"{q2['vertex_m'][0]:.6f}",
        f"{100*q3['primary_adjusted_reception_ratio']:.3f}",
        f"{100*q3['primary_baseline_reception_ratio']:.3f}",
        f"{100*q3['planar_adjusted_reception_ratio']:.3f}",
    ]
    missing_macros = [claim for claim in macro_claims if claim not in generated]
    uses_generated = "\\input{generated_numbers}" in tex
    figure_references = [
        "../figures/<SOURCE_FILE_REDACTED>",
        "../figures/<SOURCE_FILE_REDACTED>",
        "../figures/<SOURCE_FILE_REDACTED>",
        "../figures/<SOURCE_FILE_REDACTED>",
        "../figures/<SOURCE_FILE_REDACTED>",
        "../figures/<SOURCE_FILE_REDACTED>",
    ]
    missing_figure_references = [
        reference for reference in figure_references if reference not in tex
    ]
    return {
        "markdown_missing_claims": missing_markdown,
        "tex_missing_generated_claims": missing_macros,
        "tex_uses_generated_numbers": uses_generated,
        "tex_missing_figure_references": missing_figure_references,
        "status": trinary(
            not missing_markdown
            and not missing_macros
            and uses_generated
            and not missing_figure_references
        ),
    }


def collect_status_values(value, path: str = "") -> list[dict]:
    records = []
    if isinstance(value, dict):
        for key, child in value.items():
            child_path = f"{path}.{key}" if path else key
            if key == "status" or key.endswith("_status"):
                records.append({"path": child_path, "value": child})
            records.extend(collect_status_values(child, child_path))
    elif isinstance(value, list):
        for index, child in enumerate(value):
            records.extend(collect_status_values(child, f"{path}[{index}]"))
    return records


def structured_status_checks(summary: dict, validation: dict) -> dict:
    allowed = {"pass", "fail", "needs_review"}
    records = collect_status_values(summary, "summary") + collect_status_values(
        validation, "validation"
    )
    invalid = [record for record in records if record["value"] not in allowed]
    return {
        "checked_status_field_count": len(records),
        "invalid_status_fields": invalid,
        "status": trinary(not invalid),
    }


def model_and_constraint_checks(validation: dict) -> dict:
    search = pd.read_csv(RESULTS / "<SOURCE_FILE_REDACTED>", encoding="utf-8-sig")
    spot = json.loads(
        (RESULTS / "focal_spot_metrics.json").read_text(encoding="utf-8")
    )
    focal_coverage = {
        "minimum_offset_m": float(search["focal_offset_from_nominal_m"].min()),
        "maximum_offset_m": float(search["focal_offset_from_nominal_m"].max()),
    }
    focal_coverage["status"] = trinary(
        focal_coverage["minimum_offset_m"] <= -0.599999
        and focal_coverage["maximum_offset_m"] >= 0.599999
    )
    spot_closure_max = max(
        abs(item["forward_power_closure_relative_error"]) for item in spot.values()
    )
    spot_status = trinary(
        spot_closure_max <= 2e-4
        and all(
            0.0 <= item["power_fraction_within_6m"] <= 1.0
            for item in spot.values()
        )
        and all(item["max_forward_hit_radius_m"] > 6.0 for item in spot.values())
    )
    required_passes = {
        "joint_search": validation["joint_focal_search"]["status"],
        "audit_anchor": validation["audit_counterexample_anchor"][
            "selected_not_dominated_status"
        ],
        "kkt": validation["solver"]["kkt_status"],
        "stroke": validation["stroke"]["status"],
        "edge": validation["edge_strain"]["status"],
        "edge_design_reserve": validation["edge_strain"][
            "design_reserve_status"
        ],
        "cable": validation["cable_closure"]["status"],
        "rounded": validation["rounded_output"]["status"],
        "ideal_ray": validation["ideal_ray_identity"]["status"],
        "ray_numerical": validation["ray_integration"]["numerical_status"],
        "focal_coverage": focal_coverage["status"],
        "spot_power_closure": spot_status,
    }
    return {
        "required_passes": required_passes,
        "focal_search_coverage": focal_coverage,
        "focal_spot_max_forward_power_closure_relative_error": spot_closure_max,
        "focal_spot_status": spot_status,
        "status": trinary(
            all(value == "pass" for value in required_passes.values())
        ),
    }


def main() -> int:
    first_run = run_solver()
    first_hashes = hashes() if first_run["status"] == "pass" else {}
    second_run = run_solver()
    second_hashes = hashes() if second_run["status"] == "pass" else {}
    hash_equal = bool(first_hashes) and first_hashes == second_hashes
    summary = json.loads((RESULTS / "summary.json").read_text(encoding="utf-8"))
    validation = json.loads(
        (RESULTS / "validation_metrics.json").read_text(encoding="utf-8")
    )
    workbook = workbook_checks(summary)
    documents = document_integrity_checks()
    paper = paper_checks(summary)
    statuses = structured_status_checks(summary, validation)
    constraints = model_and_constraint_checks(validation)
    missing_paths = [
        relative for relative in REQUIRED_PATHS if not (ROOT / relative).exists()
    ]
    required_status = trinary(not missing_paths)
    xelatex = shutil.which("xelatex")
    latexmk = shutil.which("latexmk")
    latex_status = "needs_review"
    overall = trinary(
        first_run["status"] == "pass"
        and second_run["status"] == "pass"
        and hash_equal
        and workbook["status"] == "pass"
        and documents["status"] == "pass"
        and paper["status"] == "pass"
        and statuses["status"] == "pass"
        and constraints["status"] == "pass"
        and required_status == "pass"
    )
    payload = {
        "overall_status": overall,
        "runs": [first_run, second_run],
        "deterministic_reproduction": {
            "status": trinary(hash_equal),
            "file_count": len(DETERMINISTIC_FILES),
            "first_run_sha256": first_hashes,
            "second_run_sha256": second_hashes,
        },
        "workbook_consistency": workbook,
        "document_integrity": documents,
        "paper_result_consistency": paper,
        "structured_status_vocabulary": statuses,
        "model_and_constraint_validation": constraints,
        "required_deliverables": {
            "status": required_status,
            "missing_paths": missing_paths,
        },
        "latex": {
            "xelatex_available": bool(xelatex),
            "latexmk_available": bool(latexmk),
            "compilation_status": latex_status,
            "note": (
                "No TeX compiler is available in this environment; compilation and page-level inspection remain needs_review."
                if not (xelatex or latexmk)
                else "A compiler is available, but this verifier does not claim page-level visual inspection."
            ),
        },
        "execution_reproducibility_status": trinary(hash_equal),
        "mathematical_correctness_status": "needs_review",
        "panel_geometry_interpretation_status": "needs_review",
        "physical_external_validity_status": "needs_review",
        "notes": [
            "Binary XLSX hashes are not compared because ZIP timestamps may vary; cell values are compared instead.",
            "Deterministic reproduction and KKT residuals do not constitute a formal global proof or external physical validation.",
        ],
    }
    CHECK_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if overall == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
