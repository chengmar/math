"""Deterministic physical models for the 2019A blind solution.

All lengths are in mm, time in ms, pressure in MPa, density in mg/mm^3,
and volume flow in mm^3/ms.  With this unit system the orifice formula in
the problem statement can be evaluated directly.
"""

from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
import math
import re
from typing import Iterable, Literal

import numpy as np
from openpyxl import load_workbook
from scipy.integrate import cumulative_trapezoid
from scipy.interpolate import CubicSpline, PchipInterpolator


SEED = 2019
TWO_PI = 2.0 * math.pi


@dataclass(frozen=True)
class Constants:
    flow_coefficient: float = 0.85
    rail_length: float = 500.0
    rail_diameter: float = 10.0
    inlet_diameter: float = 1.4
    nozzle_hole_diameter: float = 1.4
    needle_diameter: float = 2.5
    seat_half_angle_deg: float = 9.0
    plunger_diameter: float = 5.0
    residual_volume: float = 20.0
    pump_supply_pressure: float = 160.0
    pump_fill_pressure: float = 0.5
    ambient_pressure: float = 0.0
    reference_pressure: float = 100.0
    reference_density: float = 0.850
    injection_period: float = 100.0
    valve_closed_time: float = 10.0

    @property
    def rail_volume(self) -> float:
        return self.rail_length * math.pi * (self.rail_diameter / 2.0) ** 2

    @property
    def inlet_area(self) -> float:
        return math.pi * (self.inlet_diameter / 2.0) ** 2

    @property
    def nozzle_hole_area(self) -> float:
        return math.pi * (self.nozzle_hole_diameter / 2.0) ** 2

    @property
    def plunger_area(self) -> float:
        return math.pi * (self.plunger_diameter / 2.0) ** 2


CONSTS = Constants()


def file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@dataclass
class InputTables:
    elasticity_path: Path
    cam_path: Path
    needle_path: Path
    pressure: np.ndarray
    elasticity: np.ndarray
    cam_angle: np.ndarray
    cam_radius: np.ndarray
    needle_time: np.ndarray
    needle_lift: np.ndarray
    needle_intervals: list[tuple[float, float, float]]


def _sheet_rows(path: Path) -> list[tuple]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    return list(workbook["Sheet1"].iter_rows(values_only=True))


def discover_input_tables(data_dir: Path) -> InputTables:
    """Identify the three workbooks by their headers, not opaque filenames."""
    found: dict[str, Path] = {}
    for path in sorted(data_dir.glob("*.xlsx")):
        rows = _sheet_rows(path)
        header = str(rows[0][0]) if rows and rows[0] else ""
        if header.startswith("压力"):
            found["elasticity"] = path
        elif header.startswith("极角"):
            found["cam"] = path
        elif header.startswith("时间"):
            found["needle"] = path
    if set(found) != {"elasticity", "cam", "needle"}:
        raise ValueError(f"Could not uniquely identify three attachments: {found}")

    elastic_rows = _sheet_rows(found["elasticity"])[1:]
    elastic = np.asarray(
        [[float(row[0]), float(row[1])] for row in elastic_rows], dtype=float
    )
    cam_rows = _sheet_rows(found["cam"])[1:]
    cam = np.asarray([[float(row[0]), float(row[1])] for row in cam_rows], dtype=float)

    needle_rows = _sheet_rows(found["needle"])[1:]
    numeric_points: list[tuple[float, float]] = []
    intervals: list[tuple[float, float, float]] = []
    interval_pattern = re.compile(
        r"^\[\s*([-+]?\d+(?:\.\d+)?)\s*,\s*([-+]?\d+(?:\.\d+)?)\s*\]$"
    )
    for row in needle_rows:
        for time_col, lift_col in ((0, 1), (3, 4)):
            time_value, lift_value = row[time_col], row[lift_col]
            if isinstance(time_value, (int, float)) and isinstance(lift_value, (int, float)):
                numeric_points.append((float(time_value), float(lift_value)))
            elif isinstance(time_value, str) and isinstance(lift_value, (int, float)):
                match = interval_pattern.match(time_value)
                if not match:
                    raise ValueError(f"Unrecognised needle interval: {time_value!r}")
                left, right = float(match.group(1)), float(match.group(2))
                intervals.append((left, right, float(lift_value)))
                numeric_points.extend([(left, float(lift_value)), (right, float(lift_value))])
    # If a numeric endpoint and an interval endpoint coincide, the interval value is authoritative.
    point_map: dict[float, float] = {time: lift for time, lift in numeric_points}
    for left, right, lift in intervals:
        point_map[left] = lift
        point_map[right] = lift
    needle = np.asarray(sorted(point_map.items()), dtype=float)

    return InputTables(
        elasticity_path=found["elasticity"],
        cam_path=found["cam"],
        needle_path=found["needle"],
        pressure=elastic[:, 0],
        elasticity=elastic[:, 1],
        cam_angle=cam[:, 0],
        cam_radius=cam[:, 1],
        needle_time=needle[:, 0],
        needle_lift=needle[:, 1],
        needle_intervals=intervals,
    )


class FuelProperties:
    """Elastic modulus and density obtained only from Attachment 3.

    The statement gives dP = E(P) d rho / rho and rho(100 MPa)=0.850.
    Density is therefore reconstructed by numerical integration of 1/E.
    """

    def __init__(self, pressure: np.ndarray, elasticity: np.ndarray):
        self.pressure = np.asarray(pressure, dtype=float)
        self.elasticity = np.asarray(elasticity, dtype=float)
        if len(self.pressure) < 2 or np.any(np.diff(self.pressure) <= 0):
            raise ValueError("Pressure table must be strictly increasing")
        self.integral = cumulative_trapezoid(
            1.0 / self.elasticity, self.pressure, initial=0.0
        )
        anchor = float(np.interp(CONSTS.reference_pressure, self.pressure, self.integral))
        self.density = CONSTS.reference_density * np.exp(self.integral - anchor)
        self.p_min = float(self.pressure[0])
        self.p_max = float(self.pressure[-1])
        differences = np.diff(self.pressure)
        self.uniform = bool(np.max(np.abs(differences - differences[0])) < 1e-12)
        self.dp = float(differences[0])

    def _scalar_lerp(self, values: np.ndarray, pressure: float) -> float:
        if not (self.p_min <= pressure <= self.p_max):
            raise ValueError(f"Pressure {pressure} MPa outside property table")
        if not self.uniform:
            return float(np.interp(pressure, self.pressure, values))
        coordinate = (pressure - self.p_min) / self.dp
        index = int(math.floor(coordinate))
        if index >= len(values) - 1:
            return float(values[-1])
        fraction = coordinate - index
        return float(values[index] + fraction * (values[index + 1] - values[index]))

    def e(self, pressure: float) -> float:
        return self._scalar_lerp(self.elasticity, pressure)

    def rho(self, pressure: float) -> float:
        return self._scalar_lerp(self.density, pressure)

    def e_array(self, pressure: np.ndarray) -> np.ndarray:
        return np.interp(pressure, self.pressure, self.elasticity)

    def rho_array(self, pressure: np.ndarray) -> np.ndarray:
        return np.interp(pressure, self.pressure, self.density)

    def compression_integral(self, pressure: float) -> float:
        """Return integral_100^P 1/E(p) dp for rail-balance checks."""
        value = float(np.interp(pressure, self.pressure, self.integral))
        anchor = float(np.interp(CONSTS.reference_pressure, self.pressure, self.integral))
        return value - anchor


class NeedleModel:
    def __init__(self, time: np.ndarray, lift: np.ndarray):
        self.time = np.asarray(time, dtype=float)
        self.lift_values = np.asarray(lift, dtype=float)
        self.interpolator = PchipInterpolator(self.time, self.lift_values, extrapolate=False)

    def lift(self, phase: float | np.ndarray) -> float | np.ndarray:
        values = np.asarray(phase)
        result = np.zeros_like(values, dtype=float)
        mask = (values >= self.time[0]) & (values <= self.time[-1])
        if np.any(mask):
            result[mask] = np.maximum(0.0, self.interpolator(values[mask]))
        return float(result) if result.ndim == 0 else result

    def effective_area(
        self, phase: float | np.ndarray, mode: Literal["bottleneck", "series"] = "bottleneck"
    ) -> float | np.ndarray:
        lift = np.asarray(self.lift(phase), dtype=float)
        seat_area = (
            math.pi
            * CONSTS.needle_diameter
            * lift
            * math.sin(math.radians(CONSTS.seat_half_angle_deg))
        )
        hole_area = CONSTS.nozzle_hole_area
        if mode == "bottleneck":
            result = np.minimum(seat_area, hole_area)
        elif mode == "series":
            result = np.zeros_like(seat_area)
            mask = seat_area > 0
            result[mask] = 1.0 / np.sqrt(
                1.0 / seat_area[mask] ** 2 + 1.0 / hole_area**2
            )
        else:
            raise ValueError(f"Unknown nozzle area mode: {mode}")
        return float(result) if result.ndim == 0 else result


class CamModel:
    def __init__(self, angle: np.ndarray, radius: np.ndarray):
        angle = np.asarray(angle, dtype=float)
        radius = np.asarray(radius, dtype=float)
        if angle[0] != 0.0 or np.any(np.diff(angle) <= 0):
            raise ValueError("Cam angle must start at zero and increase")
        closed_angle = np.append(angle, TWO_PI)
        closed_radius = np.append(radius, radius[0])
        self.spline = CubicSpline(closed_angle, closed_radius, bc_type="periodic")
        self.derivative = self.spline.derivative()
        self.radius_max = float(np.max(radius))
        self.radius_min = float(np.min(radius))
        self.stroke = self.radius_max - self.radius_min

    def geometry(self, theta: np.ndarray, omega: float) -> tuple[np.ndarray, np.ndarray]:
        theta = np.mod(theta, TWO_PI)
        radius = self.spline(theta)
        dr_dtheta = self.derivative(theta)
        volume = CONSTS.residual_volume + CONSTS.plunger_area * (self.radius_max - radius)
        dvolume_dt = -CONSTS.plunger_area * dr_dtheta * omega
        return np.asarray(volume), np.asarray(dvolume_dt)


@dataclass
class Models:
    tables: InputTables
    fuel: FuelProperties
    needle: NeedleModel
    cam: CamModel


def load_models(workspace: Path) -> Models:
    tables = discover_input_tables(workspace / "input" / "data")
    return Models(
        tables=tables,
        fuel=FuelProperties(tables.pressure, tables.elasticity),
        needle=NeedleModel(tables.needle_time, tables.needle_lift),
        cam=CamModel(tables.cam_angle, tables.cam_radius),
    )


def orifice_flow(
    high_pressure: float,
    low_pressure: float,
    area: float,
    fuel: FuelProperties,
) -> float:
    if area <= 0.0 or high_pressure <= low_pressure:
        return 0.0
    return CONSTS.flow_coefficient * area * math.sqrt(
        2.0 * (high_pressure - low_pressure) / fuel.rho(high_pressure)
    )


def prescribed_injection_rate(phase: float | np.ndarray) -> float | np.ndarray:
    """Figure 2: 0--20 ramp, plateau, then 20--0 ramp in 2.4 ms."""
    x = np.mod(np.asarray(phase, dtype=float), CONSTS.injection_period)
    result = np.zeros_like(x)
    rise = (x >= 0.0) & (x < 0.2)
    plateau = (x >= 0.2) & (x <= 2.2)
    fall = (x > 2.2) & (x <= 2.4)
    result[rise] = 100.0 * x[rise]
    result[plateau] = 20.0
    result[fall] = 100.0 * (2.4 - x[fall])
    return float(result) if result.ndim == 0 else result


def _valve_open_at(
    time: np.ndarray,
    tau_before: float,
    transition_time: float | None,
    tau_after: float | None,
) -> np.ndarray:
    time = np.asarray(time, dtype=float)
    if transition_time is None:
        return np.mod(time, tau_before + CONSTS.valve_closed_time) < tau_before
    if tau_after is None:
        raise ValueError("tau_after is required with transition_time")
    result = np.empty(time.shape, dtype=bool)
    before = time < transition_time
    result[before] = (
        np.mod(time[before], tau_before + CONSTS.valve_closed_time) < tau_before
    )
    local = time[~before] - transition_time
    result[~before] = np.mod(local, tau_after + CONSTS.valve_closed_time) < tau_after
    return result


def build_valve_schedule(
    duration: float,
    tau_command,
) -> tuple[np.ndarray, np.ndarray]:
    """Build sequential open/closed cycles for a time-varying duration command.

    ``tau_command`` is evaluated once at the start of each cycle.  The valve is
    open for the returned duration and then closed for exactly 10 ms.
    """
    starts: list[float] = []
    durations: list[float] = []
    cycle_start = 0.0
    while cycle_start <= duration + 1e-12:
        tau = float(tau_command(cycle_start))
        if not (0.0 < tau < 10.0):
            raise ValueError(f"Invalid valve duration {tau} ms at t={cycle_start} ms")
        starts.append(cycle_start)
        durations.append(tau)
        cycle_start += tau + CONSTS.valve_closed_time
    return np.asarray(starts), np.asarray(durations)


def valve_open_from_schedule(
    time: np.ndarray,
    starts: np.ndarray,
    durations: np.ndarray,
) -> np.ndarray:
    time = np.asarray(time, dtype=float)
    indices = np.searchsorted(starts, time, side="right") - 1
    result = np.zeros(time.shape, dtype=bool)
    mask = indices >= 0
    result[mask] = time[mask] < starts[indices[mask]] + durations[indices[mask]]
    return result


@dataclass
class RailSimulation:
    time: np.ndarray
    pressure: np.ndarray
    pump_flow: np.ndarray
    injection_flow: np.ndarray
    relief_flow: np.ndarray
    chamber_pressure: np.ndarray | None = None
    relief_open: np.ndarray | None = None
    interval_pump_flow: np.ndarray | None = None
    interval_injection_flow: np.ndarray | None = None
    interval_relief_flow: np.ndarray | None = None


def simulate_problem1(
    fuel: FuelProperties,
    duration: float,
    tau_before: float,
    *,
    initial_pressure: float = 100.0,
    dt: float = 0.02,
    transition_time: float | None = None,
    tau_after: float | None = None,
    valve_schedule: tuple[np.ndarray, np.ndarray] | None = None,
) -> RailSimulation:
    """Explicit midpoint integration of the scalar rail-pressure equation."""
    steps = int(math.ceil(duration / dt))
    dt = duration / steps
    time = np.linspace(0.0, duration, steps + 1)
    midpoint = time[:-1] + 0.5 * dt
    if valve_schedule is None:
        open_start = _valve_open_at(time[:-1], tau_before, transition_time, tau_after)
        open_mid = _valve_open_at(midpoint, tau_before, transition_time, tau_after)
    else:
        starts, durations = valve_schedule
        open_start = valve_open_from_schedule(time[:-1], starts, durations)
        open_mid = valve_open_from_schedule(midpoint, starts, durations)
    qout_start = np.asarray(prescribed_injection_rate(time[:-1]))
    qout_mid = np.asarray(prescribed_injection_rate(midpoint))

    pressure = np.empty(steps + 1, dtype=float)
    pressure[0] = initial_pressure
    interval_pump = np.empty(steps, dtype=float)
    interval_injection = qout_mid.copy()
    rho_supply = fuel.rho(CONSTS.pump_supply_pressure)
    coefficient = CONSTS.flow_coefficient * CONSTS.inlet_area

    def derivative(p: float, valve_open: bool, qout: float) -> float:
        qin = 0.0
        if valve_open and CONSTS.pump_supply_pressure > p:
            qin = coefficient * math.sqrt(
                2.0 * (CONSTS.pump_supply_pressure - p) / rho_supply
            )
        return fuel.e(p) * (qin - qout) / CONSTS.rail_volume

    for index in range(steps):
        p = pressure[index]
        k1 = derivative(p, bool(open_start[index]), float(qout_start[index]))
        p_mid = p + 0.5 * dt * k1
        k2 = derivative(p_mid, bool(open_mid[index]), float(qout_mid[index]))
        interval_pump[index] = 0.0
        if open_mid[index] and CONSTS.pump_supply_pressure > p_mid:
            interval_pump[index] = coefficient * math.sqrt(
                2.0 * (CONSTS.pump_supply_pressure - p_mid) / rho_supply
            )
        p_next = p + dt * k2
        if not (fuel.p_min < p_next < fuel.p_max):
            raise RuntimeError(f"Problem 1 pressure left property range at t={time[index + 1]}")
        pressure[index + 1] = p_next

    pump_flow = np.zeros_like(time)
    if valve_schedule is None:
        all_open = _valve_open_at(time, tau_before, transition_time, tau_after)
    else:
        all_open = valve_open_from_schedule(time, starts, durations)
    mask = all_open & (pressure < CONSTS.pump_supply_pressure)
    pump_flow[mask] = coefficient * np.sqrt(
        2.0 * (CONSTS.pump_supply_pressure - pressure[mask]) / rho_supply
    )
    injection_flow = np.asarray(prescribed_injection_rate(time))
    return RailSimulation(
        time=time,
        pressure=pressure,
        pump_flow=pump_flow,
        injection_flow=injection_flow,
        relief_flow=np.zeros_like(time),
        interval_pump_flow=interval_pump,
        interval_injection_flow=interval_injection,
        interval_relief_flow=np.zeros(steps, dtype=float),
    )


def nozzle_flow(
    pressure: float,
    area: float,
    fuel: FuelProperties,
    ambient_pressure: float = CONSTS.ambient_pressure,
) -> float:
    return orifice_flow(pressure, ambient_pressure, area, fuel)


def needle_pulse_volume(
    models: Models,
    pressure: float = 100.0,
    area_mode: Literal["bottleneck", "series"] = "bottleneck",
    dt: float = 0.0005,
) -> float:
    end = 2.46
    time = np.arange(0.0, end + 0.5 * dt, dt)
    areas = models.needle.effective_area(time, area_mode)
    coefficient = (
        CONSTS.flow_coefficient
        * math.sqrt(2.0 * (pressure - CONSTS.ambient_pressure) / models.fuel.rho(pressure))
    )
    return float(np.trapezoid(coefficient * areas, time))


def plunger_delivery_volume(models: Models, rail_pressure: float = 100.0) -> float:
    bottom_volume = CONSTS.residual_volume + CONSTS.plunger_area * models.cam.stroke
    return (
        bottom_volume
        * models.fuel.rho(CONSTS.pump_fill_pressure)
        / models.fuel.rho(rail_pressure)
        - CONSTS.residual_volume
    )


def mean_balance_omega(
    models: Models,
    injector_count: int = 1,
    rail_pressure: float = 100.0,
    area_mode: Literal["bottleneck", "series"] = "bottleneck",
) -> float:
    pulse_volume = needle_pulse_volume(models, rail_pressure, area_mode)
    delivery = plunger_delivery_volume(models, rail_pressure)
    cycles_per_ms = injector_count * pulse_volume / CONSTS.injection_period / delivery
    return TWO_PI * cycles_per_ms


def _flow_terms(
    rail_pressure: float,
    chamber_pressure: float,
    chamber_volume: float,
    chamber_dvdt: float,
    injection_area: float,
    relief_is_open: bool,
    models: Models,
) -> tuple[float, float, float, float, float]:
    fuel = models.fuel
    compression = chamber_dvdt < -1e-10
    pump_flow = 0.0
    if compression and chamber_pressure > rail_pressure:
        pump_flow = orifice_flow(
            chamber_pressure, rail_pressure, CONSTS.inlet_area, fuel
        )
    injection_flow = nozzle_flow(rail_pressure, injection_area, fuel)
    relief_flow = (
        nozzle_flow(rail_pressure, CONSTS.inlet_area, fuel)
        if relief_is_open
        else 0.0
    )
    rail_derivative = (
        fuel.e(rail_pressure)
        * (pump_flow - injection_flow - relief_flow)
        / CONSTS.rail_volume
    )
    if compression:
        chamber_derivative = (
            -fuel.e(chamber_pressure)
            * (pump_flow + chamber_dvdt)
            / chamber_volume
        )
    else:
        chamber_derivative = 0.0
    return rail_derivative, chamber_derivative, pump_flow, injection_flow, relief_flow


def simulate_pump_system(
    models: Models,
    omega: float,
    duration: float,
    *,
    injector_offsets: Iterable[float] = (0.0,),
    initial_pressure: float = 100.0,
    dt: float = 0.01,
    area_mode: Literal["bottleneck", "series"] = "bottleneck",
    cam_phase0: float = math.pi,
    relief_thresholds: tuple[float, float] | None = None,
) -> RailSimulation:
    """Coupled plunger-chamber/rail model using explicit midpoint integration.

    ``relief_thresholds`` is (close_pressure, open_pressure).  The valve uses
    hysteresis and its state is held over one numerical step.
    """
    if relief_thresholds is not None:
        close_pressure, open_pressure = relief_thresholds
        if close_pressure >= open_pressure:
            raise ValueError("Relief close threshold must be below open threshold")
    steps = int(math.ceil(duration / dt))
    dt = duration / steps
    time = np.linspace(0.0, duration, steps + 1)
    midpoint = time[:-1] + 0.5 * dt
    theta_start = cam_phase0 + omega * time[:-1]
    theta_mid = cam_phase0 + omega * midpoint
    volume_start, dvdt_start = models.cam.geometry(theta_start, omega)
    volume_mid, dvdt_mid = models.cam.geometry(theta_mid, omega)

    offsets = tuple(float(value) for value in injector_offsets)
    area_start = np.zeros(steps, dtype=float)
    area_mid = np.zeros(steps, dtype=float)
    for offset in offsets:
        phase_start = np.mod(time[:-1] - offset, CONSTS.injection_period)
        phase_mid = np.mod(midpoint - offset, CONSTS.injection_period)
        area_start += models.needle.effective_area(phase_start, area_mode)
        area_mid += models.needle.effective_area(phase_mid, area_mode)

    rail = np.empty(steps + 1, dtype=float)
    chamber = np.empty(steps + 1, dtype=float)
    relief_open = np.zeros(steps + 1, dtype=bool)
    interval_pump = np.empty(steps, dtype=float)
    interval_injection = np.empty(steps, dtype=float)
    interval_relief = np.empty(steps, dtype=float)
    rail[0] = initial_pressure
    chamber[0] = CONSTS.pump_fill_pressure

    for index in range(steps):
        p = rail[index]
        pc = chamber[index]
        if dvdt_start[index] >= -1e-10:
            pc = CONSTS.pump_fill_pressure
            chamber[index] = pc

        is_open = bool(relief_open[index])
        if relief_thresholds is not None:
            if is_open and p <= close_pressure:
                is_open = False
            elif (not is_open) and p >= open_pressure:
                is_open = True
        relief_open[index] = is_open

        k1_p, k1_pc, _, _, _ = _flow_terms(
            p,
            pc,
            float(volume_start[index]),
            float(dvdt_start[index]),
            float(area_start[index]),
            is_open,
            models,
        )
        p_mid = p + 0.5 * dt * k1_p
        pc_mid = pc + 0.5 * dt * k1_pc
        if dvdt_mid[index] >= -1e-10:
            pc_mid = CONSTS.pump_fill_pressure
        k2_p, k2_pc, q_pump_mid, q_injection_mid, q_relief_mid = _flow_terms(
            p_mid,
            pc_mid,
            float(volume_mid[index]),
            float(dvdt_mid[index]),
            float(area_mid[index]),
            is_open,
            models,
        )
        p_next = p + dt * k2_p
        pc_next = pc + dt * k2_pc
        if dvdt_mid[index] >= -1e-10:
            pc_next = CONSTS.pump_fill_pressure
        pc_next = max(CONSTS.pump_fill_pressure, pc_next)
        if not (models.fuel.p_min < p_next < models.fuel.p_max):
            raise RuntimeError(f"Rail pressure left property range at t={time[index + 1]}")
        if not (models.fuel.p_min <= pc_next <= models.fuel.p_max):
            raise RuntimeError(
                f"Chamber pressure left property range at t={time[index + 1]}: {pc_next}"
            )
        rail[index + 1] = p_next
        chamber[index + 1] = pc_next
        relief_open[index + 1] = is_open
        interval_pump[index] = q_pump_mid
        interval_injection[index] = q_injection_mid
        interval_relief[index] = q_relief_mid

    # Reconstruct diagnostic flows at stored nodes.
    theta_all = cam_phase0 + omega * time
    volume_all, dvdt_all = models.cam.geometry(theta_all, omega)
    total_area = np.zeros_like(time)
    for offset in offsets:
        phase = np.mod(time - offset, CONSTS.injection_period)
        total_area += models.needle.effective_area(phase, area_mode)
    pump_flow = np.zeros_like(time)
    injection_flow = np.zeros_like(time)
    relief_flow = np.zeros_like(time)
    for index in range(steps + 1):
        if dvdt_all[index] < -1e-10 and chamber[index] > rail[index]:
            pump_flow[index] = orifice_flow(
                chamber[index], rail[index], CONSTS.inlet_area, models.fuel
            )
        injection_flow[index] = nozzle_flow(
            rail[index], float(total_area[index]), models.fuel
        )
        if relief_open[index]:
            relief_flow[index] = nozzle_flow(
                rail[index], CONSTS.inlet_area, models.fuel
            )
    return RailSimulation(
        time=time,
        pressure=rail,
        pump_flow=pump_flow,
        injection_flow=injection_flow,
        relief_flow=relief_flow,
        chamber_pressure=chamber,
        relief_open=relief_open,
        interval_pump_flow=interval_pump,
        interval_injection_flow=interval_injection,
        interval_relief_flow=interval_relief,
    )


def pressure_metrics(
    simulation: RailSimulation,
    target: float,
    start: float = 0.0,
    end: float | None = None,
) -> dict[str, float]:
    if end is None:
        end = float(simulation.time[-1])
    mask = (simulation.time >= start) & (simulation.time <= end)
    values = simulation.pressure[mask]
    if not np.any(mask):
        raise ValueError("Empty metric window")
    error = values - target
    return {
        "window_start_ms": float(start),
        "window_end_ms": float(end),
        "mean_mpa": float(np.mean(values)),
        "std_mpa": float(np.std(values)),
        "rmse_mpa": float(np.sqrt(np.mean(error**2))),
        "mae_mpa": float(np.mean(np.abs(error))),
        "min_mpa": float(np.min(values)),
        "max_mpa": float(np.max(values)),
        "peak_to_peak_mpa": float(np.ptp(values)),
        "max_abs_error_mpa": float(np.max(np.abs(error))),
    }


def rail_balance_residual(
    simulation: RailSimulation,
    fuel: FuelProperties,
) -> dict[str, float]:
    if simulation.interval_pump_flow is not None:
        net_flow = (
            simulation.interval_pump_flow
            - simulation.interval_injection_flow
            - simulation.interval_relief_flow
        )
        cumulative = np.zeros_like(simulation.time)
        cumulative[1:] = np.cumsum(net_flow * np.diff(simulation.time))
        throughput = max(
            float(np.sum(np.abs(net_flow) * np.diff(simulation.time))), 1e-12
        )
    else:
        net_flow = simulation.pump_flow - simulation.injection_flow - simulation.relief_flow
        cumulative = cumulative_trapezoid(net_flow, simulation.time, initial=0.0)
        throughput = max(float(np.trapezoid(np.abs(net_flow), simulation.time)), 1e-12)
    state = CONSTS.rail_volume * np.asarray(
        [
            fuel.compression_integral(float(p))
            - fuel.compression_integral(float(simulation.pressure[0]))
            for p in simulation.pressure
        ]
    )
    residual = state - cumulative
    return {
        "max_abs_mm3": float(np.max(np.abs(residual))),
        "terminal_abs_mm3": float(abs(residual[-1])),
        "max_relative_to_throughput": float(np.max(np.abs(residual)) / throughput),
    }


def downsample_simulation(simulation: RailSimulation, stride: int) -> np.ndarray:
    indices = np.arange(0, len(simulation.time), stride, dtype=int)
    if indices[-1] != len(simulation.time) - 1:
        indices = np.append(indices, len(simulation.time) - 1)
    chamber = (
        simulation.chamber_pressure[indices]
        if simulation.chamber_pressure is not None
        else np.full(len(indices), np.nan)
    )
    relief_open = (
        simulation.relief_open[indices].astype(int)
        if simulation.relief_open is not None
        else np.zeros(len(indices), dtype=int)
    )
    return np.column_stack(
        [
            simulation.time[indices],
            simulation.pressure[indices],
            chamber,
            simulation.pump_flow[indices],
            simulation.injection_flow[indices],
            simulation.relief_flow[indices],
            relief_open,
        ]
    )
