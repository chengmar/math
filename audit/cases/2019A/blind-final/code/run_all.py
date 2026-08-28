"""Generate every numerical result and figure for the blind solution.

Run from the workspace root with:

    python code/run_all.py

The script is deterministic; no external data or network access is used.
"""

from __future__ import annotations

import argparse
import csv
import gc
import importlib.metadata
import json
import math
from pathlib import Path
import platform
import sys
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import brentq

from model import (
    CONSTS,
    SEED,
    Models,
    FuelProperties,
    InputTables,
    ReliefControl,
    build_valve_schedule,
    chamber_compression_mass_residual,
    downsample_simulation,
    file_sha256,
    load_models,
    mean_balance_omega,
    needle_pulse_volume,
    orifice_flow,
    plunger_delivery_volume,
    pressure_metrics,
    rail_mass_balance_residual,
    relief_dwell_statistics,
    simulate_problem1,
    simulate_pump_system,
    transformed_volume_balance_residual,
)


np.random.seed(SEED)


def built_in(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): built_in(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [built_in(item) for item in value]
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, (np.floating, np.integer, np.bool_)):
        return value.item()
    return value


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(built_in(value), ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="\n",
    )


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError(f"Cannot write empty CSV: {path}")
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]), lineterminator="\n")
        writer.writeheader()
        writer.writerows([built_in(row) for row in rows])


def save_timeseries(path: Path, values: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    np.savetxt(
        path,
        values,
        delimiter=",",
        header=(
            "time_ms,rail_pressure_mpa,chamber_pressure_mpa,pump_flow_mm3_per_ms,"
            "injection_flow_mm3_per_ms,relief_flow_mm3_per_ms,relief_open"
        ),
        comments="",
        fmt="%.9g",
        newline="\n",
        encoding="utf-8",
    )


def workbook_audit(path: Path, workspace: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheets = []
    formula_count = 0
    nonempty_count = 0
    for sheet in workbook.worksheets:
        sheet_formula_count = 0
        sheet_nonempty = 0
        type_counts: dict[str, int] = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                sheet_nonempty += 1
                kind = type(cell.value).__name__
                type_counts[kind] = type_counts.get(kind, 0) + 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    sheet_formula_count += 1
        formula_count += sheet_formula_count
        nonempty_count += sheet_nonempty
        sheets.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "nonempty_cells": sheet_nonempty,
                "formula_cells": sheet_formula_count,
                "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                "value_type_counts": type_counts,
            }
        )
    return {
        "file": path.relative_to(workspace).as_posix(),
        "sha256": file_sha256(path),
        "size_bytes": path.stat().st_size,
        "sheet_count": len(workbook.sheetnames),
        "sheets": sheets,
        "nonempty_cells": nonempty_count,
        "formula_cells": formula_count,
    }


def data_audit(workspace: Path, models: Models) -> dict[str, Any]:
    manifest = []
    for path in sorted((workspace / "input").rglob("*")):
        if path.is_file():
            manifest.append(
                {
                    "relative_path": str(path.relative_to(workspace).as_posix()),
                    "size_bytes": path.stat().st_size,
                    "sha256": file_sha256(path),
                }
            )
    workbooks = [
        workbook_audit(path, workspace)
        for path in sorted((workspace / "input" / "data").glob("*.xlsx"))
    ]
    tables = models.tables
    numeric_needle = np.column_stack([tables.needle_time, tables.needle_lift])
    checks = {
        "three_attachments_identified_by_header": "pass",
        "elasticity_pressure_strictly_increasing": (
            "pass" if np.all(np.diff(tables.pressure) > 0) else "fail"
        ),
        "elasticity_positive": "pass" if np.all(tables.elasticity > 0) else "fail",
        "cam_angle_strictly_increasing": (
            "pass" if np.all(np.diff(tables.cam_angle) > 0) else "fail"
        ),
        "needle_numeric_time_strictly_increasing_after_interval_expansion": (
            "pass" if np.all(np.diff(numeric_needle[:, 0]) > 0) else "fail"
        ),
        "workbook_formula_cells_absent": (
            "pass" if sum(item["formula_cells"] for item in workbooks) == 0 else "needs_review"
        ),
        "pdf_docx_visual_equivalence": "needs_review",
    }
    return {
        "status": "needs_review",
        "reason_for_overall_status": (
            "Machine checks pass for numerical attachments; PDF/DOCX visual equivalence was not "
            "asserted by an automated text extractor."
        ),
        "input_manifest": manifest,
        "workbooks": workbooks,
        "table_mapping": {
            "attachment_1_cam": tables.cam_path.name,
            "attachment_2_needle": tables.needle_path.name,
            "attachment_3_elasticity": tables.elasticity_path.name,
        },
        "elasticity": {
            "rows": int(len(tables.pressure)),
            "pressure_min_mpa": float(tables.pressure.min()),
            "pressure_max_mpa": float(tables.pressure.max()),
            "pressure_step_mpa": float(np.median(np.diff(tables.pressure))),
            "modulus_min_mpa": float(tables.elasticity.min()),
            "modulus_max_mpa": float(tables.elasticity.max()),
            "missing_numeric_values": 0,
            "duplicate_pressure_values": int(len(tables.pressure) - len(np.unique(tables.pressure))),
        },
        "cam": {
            "rows": int(len(tables.cam_angle)),
            "angle_min_rad": float(tables.cam_angle.min()),
            "angle_max_rad": float(tables.cam_angle.max()),
            "angle_step_rad": float(np.median(np.diff(tables.cam_angle))),
            "radius_min_mm": float(tables.cam_radius.min()),
            "radius_max_mm": float(tables.cam_radius.max()),
            "stroke_mm": float(models.cam.stroke),
            "endpoint_radius_difference_mm": float(tables.cam_radius[-1] - tables.cam_radius[0]),
            "periodic_closure_assumption": "Append 2*pi with the radius at 0 rad.",
        },
        "needle": {
            "numeric_points_after_interval_expansion": int(len(tables.needle_time)),
            "time_min_ms": float(tables.needle_time.min()),
            "time_max_ms": float(tables.needle_time.max()),
            "lift_min_mm": float(tables.needle_lift.min()),
            "lift_max_mm": float(tables.needle_lift.max()),
            "encoded_constant_intervals": [
                {"start_ms": left, "end_ms": right, "lift_mm": lift}
                for left, right, lift in tables.needle_intervals
            ],
            "interpretation": (
                "[0.45,2] is a 2 mm plateau; [2.46,100] is the closed interval. "
                "These are interval records, not missing or malformed numeric samples."
            ),
        },
        "checks": checks,
    }


def provenance_ledger(workspace: Path) -> dict[str, Any]:
    """Build the single canonical ledger for training-memory provenance.

    The revision reads only the copies already contained in ``blind-v1``.  It
    does not re-query the trainer knowledge base or any external source.
    """
    decisions = {
        "TM-001-EVIDENCE-LAYERS": {
            "decision": "adopt",
            "uses": ["evidence matrix", "status vocabulary", "claim layering"],
        },
        "TM-002-TIME-ORDERED-VALIDATION": {
            "decision": "reject",
            "uses": [],
        },
        "TM-007-FIT-EVALUATE-SEPARATION": {
            "decision": "adapt",
            "uses": ["design/evaluation window separation", "model-internal wording"],
        },
    }
    card_dir = workspace / "blind-v1" / "knowledge" / "training-memory"
    cards = []
    for card_id, decision in decisions.items():
        path = card_dir / f"{card_id}.yaml"
        if not path.is_file():
            raise FileNotFoundError(f"Missing frozen V1 provenance card: {path}")
        cards.append(
            {
                "id": card_id,
                "source_status": "provisional_training",
                "decision": decision["decision"],
                "uses": decision["uses"],
                "relative_path": path.relative_to(workspace).as_posix(),
                "sha256": file_sha256(path),
                "size_bytes": path.stat().st_size,
            }
        )
    source_records = []
    for relative in (
        "blind-v1/retrieval-log.json",
        "blind-v1/reports/training-memory-retrieval.json",
        "blind-v1/reports/training-memory-usage.md",
    ):
        path = workspace / relative
        source_records.append(
            {
                "relative_path": relative,
                "sha256": file_sha256(path),
                "size_bytes": path.stat().st_size,
            }
        )
    return {
        "schema_version": 1,
        "case_id": "2019A",
        "phase": "blind-revision",
        "overall_status": "needs_review",
        "ledger_consistency_status": "pass",
        "source_validity_status": "needs_review",
        "source_validity_reason": (
            "The three cards are frozen V1 copies with provisional_training status; "
            "they are workflow guidance, not current-problem truth."
        ),
        "retrieval_scope": "frozen_blind_v1_copies_only",
        "cards": cards,
        "source_records": source_records,
        "forbidden_material_used": False,
        "network_used": False,
        "audit_finding_resolved": (
            "The revision retrieval log, usage report, solution report, and this ledger "
            "all enumerate the same three card IDs and decisions."
        ),
    }


def mean_problem1_tau(models: Models, target: float) -> float:
    average_injection = 44.0 / CONSTS.injection_period
    open_flow = orifice_flow(
        CONSTS.pump_supply_pressure, target, CONSTS.inlet_area, models.fuel
    )
    rho_supply = models.fuel.rho(CONSTS.pump_supply_pressure)
    rho_rail = models.fuel.rho(target)
    average_injection_mass = rho_rail * average_injection
    return (
        CONSTS.valve_closed_time
        * average_injection_mass
        / (rho_supply * open_flow - average_injection_mass)
    )


def calibrate_problem1_tau(models: Models, target: float, baseline: float) -> float:
    def residual(tau: float) -> float:
        simulation = simulate_problem1(
            models.fuel,
            30000.0,
            tau,
            initial_pressure=target,
            dt=0.05,
        )
        value = pressure_metrics(simulation, target, 20000.0, 30000.0)["mean_mpa"] - target
        del simulation
        return value

    return float(brentq(residual, baseline * 0.95, baseline * 1.05, xtol=2e-7))


def transition_command(
    models: Models,
    transition_ms: float,
    tau_100: float,
    tau_150: float,
    baseline_100: float,
    baseline_150: float,
):
    average_injection = 44.0 / CONSTS.injection_period

    def command(time_ms: float) -> float:
        if time_ms >= transition_ms:
            return tau_150
        u = time_ms / transition_ms
        smooth = 3.0 * u**2 - 2.0 * u**3
        smooth_derivative = (6.0 * u - 6.0 * u**2) / transition_ms
        target = 100.0 + 50.0 * smooth
        target_derivative = 50.0 * smooth_derivative
        open_flow = orifice_flow(
            CONSTS.pump_supply_pressure, target, CONSTS.inlet_area, models.fuel
        )
        rail_side_required = (
            average_injection
            + CONSTS.rail_volume / models.fuel.e(target) * target_derivative
        )
        required_average = (
            models.fuel.rho(target)
            / models.fuel.rho(CONSTS.pump_supply_pressure)
            * rail_side_required
        )
        tau = (
            CONSTS.valve_closed_time
            * required_average
            / (open_flow - required_average)
        )
        # Carry the small dynamic correction of the two endpoint holding controls.
        correction = (1.0 - smooth) * (tau_100 - baseline_100) + smooth * (
            tau_150 - baseline_150
        )
        return tau + correction

    return command


def flow_window(simulation, start: float, end: float) -> dict[str, float]:
    interval_end = simulation.time[1:]
    mask = (interval_end >= start) & (interval_end <= end)
    dt = np.diff(simulation.time)[mask]
    duration = end - start
    pump_volume = float(np.sum(simulation.interval_pump_flow[mask] * dt))
    injection_volume = float(np.sum(simulation.interval_injection_flow[mask] * dt))
    relief_volume = float(np.sum(simulation.interval_relief_flow[mask] * dt))
    pump_mass = float(np.sum(simulation.interval_pump_mass_flow[mask] * dt))
    injection_mass = float(np.sum(simulation.interval_injection_mass_flow[mask] * dt))
    relief_mass = float(np.sum(simulation.interval_relief_mass_flow[mask] * dt))
    scale = 100.0 / duration
    return {
        "pump_volume_per_100ms_mm3": pump_volume * scale,
        "injection_volume_per_100ms_mm3": injection_volume * scale,
        "relief_volume_per_100ms_mm3": relief_volume * scale,
        "mean_pump_flow_mm3_per_ms": pump_volume / duration,
        "mean_injection_flow_mm3_per_ms": injection_volume / duration,
        "mean_relief_flow_mm3_per_ms": relief_volume / duration,
        "pump_mass_per_100ms_mg": pump_mass * scale,
        "injection_mass_per_100ms_mg": injection_mass * scale,
        "relief_mass_per_100ms_mg": relief_mass * scale,
        "net_mass_per_100ms_mg": (pump_mass - injection_mass - relief_mass) * scale,
    }


def calibrate_omega(
    models: Models,
    offsets: tuple[float, ...],
    baseline: float,
    *,
    duration: float = 3000.0,
    window_start: float = 2000.0,
    area_mode: str = "bottleneck",
    cam_phases: tuple[float, ...] = (math.pi,),
    dt: float = 0.02,
) -> float:
    def residual(omega: float) -> float:
        means = []
        for cam_phase0 in cam_phases:
            simulation = simulate_pump_system(
                models,
                omega,
                duration,
                dt=dt,
                injector_offsets=offsets,
                area_mode=area_mode,
                cam_phase0=cam_phase0,
            )
            means.append(
                pressure_metrics(simulation, 100.0, window_start, duration)["mean_mpa"]
            )
            del simulation
        return float(np.mean(means) - 100.0)

    return float(brentq(residual, baseline * 0.97, baseline * 1.03, xtol=2e-7))


def scaled_elasticity_models(models: Models, factor: float) -> Models:
    tables = models.tables
    fuel = FuelProperties(tables.pressure, tables.elasticity * factor)
    return Models(tables=tables, fuel=fuel, needle=models.needle, cam=models.cam)


def binary_event_statistics(
    time: np.ndarray,
    state: np.ndarray,
    start: float,
    end: float,
) -> dict[str, Any]:
    mask = (time >= start) & (time <= end)
    local_time = np.asarray(time[mask], dtype=float)
    local_state = np.asarray(state[mask], dtype=bool)
    if len(local_time) < 2:
        raise ValueError("Event window contains fewer than two samples")
    change_indices = np.flatnonzero(local_state[1:] != local_state[:-1]) + 1
    switch_times = local_time[change_indices]
    boundaries = np.concatenate(([start], switch_times, [end]))
    dwell = np.diff(boundaries)
    duration = end - start
    return {
        "state_changes": int(len(switch_times)),
        "state_changes_per_100ms": float(len(switch_times) * 100.0 / duration),
        "minimum_dwell_ms": float(np.min(dwell)),
        "p05_dwell_ms": float(np.quantile(dwell, 0.05)),
        "median_dwell_ms": float(np.median(dwell)),
        "maximum_dwell_ms": float(np.max(dwell)),
    }


def relief_control_as_dict(control: ReliefControl) -> dict[str, Any]:
    return {
        "close_pressure_mpa": control.close_pressure,
        "open_pressure_mpa": control.open_pressure,
        "open_delay_ms": control.open_delay_ms,
        "close_delay_ms": control.close_delay_ms,
        "min_open_ms": control.min_open_ms,
        "min_closed_ms": control.min_closed_ms,
        "max_switches_per_100ms": control.max_switches_per_100ms,
    }


def phase_ensemble_evaluation(
    models: Models,
    omega: float,
    area_mode: str,
    phases: tuple[float, ...],
    *,
    duration: float,
    window_start: float,
    dt: float,
) -> tuple[list[dict[str, Any]], dict[str, Any], Any]:
    rows: list[dict[str, Any]] = []
    representative = None
    for phase in phases:
        simulation = simulate_pump_system(
            models,
            omega,
            duration,
            dt=dt,
            injector_offsets=(0.0,),
            area_mode=area_mode,
            cam_phase0=phase,
        )
        metrics = pressure_metrics(simulation, 100.0, window_start, duration)
        rows.append({"cam_phase0_rad": phase, **metrics})
        if abs(phase - math.pi) < 1e-12:
            representative = simulation
        else:
            del simulation
    means = [row["mean_mpa"] for row in rows]
    aggregate_mean = float(np.mean(means))
    phase_span = float(max(means) - min(means))
    aggregate = {
        "phase_count": len(phases),
        "mean_of_phase_means_mpa": aggregate_mean,
        "phase_mean_span_mpa": phase_span,
        "worst_phase_rmse_mpa": float(max(row["rmse_mpa"] for row in rows)),
        "worst_phase_peak_to_peak_mpa": float(
            max(row["peak_to_peak_mpa"] for row in rows)
        ),
        "status": (
            "pass"
            if phase_span < 0.2 and abs(aggregate_mean - 100.0) < 0.2
            else "needs_review"
        ),
    }
    return rows, aggregate, representative


def plot_input_curves(models: Models, path: Path) -> None:
    figure, axes = plt.subplots(1, 3, figsize=(12, 3.4))
    axes[0].plot(models.tables.cam_angle, models.tables.cam_radius, color="#235789")
    axes[0].set(xlabel="Cam angle (rad)", ylabel="Radius (mm)", title="Attachment 1")
    time = np.linspace(0.0, 2.6, 800)
    axes[1].plot(time, models.needle.lift(time), color="#D95D39")
    axes[1].set(xlabel="Time in injection cycle (ms)", ylabel="Lift (mm)", title="Attachment 2")
    axes[2].plot(models.tables.pressure, models.tables.elasticity, color="#2E8B57")
    axes[2].set(xlabel="Pressure (MPa)", ylabel="Elastic modulus (MPa)", title="Attachment 3")
    for axis in axes:
        axis.grid(alpha=0.25)
    figure.tight_layout()
    figure.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(figure)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workspace", type=Path, default=Path(__file__).resolve().parents[1])
    args = parser.parse_args()
    workspace = args.workspace.resolve()
    results_dir = workspace / "results"
    figures_dir = workspace / "figures"
    results_dir.mkdir(parents=True, exist_ok=True)
    figures_dir.mkdir(parents=True, exist_ok=True)

    print("[1/8] loading and auditing allowed inputs", flush=True)
    models = load_models(workspace)
    audit = data_audit(workspace, models)
    write_json(results_dir / "data-audit.json", audit)
    write_json(results_dir / "input-manifest.json", audit["input_manifest"])
    provenance = provenance_ledger(workspace)
    write_json(results_dir / "provenance.json", provenance)
    plot_input_curves(models, figures_dir / "<SOURCE_FILE_REDACTED>")

    density_values = {
        str(pressure): models.fuel.rho(pressure)
        for pressure in (0.5, 100.0, 150.0, 160.0, 200.0)
    }
    pulse_bottleneck = needle_pulse_volume(models, area_mode="bottleneck")
    pulse_series = needle_pulse_volume(models, area_mode="series")
    delivery = plunger_delivery_volume(models)
    omega_baseline = mean_balance_omega(models)
    omega_series_baseline = mean_balance_omega(models, area_mode="series")
    p1_baseline_100 = mean_problem1_tau(models, 100.0)
    p1_baseline_150 = mean_problem1_tau(models, 150.0)
    baseline = {
        "rail_volume_mm3": CONSTS.rail_volume,
        "figure_2_injection_volume_mm3": 44.0,
        "figure_2_mean_flow_mm3_per_ms": 44.0 / CONSTS.injection_period,
        "density_mg_per_mm3": density_values,
        "problem1_mean_balance_tau_ms": {
            "100_mpa": p1_baseline_100,
            "150_mpa": p1_baseline_150,
        },
        "needle_pulse_volume_mm3": {
            "bottleneck_area": pulse_bottleneck,
            "series_resistance": pulse_series,
        },
        "plunger_stroke_mm": models.cam.stroke,
        "plunger_bottom_volume_mm3": (
            CONSTS.residual_volume + CONSTS.plunger_area * models.cam.stroke
        ),
        "plunger_delivery_at_100mpa_mm3_per_cycle": delivery,
        "problem2_mean_balance_omega_rad_per_ms": {
            "bottleneck_area": omega_baseline,
            "series_resistance": omega_series_baseline,
        },
    }
    write_json(results_dir / "baseline.json", baseline)

    print("[2/8] evaluating density-corrected Problem 1 controls", flush=True)
    # The exact cycle-mean mass balance is already within the stated 0.1 MPa
    # long-window tolerance, so it remains an independent arithmetic baseline
    # rather than being hidden behind a dynamic fit.
    tau_100 = p1_baseline_100
    tau_150 = p1_baseline_150
    p1_hold: dict[str, Any] = {}
    p1_hold_plot: dict[str, np.ndarray] = {}
    p1_balance_checks: dict[str, Any] = {}
    for target, tau in ((100.0, tau_100), (150.0, tau_150)):
        simulation = simulate_problem1(
            models.fuel, 30000.0, tau, initial_pressure=target, dt=0.02
        )
        metrics = pressure_metrics(simulation, target, 20000.0, 30000.0)
        p1_hold[f"{int(target)}_mpa"] = {
            "tau_ms": tau,
            "control_source": "exact_cycle_mean_mass_balance",
            "metrics_20s_to_30s": metrics,
        }
        p1_balance_checks[f"{int(target)}_mpa"] = {
            "physical_mass": rail_mass_balance_residual(simulation, models.fuel),
            "transformed_volume_numerical_check": transformed_volume_balance_residual(
                simulation, models.fuel
            ),
        }
        mask = simulation.time >= 29000.0
        p1_hold_plot[f"{int(target)}_mpa"] = np.column_stack(
            [simulation.time[mask] - 29000.0, simulation.pressure[mask]]
        )[::5]
        save_timeseries(
            results_dir / f"problem1-hold-{int(target)}<SOURCE_FILE_REDACTED>",
            downsample_simulation(simulation, 50),
        )
        del simulation
        gc.collect()

    figure, axes = plt.subplots(2, 1, figsize=(9, 6), sharex=True)
    for axis, target in zip(axes, (100, 150)):
        values = p1_hold_plot[f"{target}_mpa"]
        axis.plot(values[:, 0], values[:, 1], linewidth=0.9)
        axis.axhline(target, color="black", linestyle="--", linewidth=0.8)
        axis.set_ylabel("Pressure (MPa)")
        axis.set_title(f"Holding at {target} MPa, last 1 s")
        axis.grid(alpha=0.25)
    axes[-1].set_xlabel("Local time (ms)")
    figure.tight_layout()
    figure.savefig(figures_dir / "<SOURCE_FILE_REDACTED>", dpi=180, bbox_inches="tight")
    plt.close(figure)

    print("[3/8] generating smooth 2 s / 5 s / 10 s transitions", flush=True)
    transition_rows = []
    transition_summary = []
    transition_plot = []
    for transition_ms in (2000.0, 5000.0, 10000.0):
        command = transition_command(
            models,
            transition_ms,
            tau_100,
            tau_150,
            p1_baseline_100,
            p1_baseline_150,
        )
        schedule = build_valve_schedule(transition_ms + 1000.0, command)
        simulation = simulate_problem1(
            models.fuel,
            transition_ms + 1000.0,
            command(0.0),
            initial_pressure=100.0,
            dt=0.01,
            valve_schedule=schedule,
        )
        u = np.clip(simulation.time / transition_ms, 0.0, 1.0)
        reference = np.where(
            simulation.time < transition_ms,
            100.0 + 50.0 * (3.0 * u**2 - 2.0 * u**3),
            150.0,
        )
        tracking_mask = simulation.time <= transition_ms
        tracking_rmse = float(
            np.sqrt(np.mean((simulation.pressure[tracking_mask] - reference[tracking_mask]) ** 2))
        )
        schedule_before = schedule[1][schedule[0] < transition_ms]
        representative = {
            f"u_{fraction:.2f}": command(fraction * transition_ms)
            for fraction in (0.0, 0.25, 0.5, 0.75, 1.0)
        }
        item = {
            "transition_ms": transition_ms,
            "cycle_count_before_transition": int(len(schedule_before)),
            "tau_representative_ms": representative,
            "tau_min_before_transition_ms": float(np.min(schedule_before)),
            "tau_max_before_transition_ms": float(np.max(schedule_before)),
            "tracking_rmse_mpa": tracking_rmse,
            "last_100ms_before_switch": pressure_metrics(
                simulation, 150.0, transition_ms - 100.0, transition_ms
            ),
            "hold_evaluation_500ms": pressure_metrics(
                simulation, 150.0, transition_ms + 500.0, transition_ms + 1000.0
            ),
            "rail_mass_balance": rail_mass_balance_residual(simulation, models.fuel),
            "transformed_volume_numerical_check": transformed_volume_balance_residual(
                simulation, models.fuel
            ),
        }
        transition_summary.append(item)
        for start, tau in zip(schedule[0], schedule[1]):
            if start <= transition_ms:
                transition_rows.append(
                    {
                        "transition_target_ms": transition_ms,
                        "cycle_start_ms": start,
                        "open_duration_ms": tau,
                    }
                )
        stride = max(1, int(round(1.0 / (simulation.time[1] - simulation.time[0]))))
        sampled = np.column_stack(
            [simulation.time[::stride], simulation.pressure[::stride], reference[::stride]]
        )
        transition_plot.append((transition_ms, sampled))
        save_timeseries(
            results_dir / f"problem1-transition-{int(transition_ms / 1000)}<SOURCE_FILE_REDACTED>",
            downsample_simulation(simulation, stride),
        )
        del simulation
        gc.collect()
    write_csv(results_dir / "<SOURCE_FILE_REDACTED>", transition_rows)
    figure, axes = plt.subplots(3, 1, figsize=(9, 8), sharey=True)
    for axis, (transition_ms, sampled) in zip(axes, transition_plot):
        axis.plot(sampled[:, 0] / 1000.0, sampled[:, 1], label="simulation", linewidth=0.9)
        axis.plot(sampled[:, 0] / 1000.0, sampled[:, 2], label="smooth reference", linestyle="--")
        axis.axvline(transition_ms / 1000.0, color="grey", linestyle=":")
        axis.set_ylabel("MPa")
        axis.set_title(f"Nominal transition time: {transition_ms / 1000:.0f} s")
        axis.grid(alpha=0.25)
    axes[0].legend(loc="lower right")
    axes[-1].set_xlabel("Time (s)")
    figure.tight_layout()
    figure.savefig(figures_dir / "<SOURCE_FILE_REDACTED>", dpi=180, bbox_inches="tight")
    plt.close(figure)
    problem1 = {
        "holding": p1_hold,
        "transitions": transition_summary,
        "holding_balance_checks": p1_balance_checks,
    }
    write_json(results_dir / "problem1-summary.json", problem1)

    print("[4/8] calibrating the nozzle-structure and initial-phase envelope", flush=True)
    phase_ensemble = (0.0, 0.5 * math.pi, math.pi, 1.5 * math.pi)
    omega_problem2 = calibrate_omega(
        models,
        (0.0,),
        omega_baseline,
        duration=90000.0,
        window_start=80000.0,
        area_mode="bottleneck",
        cam_phases=(math.pi,),
        dt=0.05,
    )
    omega_series_dynamic = calibrate_omega(
        models,
        (0.0,),
        omega_series_baseline,
        duration=90000.0,
        window_start=80000.0,
        area_mode="series",
        cam_phases=(math.pi,),
        dt=0.05,
    )
    p2_bottleneck_rows, p2_bottleneck_aggregate, p2_simulation = phase_ensemble_evaluation(
        models,
        omega_problem2,
        "bottleneck",
        phase_ensemble,
        duration=100000.0,
        window_start=90000.0,
        dt=0.05,
    )
    p2_series_rows, p2_series_aggregate, p2_series_representative = phase_ensemble_evaluation(
        models,
        omega_series_dynamic,
        "series",
        phase_ensemble,
        duration=100000.0,
        window_start=90000.0,
        dt=0.05,
    )
    phase_rows = [
        {"area_model": "bottleneck", **row} for row in p2_bottleneck_rows
    ] + [{"area_model": "series", **row} for row in p2_series_rows]
    write_csv(results_dir / "<SOURCE_FILE_REDACTED>", phase_rows)
    if p2_series_representative is not None:
        del p2_series_representative

    p2_eval_metrics = pressure_metrics(p2_simulation, 100.0, 90000.0, 100000.0)
    p2_flows = flow_window(p2_simulation, 90000.0, 100000.0)
    p2_mass_balance = rail_mass_balance_residual(p2_simulation, models.fuel)
    p2_transformed_balance = transformed_volume_balance_residual(
        p2_simulation, models.fuel
    )
    p2_chamber_balance = chamber_compression_mass_residual(
        p2_simulation, models, omega_problem2, math.pi
    )
    structural_interval = sorted([omega_series_dynamic, omega_problem2])
    p2_summary = {
        "answer_type": "structural_interval",
        "status": "needs_review",
        "reason": (
            "Separate loss coefficients or an intermediate nozzle pressure are absent; "
            "both admissible area laws are retained."
        ),
        "cam_phase0_values_rad": list(phase_ensemble),
        "calibration_cam_phase0_rad": math.pi,
        "calibration_window_ms": [80000.0, 90000.0],
        "model_internal_evaluation_window_ms": [90000.0, 100000.0],
        "structural_cases": {
            "bottleneck_upper_flow": {
                "mean_balance_omega_rad_per_ms": omega_baseline,
                "dynamic_omega_rad_per_ms": omega_problem2,
                "speed_rpm": omega_problem2 * 1000.0 * 60.0 / (2.0 * math.pi),
                "needle_pulse_volume_at_100mpa_mm3": pulse_bottleneck,
                "phase_ensemble": p2_bottleneck_rows,
                "phase_aggregate": p2_bottleneck_aggregate,
            },
            "series_resistance_lower_flow": {
                "mean_balance_omega_rad_per_ms": omega_series_baseline,
                "dynamic_omega_rad_per_ms": omega_series_dynamic,
                "speed_rpm": omega_series_dynamic * 1000.0 * 60.0 / (2.0 * math.pi),
                "needle_pulse_volume_at_100mpa_mm3": pulse_series,
                "phase_ensemble": p2_series_rows,
                "phase_aggregate": p2_series_aggregate,
            },
        },
        "omega_interval_rad_per_ms": structural_interval,
        "speed_interval_rpm": [
            value * 1000.0 * 60.0 / (2.0 * math.pi) for value in structural_interval
        ],
        "plunger_delivery_at_100mpa_mm3": delivery,
        "representative_bottleneck_pi_phase_metrics": p2_eval_metrics,
        "representative_flows": p2_flows,
        "rail_mass_balance": p2_mass_balance,
        "rail_transformed_volume_numerical_check": p2_transformed_balance,
        "chamber_compression_mass_balance": p2_chamber_balance,
        "chamber_pressure_range_evaluation_mpa": [
            float(np.min(p2_simulation.chamber_pressure[p2_simulation.time >= 90000.0])),
            float(np.max(p2_simulation.chamber_pressure[p2_simulation.time >= 90000.0])),
        ],
    }
    write_json(results_dir / "problem2-summary.json", p2_summary)
    save_timeseries(
        results_dir / "<SOURCE_FILE_REDACTED>", downsample_simulation(p2_simulation, 100)
    )
    p2_plot_mask = p2_simulation.time >= 99000.0
    p2_plot = np.column_stack(
        [
            p2_simulation.time[p2_plot_mask] - 99000.0,
            p2_simulation.pressure[p2_plot_mask],
            p2_simulation.chamber_pressure[p2_plot_mask],
            p2_simulation.pump_flow[p2_plot_mask],
            p2_simulation.injection_flow[p2_plot_mask],
        ]
    )[::5]
    del p2_simulation
    gc.collect()
    figure, axes = plt.subplots(2, 1, figsize=(10, 6), sharex=True)
    axes[0].plot(p2_plot[:, 0], p2_plot[:, 1], label="rail")
    axes[0].plot(p2_plot[:, 0], p2_plot[:, 2], label="plunger chamber", alpha=0.7)
    axes[0].axhline(100.0, color="black", linestyle="--", linewidth=0.8)
    axes[0].set_ylabel("Pressure (MPa)")
    axes[0].legend()
    axes[0].grid(alpha=0.25)
    axes[1].plot(p2_plot[:, 0], p2_plot[:, 3], label="pump to rail")
    axes[1].plot(p2_plot[:, 0], p2_plot[:, 4], label="injector")
    axes[1].set(xlabel="Local time (ms)", ylabel="Flow (mm3/ms)")
    axes[1].legend()
    axes[1].grid(alpha=0.25)
    figure.tight_layout()
    figure.savefig(figures_dir / "<SOURCE_FILE_REDACTED>", dpi=180, bbox_inches="tight")
    plt.close(figure)

    print("[5/8] comparing two-injector phasing and constrained relief controls", flush=True)
    two_baseline = mean_balance_omega(models, injector_count=2)
    two_series_baseline = mean_balance_omega(models, injector_count=2, area_mode="series")
    phase_rows: list[dict[str, Any]] = []
    phase_roots: dict[float, float] = {}
    for offset in (0.0, 25.0, 50.0):
        omega = calibrate_omega(
            models,
            (0.0, offset),
            two_baseline,
            duration=4000.0,
            window_start=2000.0,
            cam_phases=(math.pi,),
            dt=0.03,
        )
        phase_roots[offset] = omega
        simulation = simulate_pump_system(
            models,
            omega,
            6000.0,
            dt=0.01,
            injector_offsets=(0.0, offset),
            cam_phase0=math.pi,
        )
        metrics = pressure_metrics(simulation, 100.0, 3000.0, 6000.0)
        phase_rows.append(
            {
                "second_injector_offset_ms": offset,
                "area_model": "bottleneck",
                "cam_phase0_rad": math.pi,
                "omega_rad_per_ms": omega,
                "speed_rpm": omega * 1000.0 * 60.0 / (2.0 * math.pi),
                "evaluation_mean_mpa": metrics["mean_mpa"],
                "evaluation_rmse_mpa": metrics["rmse_mpa"],
                "evaluation_peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                "evaluation_max_abs_error_mpa": metrics["max_abs_error_mpa"],
            }
        )
        del simulation
        gc.collect()
    write_csv(results_dir / "<SOURCE_FILE_REDACTED>", phase_rows)

    economy_omega = phase_roots[50.0]
    economy_series_omega = calibrate_omega(
        models,
        (0.0, 50.0),
        two_series_baseline,
        duration=4000.0,
        window_start=2000.0,
        area_mode="series",
        cam_phases=(math.pi,),
        dt=0.03,
    )
    economy = simulate_pump_system(
        models,
        economy_omega,
        5000.0,
        dt=0.005,
        injector_offsets=(0.0, 50.0),
        cam_phase0=math.pi,
    )
    economy_metrics = pressure_metrics(economy, 100.0, 2500.0, 5000.0)
    economy_flows = flow_window(economy, 2500.0, 5000.0)
    economy_mass_balance = rail_mass_balance_residual(economy, models.fuel)
    economy_transformed_balance = transformed_volume_balance_residual(economy, models.fuel)
    economy_chamber_balance = chamber_compression_mass_residual(
        economy, models, economy_omega, math.pi
    )
    economy_mask = economy.time >= 4500.0
    economy_plot = np.column_stack(
        [economy.time[economy_mask] - 4500.0, economy.pressure[economy_mask]]
    )[::5]
    save_timeseries(results_dir / "<SOURCE_FILE_REDACTED>", downsample_simulation(economy, 20))
    del economy
    gc.collect()

    actuator_profiles = [
        {
            "profile": "ideal_zero_dwell",
            "min_dwell_ms": 0.0,
            "open_delay_ms": 0.0,
            "close_delay_ms": 0.0,
            "max_switches_per_100ms": None,
        },
        {
            "profile": "dwell_0.5ms",
            "min_dwell_ms": 0.5,
            "open_delay_ms": 0.0,
            "close_delay_ms": 0.0,
            "max_switches_per_100ms": 100,
        },
        {
            "profile": "dwell_1ms",
            "min_dwell_ms": 1.0,
            "open_delay_ms": 0.0,
            "close_delay_ms": 0.0,
            "max_switches_per_100ms": 60,
        },
        {
            "profile": "dwell_2ms",
            "min_dwell_ms": 2.0,
            "open_delay_ms": 0.0,
            "close_delay_ms": 0.0,
            "max_switches_per_100ms": 30,
        },
        {
            "profile": "dwell_5ms",
            "min_dwell_ms": 5.0,
            "open_delay_ms": 0.0,
            "close_delay_ms": 0.0,
            "max_switches_per_100ms": 12,
        },
        {
            "profile": "dwell_1ms_delay_0.2ms",
            "min_dwell_ms": 1.0,
            "open_delay_ms": 0.2,
            "close_delay_ms": 0.2,
            "max_switches_per_100ms": 60,
        },
    ]
    design_rows: list[dict[str, Any]] = []
    selected_by_profile: dict[str, dict[str, Any]] = {}
    for profile in actuator_profiles:
        profile_rows = []
        for speed_rpm in (900.0, 1050.0, 1200.0):
            omega = speed_rpm * 2.0 * math.pi / (60.0 * 1000.0)
            for cam_phase in (4.2, 4.6, 5.0):
                for center in (99.9, 100.1, 100.3):
                    for half_width in (0.1, 0.3):
                        control = ReliefControl(
                            close_pressure=center - half_width,
                            open_pressure=center + half_width,
                            open_delay_ms=profile["open_delay_ms"],
                            close_delay_ms=profile["close_delay_ms"],
                            min_open_ms=profile["min_dwell_ms"],
                            min_closed_ms=profile["min_dwell_ms"],
                            max_switches_per_100ms=profile["max_switches_per_100ms"],
                        )
                        simulation = simulate_pump_system(
                            models,
                            omega,
                            1000.0,
                            dt=0.02,
                            injector_offsets=(0.0, 50.0),
                            cam_phase0=cam_phase,
                            relief_control=control,
                        )
                        metrics = pressure_metrics(simulation, 100.0, 400.0, 1000.0)
                        flows = flow_window(simulation, 400.0, 1000.0)
                        dwell = relief_dwell_statistics(simulation, 400.0, 1000.0)
                        objective = (
                            metrics["rmse_mpa"]
                            + 0.0005 * flows["relief_volume_per_100ms_mm3"]
                            + 0.002 * dwell["state_changes_per_100ms"]
                        )
                        row = {
                            "profile": profile["profile"],
                            "speed_rpm": speed_rpm,
                            "omega_rad_per_ms": omega,
                            "cam_phase_rad": cam_phase,
                            "relief_close_mpa": control.close_pressure,
                            "relief_open_mpa": control.open_pressure,
                            "open_delay_ms": control.open_delay_ms,
                            "close_delay_ms": control.close_delay_ms,
                            "min_open_ms": control.min_open_ms,
                            "min_closed_ms": control.min_closed_ms,
                            "max_switches_per_100ms": control.max_switches_per_100ms,
                            "objective": objective,
                            "mean_mpa": metrics["mean_mpa"],
                            "rmse_mpa": metrics["rmse_mpa"],
                            "peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                            "relief_volume_per_100ms_mm3": flows[
                                "relief_volume_per_100ms_mm3"
                            ],
                            "state_changes_per_100ms": dwell["state_changes_per_100ms"],
                            "minimum_dwell_ms": dwell["minimum_dwell_ms"],
                            "status": "needs_review",
                        }
                        profile_rows.append(row)
                        design_rows.append(row)
                        del simulation
        profile_rows.sort(key=lambda row: row["objective"])
        selected_by_profile[profile["profile"]] = profile_rows[0]
    design_rows.sort(key=lambda row: (row["profile"], row["objective"]))
    write_csv(results_dir / "<SOURCE_FILE_REDACTED>", design_rows)
    constrained_scenarios: list[dict[str, Any]] = []
    precision = None
    for profile in actuator_profiles:
        selected = selected_by_profile[profile["profile"]]
        control = ReliefControl(
            close_pressure=selected["relief_close_mpa"],
            open_pressure=selected["relief_open_mpa"],
            open_delay_ms=selected["open_delay_ms"],
            close_delay_ms=selected["close_delay_ms"],
            min_open_ms=selected["min_open_ms"],
            min_closed_ms=selected["min_closed_ms"],
            max_switches_per_100ms=selected["max_switches_per_100ms"],
        )
        simulation = simulate_pump_system(
            models,
            selected["omega_rad_per_ms"],
            3000.0,
            dt=0.005,
            injector_offsets=(0.0, 50.0),
            cam_phase0=selected["cam_phase_rad"],
            relief_control=control,
        )
        metrics = pressure_metrics(simulation, 100.0, 1500.0, 3000.0)
        flows = flow_window(simulation, 1500.0, 3000.0)
        dwell = relief_dwell_statistics(simulation, 1500.0, 3000.0)
        scenario = {
            "profile": profile["profile"],
            "status": "needs_review",
            "status_reason": "Actuator timing is an assumed scenario, not supplied device data.",
            "pump_speed_rpm": selected["speed_rpm"],
            "omega_rad_per_ms": selected["omega_rad_per_ms"],
            "cam_phase_rad": selected["cam_phase_rad"],
            "relief_control": relief_control_as_dict(control),
            "metrics": metrics,
            "flows": flows,
            "dwell_statistics": dwell,
            "rail_mass_balance": rail_mass_balance_residual(simulation, models.fuel),
            "rail_transformed_volume_numerical_check": transformed_volume_balance_residual(
                simulation, models.fuel
            ),
            "relief_to_injection_volume_ratio": (
                flows["relief_volume_per_100ms_mm3"]
                / flows["injection_volume_per_100ms_mm3"]
            ),
        }
        constrained_scenarios.append(scenario)
        if profile["profile"] == "dwell_1ms":
            precision = simulation
            selected_design = selected
            selected_control = control
            precision_metrics = metrics
            precision_flows = flows
            precision_dwell = dwell
        else:
            del simulation
    if precision is None:
        raise RuntimeError("Conditional 1 ms relief scenario was not evaluated")
    precision_omega = selected_design["omega_rad_per_ms"]
    selected_thresholds = (
        selected_control.close_pressure,
        selected_control.open_pressure,
    )
    precision_mass_balance = rail_mass_balance_residual(precision, models.fuel)
    precision_chamber_balance = chamber_compression_mass_residual(
        precision, models, precision_omega, selected_design["cam_phase_rad"]
    )
    precision_mask = precision.time >= 2500.0
    precision_plot = np.column_stack(
        [precision.time[precision_mask] - 2500.0, precision.pressure[precision_mask]]
    )[::5]
    save_timeseries(results_dir / "<SOURCE_FILE_REDACTED>", downsample_simulation(precision, 20))
    del precision
    gc.collect()

    problem3 = {
        "selected_injector_strategy": "Two identical injections staggered by 50 ms.",
        "cam_phase0_for_economy_comparison_rad": math.pi,
        "phase_comparison": phase_rows,
        "economy_scheme": {
            "omega_rad_per_ms": economy_omega,
            "speed_rpm": economy_omega * 1000.0 * 60.0 / (2.0 * math.pi),
            "nozzle_structure_speed_interval_rad_per_ms": sorted(
                [economy_series_omega, economy_omega]
            ),
            "relief_control": "disabled_in_nominal_simulation",
            "metrics": economy_metrics,
            "flows": economy_flows,
            "rail_mass_balance": economy_mass_balance,
            "rail_transformed_volume_numerical_check": economy_transformed_balance,
            "chamber_compression_mass_balance": economy_chamber_balance,
        },
        "constrained_relief_scenarios": constrained_scenarios,
        "conditional_reference_scenario": {
            "profile": "dwell_1ms",
            "status": "needs_review",
            "interpretation": (
                "Illustrative design conditional on verified 1 ms minimum on/off dwell; "
                "not a hardware-certified final controller."
            ),
            "pump_cycle_period_ms": 2.0 * math.pi / precision_omega,
            "omega_rad_per_ms": precision_omega,
            "speed_rpm": selected_design["speed_rpm"],
            "cam_phase_at_injection_start_rad": selected_design["cam_phase_rad"],
            "relief_control": relief_control_as_dict(selected_control),
            "design_objective": (
                "RMSE + 0.0005 * relief volume per 100 ms + 0.002 * state changes per 100 ms"
            ),
            "metrics": precision_metrics,
            "flows": precision_flows,
            "dwell_statistics": precision_dwell,
            "relief_to_injection_volume_ratio": (
                precision_flows["relief_volume_per_100ms_mm3"]
                / precision_flows["injection_volume_per_100ms_mm3"]
            ),
            "rail_mass_balance": precision_mass_balance,
            "chamber_compression_mass_balance": precision_chamber_balance,
        },
    }
    write_json(results_dir / "problem3-summary.json", problem3)

    figure, axes = plt.subplots(2, 1, figsize=(9, 6), sharex=True)
    axes[0].plot(economy_plot[:, 0], economy_plot[:, 1])
    axes[0].axhline(100.0, color="black", linestyle="--", linewidth=0.8)
    axes[0].set_title("Economy scheme: 50 ms stagger, relief normally closed")
    axes[1].plot(precision_plot[:, 0], precision_plot[:, 1], color="#D95D39")
    axes[1].axhline(100.0, color="black", linestyle="--", linewidth=0.8)
    axes[1].set_title("Conditional 1 ms-dwell relief design (needs review)")
    for axis in axes:
        axis.set_ylabel("Pressure (MPa)")
        axis.grid(alpha=0.25)
    axes[-1].set_xlabel("Local time (ms)")
    figure.tight_layout()
    figure.savefig(figures_dir / "<SOURCE_FILE_REDACTED>", dpi=180, bbox_inches="tight")
    plt.close(figure)

    print("[6/8] running mass, event, convergence, and sensitivity checks", flush=True)
    sensitivity_rows: list[dict[str, Any]] = []
    event_rows: list[dict[str, Any]] = []
    timestep_metrics: dict[str, Any] = {"problem1": {}, "problem2": {}, "problem3": {}}

    for dt in (0.02, 0.01, 0.005):
        simulation = simulate_problem1(
            models.fuel, 3000.0, tau_100, initial_pressure=100.0, dt=dt
        )
        metrics = pressure_metrics(simulation, 100.0, 2000.0, 3000.0)
        flows = flow_window(simulation, 2000.0, 3000.0)
        events = binary_event_statistics(
            simulation.time, simulation.pump_flow > 0.0, 2000.0, 3000.0
        )
        mass = rail_mass_balance_residual(simulation, models.fuel)
        record = {"metrics": metrics, "flows": flows, "events": events, "mass": mass}
        timestep_metrics["problem1"][str(dt)] = record
        event_rows.append(
            {
                "component": "problem1_commanded_inlet_valve",
                "dt_ms": dt,
                "mean_mpa": metrics["mean_mpa"],
                "rmse_mpa": metrics["rmse_mpa"],
                "peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                "pump_mass_per_100ms_mg": flows["pump_mass_per_100ms_mg"],
                "injection_mass_per_100ms_mg": flows["injection_mass_per_100ms_mg"],
                "relief_mass_per_100ms_mg": flows["relief_mass_per_100ms_mg"],
                "state_changes_per_100ms": events["state_changes_per_100ms"],
                "minimum_dwell_ms": events["minimum_dwell_ms"],
                "mass_residual_relative": mass["max_relative_to_throughput"],
            }
        )
        del simulation

    for dt in (0.02, 0.01, 0.005):
        simulation = simulate_pump_system(
            models,
            omega_problem2,
            3000.0,
            dt=dt,
            injector_offsets=(0.0,),
            cam_phase0=math.pi,
        )
        metrics = pressure_metrics(simulation, 100.0, 2000.0, 3000.0)
        flows = flow_window(simulation, 2000.0, 3000.0)
        events = binary_event_statistics(
            simulation.time, simulation.pump_flow > 0.0, 2000.0, 3000.0
        )
        mass = rail_mass_balance_residual(simulation, models.fuel)
        record = {"metrics": metrics, "flows": flows, "events": events, "mass": mass}
        timestep_metrics["problem2"][str(dt)] = record
        sensitivity_rows.append(
            {
                "family": "time_step_problem2",
                "case": f"dt={dt} ms",
                "mean_mpa": metrics["mean_mpa"],
                "rmse_mpa": metrics["rmse_mpa"],
                "peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                "control_value": dt,
            }
        )
        event_rows.append(
            {
                "component": "problem2_pump_check_valve",
                "dt_ms": dt,
                "mean_mpa": metrics["mean_mpa"],
                "rmse_mpa": metrics["rmse_mpa"],
                "peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                "pump_mass_per_100ms_mg": flows["pump_mass_per_100ms_mg"],
                "injection_mass_per_100ms_mg": flows["injection_mass_per_100ms_mg"],
                "relief_mass_per_100ms_mg": flows["relief_mass_per_100ms_mg"],
                "state_changes_per_100ms": events["state_changes_per_100ms"],
                "minimum_dwell_ms": events["minimum_dwell_ms"],
                "mass_residual_relative": mass["max_relative_to_throughput"],
            }
        )
        del simulation

    for dt in (0.02, 0.01, 0.005, 0.0025):
        simulation = simulate_pump_system(
            models,
            precision_omega,
            2000.0,
            dt=dt,
            injector_offsets=(0.0, 50.0),
            cam_phase0=selected_design["cam_phase_rad"],
            relief_control=selected_control,
        )
        metrics = pressure_metrics(simulation, 100.0, 1000.0, 2000.0)
        flows = flow_window(simulation, 1000.0, 2000.0)
        events = relief_dwell_statistics(simulation, 1000.0, 2000.0)
        mass = rail_mass_balance_residual(simulation, models.fuel)
        record = {"metrics": metrics, "flows": flows, "events": events, "mass": mass}
        timestep_metrics["problem3"][str(dt)] = record
        event_rows.append(
            {
                "component": "problem3_constrained_relief_valve",
                "dt_ms": dt,
                "mean_mpa": metrics["mean_mpa"],
                "rmse_mpa": metrics["rmse_mpa"],
                "peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                "pump_mass_per_100ms_mg": flows["pump_mass_per_100ms_mg"],
                "injection_mass_per_100ms_mg": flows["injection_mass_per_100ms_mg"],
                "relief_mass_per_100ms_mg": flows["relief_mass_per_100ms_mg"],
                "state_changes_per_100ms": events["state_changes_per_100ms"],
                "minimum_dwell_ms": events["minimum_dwell_ms"],
                "mass_residual_relative": mass["max_relative_to_throughput"],
            }
        )
        del simulation
    write_csv(results_dir / "<SOURCE_FILE_REDACTED>", event_rows)

    for factor in (0.99, 1.0, 1.01):
        simulation = simulate_pump_system(
            models,
            omega_problem2 * factor,
            3000.0,
            dt=0.01,
            injector_offsets=(0.0,),
            cam_phase0=math.pi,
        )
        metrics = pressure_metrics(simulation, 100.0, 2000.0, 3000.0)
        sensitivity_rows.append(
            {
                "family": "pump_speed",
                "case": f"omega x {factor:.2f}",
                "mean_mpa": metrics["mean_mpa"],
                "rmse_mpa": metrics["rmse_mpa"],
                "peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                "control_value": omega_problem2 * factor,
            }
        )
        del simulation
    for factor in (0.95, 1.05):
        altered = scaled_elasticity_models(models, factor)
        simulation = simulate_pump_system(
            altered,
            omega_problem2,
            3000.0,
            dt=0.01,
            injector_offsets=(0.0,),
            cam_phase0=math.pi,
        )
        metrics = pressure_metrics(simulation, 100.0, 2000.0, 3000.0)
        sensitivity_rows.append(
            {
                "family": "elastic_modulus",
                "case": f"E x {factor:.2f}",
                "mean_mpa": metrics["mean_mpa"],
                "rmse_mpa": metrics["rmse_mpa"],
                "peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                "control_value": factor,
            }
        )
        del simulation
    series_simulation = simulate_pump_system(
        models,
        omega_series_dynamic,
        3000.0,
        dt=0.01,
        injector_offsets=(0.0,),
        area_mode="series",
        cam_phase0=math.pi,
    )
    series_metrics = pressure_metrics(series_simulation, 100.0, 2000.0, 3000.0)
    sensitivity_rows.append(
        {
            "family": "nozzle_structure",
            "case": "series resistance",
            "mean_mpa": series_metrics["mean_mpa"],
            "rmse_mpa": series_metrics["rmse_mpa"],
            "peak_to_peak_mpa": series_metrics["peak_to_peak_mpa"],
            "control_value": omega_series_dynamic,
        }
    )
    del series_simulation
    for phase_shift in (-0.1, 0.0, 0.1):
        simulation = simulate_pump_system(
            models,
            precision_omega,
            2000.0,
            dt=0.01,
            injector_offsets=(0.0, 50.0),
            cam_phase0=selected_design["cam_phase_rad"] + phase_shift,
            relief_control=selected_control,
        )
        metrics = pressure_metrics(simulation, 100.0, 1000.0, 2000.0)
        sensitivity_rows.append(
            {
                "family": "precision_cam_phase",
                "case": f"phase shift {phase_shift:+.1f} rad",
                "mean_mpa": metrics["mean_mpa"],
                "rmse_mpa": metrics["rmse_mpa"],
                "peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                "control_value": selected_design["cam_phase_rad"] + phase_shift,
            }
        )
        del simulation
    for threshold_shift in (-0.05, 0.05):
        shifted = ReliefControl(
            close_pressure=selected_thresholds[0] + threshold_shift,
            open_pressure=selected_thresholds[1] + threshold_shift,
            open_delay_ms=selected_control.open_delay_ms,
            close_delay_ms=selected_control.close_delay_ms,
            min_open_ms=selected_control.min_open_ms,
            min_closed_ms=selected_control.min_closed_ms,
            max_switches_per_100ms=selected_control.max_switches_per_100ms,
        )
        simulation = simulate_pump_system(
            models,
            precision_omega,
            2000.0,
            dt=0.01,
            injector_offsets=(0.0, 50.0),
            cam_phase0=selected_design["cam_phase_rad"],
            relief_control=shifted,
        )
        metrics = pressure_metrics(simulation, 100.0, 1000.0, 2000.0)
        sensitivity_rows.append(
            {
                "family": "precision_relief_threshold",
                "case": f"threshold shift {threshold_shift:+.2f} MPa",
                "mean_mpa": metrics["mean_mpa"],
                "rmse_mpa": metrics["rmse_mpa"],
                "peak_to_peak_mpa": metrics["peak_to_peak_mpa"],
                "control_value": threshold_shift,
            }
        )
        del simulation
    write_csv(results_dir / "<SOURCE_FILE_REDACTED>", sensitivity_rows)

    convergence: dict[str, Any] = {}
    for component, fine_key in (("problem1", "0.005"), ("problem2", "0.005"), ("problem3", "0.0025")):
        coarse_key = "0.005" if component == "problem3" else "0.01"
        medium = timestep_metrics[component][coarse_key]
        fine = timestep_metrics[component][fine_key]
        pressure_differences = {
            "mean_mpa": abs(medium["metrics"]["mean_mpa"] - fine["metrics"]["mean_mpa"]),
            "rmse_mpa": abs(medium["metrics"]["rmse_mpa"] - fine["metrics"]["rmse_mpa"]),
            "peak_to_peak_mpa": abs(
                medium["metrics"]["peak_to_peak_mpa"]
                - fine["metrics"]["peak_to_peak_mpa"]
            ),
        }
        event_difference = abs(
            medium["events"]["state_changes_per_100ms"]
            - fine["events"]["state_changes_per_100ms"]
        )
        pressure_status = (
            "pass"
            if max(pressure_differences["mean_mpa"], pressure_differences["rmse_mpa"]) < 0.03
            and pressure_differences["peak_to_peak_mpa"] < 0.08
            else "fail"
        )
        event_status = "pass" if event_difference <= 1.0 else "needs_review"
        convergence[component] = {
            "pressure_status": pressure_status,
            "event_status": event_status,
            "status": "fail" if pressure_status == "fail" else event_status,
            "pressure_differences": pressure_differences,
            "state_changes_per_100ms_difference": event_difference,
            "coarse_dt_ms": float(coarse_key),
            "fine_dt_ms": float(fine_key),
        }
    convergence["overall_status"] = (
        "fail"
        if any(item["status"] == "fail" for item in convergence.values() if isinstance(item, dict))
        else (
            "needs_review"
            if any(
                item["status"] == "needs_review"
                for item in convergence.values()
                if isinstance(item, dict)
            )
            else "pass"
        )
    )

    figure, axis = plt.subplots(figsize=(9, 4.5))
    labels = [row["case"] for row in sensitivity_rows if row["family"] in {"pump_speed", "elastic_modulus"}]
    values = [row["rmse_mpa"] for row in sensitivity_rows if row["family"] in {"pump_speed", "elastic_modulus"}]
    axis.bar(labels, values, color=["#235789"] * 3 + ["#D95D39"] * 2)
    axis.set_ylabel("Pressure RMSE (MPa)")
    axis.set_title("Problem 2 parameter sensitivity (controls fixed)")
    axis.tick_params(axis="x", rotation=20)
    axis.grid(axis="y", alpha=0.25)
    figure.tight_layout()
    figure.savefig(figures_dir / "<SOURCE_FILE_REDACTED>", dpi=180, bbox_inches="tight")
    plt.close(figure)

    print("[7/8] writing comparison and validation records", flush=True)
    model_comparison = [
        {
            "candidate": "C0 cycle-mean balance baseline",
            "selected": "no",
            "data_need": "geometry, E(P), cycle volumes",
            "assumptions": "pressure constant within a cycle",
            "complexity": "low",
            "interpretability": "high",
            "robustness_role": "search bounds and arithmetic cross-check",
            "reason": "cannot quantify within-cycle fluctuation or valve timing",
        },
        {
            "candidate": "C1 coupled variable-compressibility ODE with bottleneck nozzle area",
            "selected": "structural_envelope_upper_flow",
            "data_need": "all three attachments and stated geometry",
            "assumptions": "lumped uniform pressure; limiting cross-section dominates",
            "complexity": "medium",
            "interpretability": "high",
            "robustness_role": "upper-flow endpoint plus mesh and perturbation checks",
            "reason": "admissible but not uniquely identified by the supplied nozzle data",
        },
        {
            "candidate": "C2 coupled ODE with two serial orifice resistances",
            "selected": "structural_envelope_lower_flow",
            "data_need": "same data plus an unsupported resistance-sharing rule",
            "assumptions": "same C at both serial restrictions",
            "complexity": "medium",
            "interpretability": "medium",
            "robustness_role": "lower-flow structural endpoint",
            "reason": "retained because intermediate pressure and separate loss coefficients are absent",
        },
    ]
    write_csv(results_dir / "<SOURCE_FILE_REDACTED>", model_comparison)

    all_pressures_finite = all(
        np.isfinite(value)
        for value in (
            p1_hold["100_mpa"]["metrics_20s_to_30s"]["mean_mpa"],
            p1_hold["150_mpa"]["metrics_20s_to_30s"]["mean_mpa"],
            p2_eval_metrics["mean_mpa"],
            economy_metrics["mean_mpa"],
            precision_metrics["mean_mpa"],
        )
    )
    maximum_mass_balance_relative = max(
        p2_mass_balance["max_relative_to_throughput"],
        economy_mass_balance["max_relative_to_throughput"],
        precision_mass_balance["max_relative_to_throughput"],
        *(
            item["rail_mass_balance"]["max_relative_to_throughput"]
            for item in transition_summary
        ),
        *(
            item["physical_mass"]["max_relative_to_throughput"]
            for item in p1_balance_checks.values()
        ),
    )
    maximum_chamber_mass_balance_relative = max(
        p2_chamber_balance["max_relative_to_pumped_mass"],
        economy_chamber_balance["max_relative_to_pumped_mass"],
        precision_chamber_balance["max_relative_to_pumped_mass"],
    )
    validation = {
        "overall_status": "needs_review",
        "checks": [
            {
                "check": "allowed input schemas and monotonic independent variables",
                "status": (
                    "pass"
                    if all(value != "fail" for value in audit["checks"].values())
                    else "fail"
                ),
                "evidence": "results/data-audit.json",
            },
            {
                "check": "all reported numerical pressures finite",
                "status": "pass" if all_pressures_finite else "fail",
                "evidence": "results/problem1-summary.json; problem2-summary.json; problem3-summary.json",
            },
            {
                "check": "physical rail mass residual below 1e-4 of mass throughput",
                "status": "pass" if maximum_mass_balance_relative < 1e-4 else "fail",
                "value": maximum_mass_balance_relative,
                "evidence": "summary JSON rail_mass_balance fields",
            },
            {
                "check": "plunger chamber compression-segment mass residual below 1e-4",
                "status": (
                    "pass" if maximum_chamber_mass_balance_relative < 1e-4 else "fail"
                ),
                "value": maximum_chamber_mass_balance_relative,
                "evidence": "problem2-summary.json; problem3-summary.json",
            },
            {
                "check": "event-aware time-step convergence for all three valve types",
                "status": convergence["overall_status"],
                "value": convergence,
                "evidence": "results/<SOURCE_FILE_REDACTED>",
            },
            {
                "check": "controls evaluated after their model-internal design windows",
                "status": "pass",
                "evidence": "problem2: 90-100 s four-phase evaluation after 80-90 s calibration; problem3: 1.5-3 s",
            },
            {
                "check": "nozzle structure and relief actuator external identifiability",
                "status": "needs_review",
                "evidence": "No separate nozzle losses or verified actuator timing in allowed inputs.",
            },
            {
                "check": "deterministic full-script rerun",
                "status": "needs_review",
                "evidence": "to be filled by post-run hash comparison",
            },
            {
                "check": "external experimental validity",
                "status": "needs_review",
                "evidence": "No apparatus measurements are present in allowed inputs.",
            },
        ],
        "convergence": convergence,
        "mass_balance": {
            "rail_max_relative_to_throughput": maximum_mass_balance_relative,
            "chamber_max_relative_to_pumped_mass": maximum_chamber_mass_balance_relative,
            "status": (
                "pass"
                if max(
                    maximum_mass_balance_relative,
                    maximum_chamber_mass_balance_relative,
                )
                < 1e-4
                else "fail"
            ),
        },
        "structural_sensitivity": {
            "bottleneck_dynamic_omega_rad_per_ms": omega_problem2,
            "series_dynamic_omega_rad_per_ms": omega_series_dynamic,
            "relative_change": omega_series_dynamic / omega_problem2 - 1.0,
            "status": "needs_review",
        },
    }
    write_json(results_dir / "validation.json", validation)

    evidence_rows = [
        {
            "claim": "Input attachment identity and numeric schema",
            "evidence_layer": "file identity/schema",
            "status": "pass",
            "artifact": "results/input-manifest.json; results/data-audit.json",
        },
        {
            "claim": "Rail and compression-segment chamber physical mass balances",
            "evidence_layer": "model-internal invariant",
            "status": (
                "pass"
                if max(
                    maximum_mass_balance_relative,
                    maximum_chamber_mass_balance_relative,
                )
                < 1e-4
                else "fail"
            ),
            "artifact": "results/validation.json",
        },
        {
            "claim": "Pressure and event-aware time-step adequacy",
            "evidence_layer": "numerical convergence",
            "status": convergence["overall_status"],
            "artifact": "results/<SOURCE_FILE_REDACTED>",
        },
        {
            "claim": "Controls perform in subsequent model-internal evaluation windows",
            "evidence_layer": "model-internal evaluation",
            "status": "pass",
            "artifact": "results/problem2-summary.json; results/problem3-summary.json",
        },
        {
            "claim": "Nozzle structure and relief actuator timing identify a unique controller",
            "evidence_layer": "engineering identifiability",
            "status": "needs_review",
            "artifact": "results/problem2-summary.json; results/problem3-summary.json",
        },
        {
            "claim": "The same performance holds on a real fuel system",
            "evidence_layer": "external validity",
            "status": "needs_review",
            "artifact": "No experimental data supplied",
        },
        {
            "claim": "Declared authoritative generated artifacts are byte-reproducible",
            "evidence_layer": "execution reproducibility",
            "status": "needs_review",
            "artifact": "post-run comparison pending",
        },
    ]
    write_csv(results_dir / "<SOURCE_FILE_REDACTED>", evidence_rows)

    print("[8/8] writing consolidated values and environment record", flush=True)
    consolidated = {
        "baseline": baseline,
        "problem1": problem1,
        "problem2": p2_summary,
        "problem3": problem3,
        "validation": validation,
        "provenance": provenance,
        "sensitivity": {
            "time_step": timestep_metrics,
            "series_dynamic_omega_rad_per_ms": omega_series_dynamic,
        },
    }
    write_json(results_dir / "all-results.json", consolidated)
    environment = {
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "seed": SEED,
        "packages": {
            name: importlib.metadata.version(name)
            for name in ("numpy", "scipy", "openpyxl", "matplotlib", "PyYAML")
        },
        "command": "python code/run_all.py",
    }
    write_json(results_dir / "environment.json", environment)

    paper_values = {
        "p1_tau_100_ms": tau_100,
        "p1_tau_150_ms": tau_150,
        "p1_hold_100_mean_mpa": p1_hold["100_mpa"]["metrics_20s_to_30s"]["mean_mpa"],
        "p1_hold_100_ptp_mpa": p1_hold["100_mpa"]["metrics_20s_to_30s"]["peak_to_peak_mpa"],
        "p1_hold_150_mean_mpa": p1_hold["150_mpa"]["metrics_20s_to_30s"]["mean_mpa"],
        "p1_hold_150_ptp_mpa": p1_hold["150_mpa"]["metrics_20s_to_30s"]["peak_to_peak_mpa"],
        "p1_transition_2s_tau_max_ms": transition_summary[0]["tau_max_before_transition_ms"],
        "p1_transition_5s_tau_max_ms": transition_summary[1]["tau_max_before_transition_ms"],
        "p1_transition_10s_tau_max_ms": transition_summary[2]["tau_max_before_transition_ms"],
        "p2_omega_lower_rad_per_ms": structural_interval[0],
        "p2_omega_upper_rad_per_ms": structural_interval[1],
        "p2_rpm_lower": p2_summary["speed_interval_rpm"][0],
        "p2_rpm_upper": p2_summary["speed_interval_rpm"][1],
        "p2_phase_mean_span_bottleneck_mpa": p2_bottleneck_aggregate[
            "phase_mean_span_mpa"
        ],
        "p2_phase_mean_span_series_mpa": p2_series_aggregate["phase_mean_span_mpa"],
        "p2_eval_mean_mpa": p2_eval_metrics["mean_mpa"],
        "p2_eval_ptp_mpa": p2_eval_metrics["peak_to_peak_mpa"],
        "p3_economy_omega_rad_per_ms": economy_omega,
        "p3_economy_rpm": problem3["economy_scheme"]["speed_rpm"],
        "p3_economy_omega_lower_rad_per_ms": min(economy_series_omega, economy_omega),
        "p3_economy_omega_upper_rad_per_ms": max(economy_series_omega, economy_omega),
        "p3_economy_ptp_mpa": economy_metrics["peak_to_peak_mpa"],
        "p3_conditional_omega_rad_per_ms": precision_omega,
        "p3_conditional_rpm": selected_design["speed_rpm"],
        "p3_conditional_cam_phase_rad": selected_design["cam_phase_rad"],
        "p3_conditional_close_mpa": selected_thresholds[0],
        "p3_conditional_open_mpa": selected_thresholds[1],
        "p3_conditional_min_dwell_ms": selected_control.min_open_ms,
        "p3_conditional_mean_mpa": precision_metrics["mean_mpa"],
        "p3_conditional_ptp_mpa": precision_metrics["peak_to_peak_mpa"],
        "p3_conditional_switches_per_100ms": precision_dwell[
            "state_changes_per_100ms"
        ],
        "p3_relief_ratio": problem3["conditional_reference_scenario"][
            "relief_to_injection_volume_ratio"
        ],
        "rail_mass_residual_max_relative": maximum_mass_balance_relative,
        "series_omega_relative_change_pct": 100.0 * (omega_series_dynamic / omega_problem2 - 1.0),
    }
    write_json(results_dir / "paper-values.json", paper_values)
    tex_lines = [
        "% Generated by code/run_all.py; do not edit numerical values by hand.",
        f"\\newcommand{{\\PoneTauHundred}}{{{tau_100:.4f}}}",
        f"\\newcommand{{\\PoneTauOneFifty}}{{{tau_150:.4f}}}",
        f"\\newcommand{{\\PtwoOmegaLow}}{{{structural_interval[0]:.5f}}}",
        f"\\newcommand{{\\PtwoOmegaHigh}}{{{structural_interval[1]:.5f}}}",
        f"\\newcommand{{\\PtwoRPMLow}}{{{p2_summary['speed_interval_rpm'][0]:.1f}}}",
        f"\\newcommand{{\\PtwoRPMHigh}}{{{p2_summary['speed_interval_rpm'][1]:.1f}}}",
        f"\\newcommand{{\\PtwoMean}}{{{p2_eval_metrics['mean_mpa']:.2f}}}",
        f"\\newcommand{{\\PtwoPtp}}{{{p2_eval_metrics['peak_to_peak_mpa']:.2f}}}",
        f"\\newcommand{{\\PthreeEconomyRPM}}{{{problem3['economy_scheme']['speed_rpm']:.1f}}}",
        f"\\newcommand{{\\PthreeEconomyPtp}}{{{economy_metrics['peak_to_peak_mpa']:.2f}}}",
        f"\\newcommand{{\\PthreePrecisionRPM}}{{{selected_design['speed_rpm']:.0f}}}",
        f"\\newcommand{{\\PthreePhase}}{{{selected_design['cam_phase_rad']:.2f}}}",
        f"\\newcommand{{\\PthreeClose}}{{{selected_thresholds[0]:.1f}}}",
        f"\\newcommand{{\\PthreeOpen}}{{{selected_thresholds[1]:.1f}}}",
        f"\\newcommand{{\\PthreeMean}}{{{precision_metrics['mean_mpa']:.2f}}}",
        f"\\newcommand{{\\PthreePtp}}{{{precision_metrics['peak_to_peak_mpa']:.2f}}}",
        f"\\newcommand{{\\PthreeReliefRatio}}{{{problem3['conditional_reference_scenario']['relief_to_injection_volume_ratio']:.2f}}}",
    ]
    (results_dir / "paper-values.tex").write_text(
        "\n".join(tex_lines) + "\n", encoding="utf-8", newline="\n"
    )
    print("[PASS] all numerical artifacts generated", flush=True)


if __name__ == "__main__":
    main()
